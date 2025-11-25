#!/usr/bin/env python3
"""
Generate Excel Report with Multi-Level Headers (BU → SERVICE_GROUP → PRODUCT)
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.data_loader import CSVLoader, DataProcessor, DataAggregator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.row_order import ROW_ORDER
from config.data_mapping import get_group_sub_group, is_calculated_row
from config.settings import settings
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def generate_report_with_products(csv_path: Path, output_path: Path):
    """
    Generate P&L Excel report with multi-level headers

    Headers structure (4 rows):
    - Row 1: กลุ่มธุรกิจ (BU)
    - Row 2: กลุ่มบริการ (SERVICE_GROUP)
    - Row 3: รหัส (PRODUCT_KEY)
    - Row 4: ชื่อบริการ (PRODUCT_NAME)
    """
    logger.info(f"Loading CSV: {csv_path}")

    # Load CSV
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)

    logger.info(f"Loaded {len(df)} rows")

    # Process data
    data_processor = DataProcessor()
    df = data_processor.process_data(df)

    # Create aggregator
    aggregator = DataAggregator(df)

    # Get unique BUs
    bu_list = data_processor.get_unique_business_units(df)
    logger.info(f"BUs: {bu_list}")

    # Build service group dict
    service_group_dict = {}
    for bu in bu_list:
        service_group_dict[bu] = data_processor.get_unique_service_groups(df, bu)

    # Build product dict (BU → SERVICE_GROUP → [(PRODUCT_KEY, PRODUCT_NAME)])
    product_dict = {}
    for bu in bu_list:
        product_dict[bu] = {}
        for sg in service_group_dict[bu]:
            # Get unique products for this SERVICE_GROUP
            products = df[(df['BU'] == bu) & (df['SERVICE_GROUP'] == sg)][['PRODUCT_KEY', 'PRODUCT_NAME']].drop_duplicates()
            product_dict[bu][sg] = list(products.itertuples(index=False, name=None))

    logger.info(f"Products per BU/SG: {[(bu, sg, len(prods)) for bu in product_dict for sg in product_dict[bu] for prods in [product_dict[bu][sg]]][:10]}")

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    # Font settings
    font_name = settings.excel_font_name
    font_size = settings.excel_font_size

    # Start positions
    start_row = 5  # Row 6 (0-indexed = 5)
    start_col = 1  # Column B (0-indexed = 1)
    header_rows = 4  # 4 rows for multi-level headers

    # Header rows (rows 1-5)
    # Row 2: Company name
    cell = ws.cell(row=2, column=2)
    cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Row 3: Report name
    cell = ws.cell(row=3, column=2)
    period_desc, months = data_processor.get_period_description(df, "MTH" if "MTH" in csv_path.name else "YTD")
    report_type = "มิติประเภทต้นทุน" if "COSTTYPE" in csv_path.name else "มิติหมวดบัญชี"
    cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {report_type}"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Row 4: Period
    cell = ws.cell(row=4, column=2)
    cell.value = period_desc
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Info box (Column G onwards, row 2-5)
    info_box_content = """วัตถุประสงค์ของรายงาน :
เพื่อทราบผลดำเนินงานรายเดือนหลังหักค่าใช้จ่ายรวมทั้งหมด ประกอบด้วย ประเภทต้นทุน ขายและบริหาร
ใช้รายงานเพื่อการบริหารค่าใช้จ่ายแต่ละประเภทให้มีประสิทธิภาพ
ไม่เหมาะสมที่จะใช้เพื่อวัด Performance ส่วนงาน เนื่องจากเป็นการแสดงมิติค่าใช้จ่ายแบบรวม ไม่ได้แสดงเป็นค่าใช้จ่ายที่ส่วนงานมีอำนาจควบคุมได้หรือไม่ได้
ไม่เหมาะสมที่จะใช้เพื่อพิจารณาการยกเลิกบริการ เนื่องจากไม่ได้แสดงมิติต้นทุนผันแปรและคงที่ หากยกเลิกบริการแล้ว ควรทราบว่าต้นทุนใดบ้างลดลงได้ ต้นทุนใดบ้างยังคงอยู่ เพื่อเปรียบกับรายได้ที่จะหายไป"""

    info_cell = ws.cell(row=2, column=7)
    info_cell.value = info_box_content
    info_cell.font = Font(name=font_name, size=14)  # Size 14
    info_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    info_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # No wrap
    info_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=2, start_column=7, end_row=5, end_column=10)

    # Build column structure with multi-level headers
    current_col = start_col
    column_mapping = {}  # Map column index to data key

    # Column headers structure
    bu_headers = {}  # {col_start: (col_end, bu_name)}
    sg_headers = {}  # {col_start: (col_end, sg_name)}
    pk_headers = {}  # {col: product_key}
    pn_headers = {}  # {col: product_name}

    # Column B: รายละเอียด (spans all 4 header rows)
    ws.cell(row=start_row + 1, column=current_col + 1).value = "รายละเอียด"
    ws.column_dimensions[get_column_letter(current_col + 1)].width = 50
    ws.merge_cells(start_row=start_row + 1, start_column=current_col + 1,
                   end_row=start_row + header_rows, end_column=current_col + 1)
    current_col += 1

    # Column C: รวมทั้งสิ้น (GRAND TOTAL - ย้ายมาด้านซ้าย!)
    grand_total_col = current_col
    ws.cell(row=start_row + 1, column=grand_total_col + 1).value = "รวมทั้งสิ้น"
    ws.column_dimensions[get_column_letter(grand_total_col + 1)].width = 20
    ws.merge_cells(start_row=start_row + 1, start_column=grand_total_col + 1,
                   end_row=start_row + header_rows, end_column=grand_total_col + 1)
    column_mapping[grand_total_col] = ("GRAND_TOTAL", None, None, None, None)
    current_col += 1

    # For each BU
    for bu in bu_list:
        bu_start_col = current_col

        # BU total column
        bu_total_col = current_col
        ws.cell(row=start_row + 1, column=bu_total_col + 1).value = f"รวม {bu}"
        ws.column_dimensions[get_column_letter(bu_total_col + 1)].width = 18
        ws.merge_cells(start_row=start_row + 1, start_column=bu_total_col + 1,
                       end_row=start_row + header_rows, end_column=bu_total_col + 1)
        column_mapping[bu_total_col] = ("BU_TOTAL", bu, None, None, None)
        current_col += 1

        # Service groups
        for sg in service_group_dict.get(bu, []):
            sg_start_col = current_col

            # Products in this SERVICE_GROUP
            products = product_dict[bu][sg]

            if products:
                # Write products
                for product_key, product_name in products:
                    col = current_col

                    # Row 3: PRODUCT_KEY
                    ws.cell(row=start_row + 3, column=col + 1).value = product_key
                    # Row 4: PRODUCT_NAME
                    ws.cell(row=start_row + 4, column=col + 1).value = product_name

                    ws.column_dimensions[get_column_letter(col + 1)].width = 18
                    column_mapping[col] = ("PRODUCT", bu, sg, product_key, product_name)

                    pk_headers[col] = product_key
                    pn_headers[col] = product_name

                    current_col += 1

                sg_end_col = current_col - 1

                # Row 2: SERVICE_GROUP (merge across all products)
                if sg_start_col <= sg_end_col:
                    ws.cell(row=start_row + 2, column=sg_start_col + 1).value = sg
                    if sg_end_col > sg_start_col:
                        ws.merge_cells(start_row=start_row + 2, start_column=sg_start_col + 1,
                                       end_row=start_row + 2, end_column=sg_end_col + 1)
                    sg_headers[sg_start_col] = (sg_end_col, sg)

        bu_end_col = current_col - 1

        # Row 1: BU (merge across all service groups)
        if bu_start_col <= bu_end_col:
            ws.cell(row=start_row + 1, column=bu_start_col + 1).value = bu
            if bu_end_col > bu_start_col:
                ws.merge_cells(start_row=start_row + 1, start_column=bu_start_col + 1,
                               end_row=start_row + 1, end_column=bu_end_col + 1)
            bu_headers[bu_start_col] = (bu_end_col, bu)

    # Format all header cells
    for row in range(start_row + 1, start_row + header_rows + 1):
        for col in range(start_col + 1, current_col + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            cell.font = Font(name=font_name, size=font_size, bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # Write data rows
    current_row = start_row + header_rows
    all_row_data = {}
    # Track the current main group context
    current_main_group_label = None

    for level, label, is_calc, formula, is_bold in ROW_ORDER:
        if not label:
            current_row += 1
            continue

        # Update the main group context if this is a level 0 row
        if level == 0:
            current_main_group_label = label

        # Write label
        cell = ws.cell(row=current_row + 1, column=start_col + 1)
        cell.value = label
        cell.font = Font(name=font_name, size=font_size, bold=is_bold)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Color for section headers
        if is_bold and level == 0:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        # Get row data
        if is_calculated_row(label):
            row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
        else:
            row_data = aggregator.get_row_data(label, current_main_group_label, bu_list, service_group_dict)

        all_row_data[label] = row_data

        # Get GROUP and SUB_GROUP for this row
        group, sub_group = get_group_sub_group(label, current_main_group_label)

        # Write values
        for col_idx, (col_type, bu, sg, pk, pn) in column_mapping.items():
            if col_type == "GRAND_TOTAL":
                key = "GRAND_TOTAL"
                value = row_data.get(key, 0)
            elif col_type == "BU_TOTAL":
                key = f"BU_TOTAL_{bu}"
                value = row_data.get(key, 0)
            elif col_type == "PRODUCT":
                # For product level, use the new calculate_product_value method
                value = aggregator.calculate_product_value(label, bu, sg, pk, all_row_data, current_main_group_label)
                # Store product-level values for calculated rows
                product_key_str = f"{bu}_{sg}_{pk}"
                if product_key_str not in all_row_data.get(label, {}):
                    if label not in all_row_data:
                        all_row_data[label] = {}
                    all_row_data[label][product_key_str] = value
            else:
                continue
            cell = ws.cell(row=current_row + 1, column=col_idx + 1)

            if value is None:
                cell.value = ""
            elif isinstance(value, float) and "สัดส่วน" in label:
                cell.value = value
                cell.number_format = '0.00%'
            else:
                cell.value = value
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'

            cell.font = Font(name=font_name, size=font_size, bold=is_bold)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        current_row += 1

    # Freeze panes (at first data row, after "รวมทั้งสิ้น" column)
    ws.freeze_panes = f"{get_column_letter(grand_total_col + 2)}{start_row + header_rows + 1}"

    # Add remarks at bottom
    remark_file = csv_path.parent / f"remark_{csv_path.stem.split('_')[-1]}.txt"
    if remark_file.exists():
        with open(remark_file, 'r', encoding='utf-8') as f:
            remark_content = f.read()

        current_row += 2
        remark_cell = ws.cell(row=current_row + 1, column=start_col + 1)
        remark_cell.value = "หมายเหตุ"
        remark_cell.font = Font(name=font_name, size=14, bold=True)

        current_row += 1
        # Format remark content (no wrap text)
        remark_cell = ws.cell(row=current_row + 1, column=start_col + 1)
        remark_cell.value = remark_content
        remark_cell.font = Font(name=font_name, size=14)
        remark_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # No wrap

    # Save
    wb.save(output_path)
    logger.info(f"Report saved to: {output_path}")


if __name__ == "__main__":
    csv_path = Path("../data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    output_path = Path("./output/test_report_with_products.xlsx")

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)

    output_path.parent.mkdir(exist_ok=True)

    generate_report_with_products(csv_path, output_path)
    logger.info("Done!")
