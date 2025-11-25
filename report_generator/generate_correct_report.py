#!/usr/bin/env python3
"""
Generate Correct Excel Report with Multi-Level Headers
- BU Total
- Service Group Total (NEW!)
- Products
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.data_loader import CSVLoader, DataProcessor, DataAggregator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.row_order import ROW_ORDER
from config.data_mapping import get_group_sub_group, is_calculated_row
from config.settings import settings
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# BU Colors
BU_COLORS = {
    '1.กลุ่มธุรกิจ HARD INFRASTRUCTURE': 'E2EFDA',
    '2.กลุ่มธุรกิจ INTERNATIONAL': 'DDEBF7',
    '3.กลุ่มธุรกิจ MOBILE': 'DBD3E5',
    '4.กลุ่มธุรกิจ FIXED LINE & BROADBAND': 'FCE4D6',
    '5.กลุ่มธุรกิจ DIGITAL': 'D9E1F2',
    '6.กลุ่มธุรกิจ ICT SOLUTION': 'C6E0B4',
    '7.กลุ่มบริการอื่นไม่ใช่โทรคมนาคม': 'BDD7EE',
    '8.รายได้อื่น/ค่าใช้จ่ายอื่น': 'EAC1C0',
}


def generate_correct_report(csv_path: Path, output_path: Path):
    """Generate P&L Excel report with correct structure"""
    logger.info(f"Loading CSV: {csv_path}")

    # Load CSV
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)

    logger.info(f"Loaded {len(df)} rows")

    # Process data
    data_processor = DataProcessor()
    df = data_processor.process_data(df)

    # Create aggregator
    aggregator = DataAggregator(df)

    # Get unique BUs
    bu_list = data_processor.get_unique_business_units(df)
    logger.info(f"BUs: {bu_list}")

    # Build service group dict
    service_group_dict = {}
    for bu in bu_list:
        service_group_dict[bu] = data_processor.get_unique_service_groups(df, bu)

    # Build product dict
    product_dict = {}
    for bu in bu_list:
        product_dict[bu] = {}
        for sg in service_group_dict[bu]:
            products = df[(df['BU'] == bu) & (df['SERVICE_GROUP'] == sg)][['PRODUCT_KEY', 'PRODUCT_NAME']].sort_values(by='PRODUCT_KEY').drop_duplicates()
            product_dict[bu][sg] = list(products.itertuples(index=False, name=None))

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    font_name = settings.excel_font_name
    font_size = settings.excel_font_size

    start_row = 5  # Row 6 - data headers start here (info box ends at row 5)
    start_col = 1
    header_rows = 4

    # Header rows
    cell = ws.cell(row=2, column=2)
    cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=3, column=2)
    period_desc, months = data_processor.get_period_description(df, "MTH" if "MTH" in csv_path.name else "YTD")
    report_type = "มิติประเภทต้นทุน" if "COSTTYPE" in csv_path.name else "มิติหมวดบัญชี"
    cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {report_type}"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=4, column=2)
    cell.value = period_desc
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Info box - แยกบรรทัดจริง (แต่ละบรรทัดเป็น cell แยก)
    info_lines = [
        "วัตถุประสงค์ของรายงาน :",
        "เพื่อทราบผลดำเนินงานรายเดือนหลังหักค่าใช้จ่ายรวมทั้งหมด ประกอบด้วย ประเภทต้นทุน ขายและบริหาร",
        "ใช้รายงานเพื่อการบริหารค่าใช้จ่ายแต่ละประเภทให้มีประสิทธิภาพ",
        "ไม่เหมาะสมที่จะใช้เพื่อวัด Performance ส่วนงาน เนื่องจากเป็นการแสดงมิติค่าใช้จ่ายแบบรวม ไม่ได้แสดงเป็นค่าใช้จ่ายที่ส่วนงานมีอำนาจควบคุมได้หรือไม่ได้",
        "ไม่เหมาะสมที่จะใช้เพื่อพิจารณาการยกเลิกบริการ เนื่องจากไม่ได้แสดงมิติต้นทุนผันแปรและคงที่ หากยกเลิกบริการแล้ว ควรทราบว่าต้นทุนใดบ้างลดลงได้ ต้นทุนใดบ้างยังคงอยู่ เพื่อเปรียบกับรายได้ที่จะหายไป"
    ]

    # เขียนแต่ละบรรทัดใน cell แยก (rows 1-5)
    for idx, line in enumerate(info_lines):
        info_cell = ws.cell(row=1 + idx, column=7)
        info_cell.value = line
        info_cell.font = Font(name=font_name, size=14, bold=(idx == 0))  # Bold for title
        info_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        info_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # No wrap
        info_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Merge columns G-J for each line (rows 1-5, no overlap with data headers at row 6)
        ws.merge_cells(start_row=1 + idx, start_column=7, end_row=1 + idx, end_column=10)
        ws.row_dimensions[1 + idx].height = 20  # Adjust row height

    # Build column structure
    current_col = start_col
    column_mapping = {}

    # Column B: รายละเอียด
    detail_cell = ws.cell(row=start_row + 1, column=current_col + 1)
    detail_cell.value = "รายละเอียด"
    detail_cell.font = Font(name=font_name, size=font_size, bold=True)
    detail_cell.fill = PatternFill(start_color="F4DEDC", end_color="F4DEDC", fill_type="solid")
    detail_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=start_row + 1, start_column=current_col + 1,
                   end_row=start_row + header_rows, end_column=current_col + 1)
    ws.column_dimensions[get_column_letter(current_col + 1)].width = 50
    current_col += 1

    # Column C: รวมทั้งสิ้น
    grand_total_col = current_col
    gt_cell = ws.cell(row=start_row + 1, column=grand_total_col + 1)
    gt_cell.value = "รวมทั้งสิ้น"
    gt_cell.font = Font(name=font_name, size=font_size, bold=True)
    gt_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    gt_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gt_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=start_row + 1, start_column=grand_total_col + 1,
                   end_row=start_row + header_rows, end_column=grand_total_col + 1)
    ws.column_dimensions[get_column_letter(grand_total_col + 1)].width = 20
    column_mapping[grand_total_col] = ("GRAND_TOTAL", None, None, None, None)
    current_col += 1

    # For each BU
    for bu in bu_list:
        bu_start_col = current_col
        bu_color = BU_COLORS.get(bu, 'FFFFFF')

        # BU total column (รวม + BU name)
        bu_total_col = current_col
        bu_total_cell = ws.cell(row=start_row + 1, column=bu_total_col + 1)
        bu_total_cell.value = f"รวม {bu}"
        bu_total_cell.font = Font(name=font_name, size=font_size, bold=True)
        bu_total_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
        bu_total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bu_total_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        ws.merge_cells(start_row=start_row + 1, start_column=bu_total_col + 1,
                       end_row=start_row + header_rows, end_column=bu_total_col + 1)
        ws.column_dimensions[get_column_letter(bu_total_col + 1)].width = 18
        column_mapping[bu_total_col] = ("BU_TOTAL", bu, None, None, None)
        current_col += 1

        # Track first SG column for BU header merge
        bu_first_sg_col = current_col

        # Service groups
        for sg in service_group_dict.get(bu, []):
            sg_start_col = current_col

            # Service Group total column (รวม + SG name) - NEW!
            # Merge rows 7-9 (start_row + 2 to start_row + 4)
            sg_total_col = current_col
            sg_total_cell = ws.cell(row=start_row + 2, column=sg_total_col + 1)
            sg_total_cell.value = f"รวม {sg}"
            sg_total_cell.font = Font(name=font_name, size=font_size, bold=True)
            sg_total_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
            sg_total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sg_total_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # Merge rows 2-4 (7-9 in Excel)
            ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 1,
                           end_row=start_row + 4, end_column=sg_total_col + 1)
            ws.column_dimensions[get_column_letter(sg_total_col + 1)].width = 18
            column_mapping[sg_total_col] = ("SG_TOTAL", bu, sg, None, None)
            current_col += 1

            products = product_dict[bu][sg]

            if products:
                # Write products
                for product_key, product_name in products:
                    col = current_col

                    # Row 3: PRODUCT_KEY
                    pk_cell = ws.cell(row=start_row + 3, column=col + 1)
                    pk_cell.value = product_key
                    pk_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pk_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pk_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    # Row 4: PRODUCT_NAME
                    pn_cell = ws.cell(row=start_row + 4, column=col + 1)
                    pn_cell.value = product_name
                    pn_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pn_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pn_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pn_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    ws.column_dimensions[get_column_letter(col + 1)].width = 18
                    column_mapping[col] = ("PRODUCT", bu, sg, product_key, product_name)
                    current_col += 1

                sg_end_col = current_col - 1

                # Row 2: SERVICE_GROUP header (merge from first product to last product, row 7 only)
                if sg_end_col > sg_total_col:
                    sg_header_cell = ws.cell(row=start_row + 2, column=sg_total_col + 2)
                    sg_header_cell.value = sg
                    sg_header_cell.font = Font(name=font_name, size=font_size, bold=True)
                    sg_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    sg_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sg_header_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Merge only row 7 (start_row + 2)
                    ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 2,
                                   end_row=start_row + 2, end_column=sg_end_col + 1)

        bu_end_col = current_col - 1

        # Row 1: BU header (merge from first SG to last column of BU, row 6 only)
        if bu_end_col >= bu_first_sg_col:
            bu_header_cell = ws.cell(row=start_row + 1, column=bu_first_sg_col + 1)
            bu_header_cell.value = bu
            bu_header_cell.font = Font(name=font_name, size=font_size, bold=True)
            bu_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
            bu_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bu_header_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # Merge only row 6 (start_row + 1)
            ws.merge_cells(start_row=start_row + 1, start_column=bu_first_sg_col + 1,
                           end_row=start_row + 1, end_column=bu_end_col + 1)

    # Write data rows
    current_row = start_row + header_rows
    all_row_data = {}
    previous_label = None
    # New: Track the current main group context for sub-items
    current_main_group_label = None

    for level, label, is_calc, formula, is_bold in ROW_ORDER:
        # New: Update the main group context if this is a level 0 row
        if level == 0:
            current_main_group_label = label

        if not label:
            current_row += 1
            continue

        # Write label
        cell = ws.cell(row=current_row + 1, column=start_col + 1)
        cell.value = label
        cell.font = Font(name=font_name, size=font_size, bold=is_bold)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Get row data - handle ratio rows with context
        is_ratio_row = (label == "         สัดส่วนต่อรายได้")
        ratio_calc_type = None

        if is_ratio_row:
            # Determine which ratio to calculate based on previous row
            if previous_label == "     1. ต้นทุนบริการรวม":
                ratio_calc_type = "total_service_cost_ratio"
            elif previous_label == "     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ":
                ratio_calc_type = "service_cost_no_depreciation_ratio"
            elif previous_label == "     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ":
                ratio_calc_type = "service_cost_no_personnel_depreciation_ratio"

            if ratio_calc_type:
                row_data = aggregator._calculate_ratio_by_type(ratio_calc_type, all_row_data, bu_list, service_group_dict)
            else:
                row_data = {}
        elif is_calculated_row(label):
            row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
        else:
            # Modified: Pass the main group context to the aggregator
            row_data = aggregator.get_row_data(label, current_main_group_label, bu_list, service_group_dict)

        all_row_data[label] = row_data
        # Modified: Pass context to get_group_sub_group
        group, sub_group = get_group_sub_group(label, current_main_group_label)
        previous_label = label

        # Determine row color
        section_header_color = None
        if is_bold and level == 0:
            section_header_color = "F8CBAD"  # Section headers (1.รายได้, 2.ต้นทุน, etc.)
            cell.fill = PatternFill(start_color=section_header_color, end_color=section_header_color, fill_type="solid")

        # Write values
        for col_idx, (col_type, bu, sg, pk, pn) in column_mapping.items():
            if col_type == "GRAND_TOTAL":
                value = row_data.get("GRAND_TOTAL", 0)
            elif col_type == "BU_TOTAL":
                value = row_data.get(f"BU_TOTAL_{bu}", 0)
            elif col_type == "SG_TOTAL":
                # Service Group total = sum of all products in this SG
                value = row_data.get(f"{bu}_{sg}", 0)
            elif col_type == "PRODUCT":
                # For product level, use the new calculate_product_value method
                # Special handling for ratio rows
                if is_ratio_row and ratio_calc_type:
                    # Calculate ratio for this specific product
                    product_key_str = f"{bu}_{sg}_{pk}"
                    if ratio_calc_type == "total_service_cost_ratio":
                        service_revenue = all_row_data.get("รายได้บริการ", {}).get(product_key_str, 0)
                        total_cost = all_row_data.get("     1. ต้นทุนบริการรวม", {}).get(product_key_str, 0)
                        value = total_cost / service_revenue if abs(service_revenue) >= 1e-9 else None
                    elif ratio_calc_type == "service_cost_no_depreciation_ratio":
                        service_revenue = all_row_data.get("รายได้บริการ", {}).get(product_key_str, 0)
                        cost_no_dep = all_row_data.get("     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ", {}).get(product_key_str, 0)
                        value = cost_no_dep / service_revenue if abs(service_revenue) >= 1e-9 else None
                    elif ratio_calc_type == "service_cost_no_personnel_depreciation_ratio":
                        service_revenue = all_row_data.get("รายได้บริการ", {}).get(product_key_str, 0)
                        cost_no_pers_dep = all_row_data.get("     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ", {}).get(product_key_str, 0)
                        total_cost_check = all_row_data.get("     1. ต้นทุนบริการรวม", {}).get(product_key_str, 0)
                        value = cost_no_pers_dep / service_revenue if abs(service_revenue) >= 1e-9 else None
                        # Debug logging
                        print(f"DEBUG RETRIEVE [{pk}]: key='{product_key_str}'")
                        print(f"  revenue={service_revenue}, cost_row3={cost_no_pers_dep}, cost_row1={total_cost_check}, ratio={value}")
                    else:
                        value = None
                else:
                    value = aggregator.calculate_product_value(label, bu, sg, pk, all_row_data, current_main_group_label)

                # Store product-level values for calculated rows
                product_key_str = f"{bu}_{sg}_{pk}"
                if label not in all_row_data:
                    all_row_data[label] = {}
                all_row_data[label][product_key_str] = value

                # Debug logging - check what we stored
                if "ต้นทุนบริการ" in label or label == "รายได้บริการ":
                    print(f"DEBUG STORE [{pk}]: label='{label}', key='{product_key_str}', value={value}")
            else:
                continue

            cell = ws.cell(row=current_row + 1, column=col_idx + 1)

            if value is None:
                cell.value = ""
                # Apply dark gray background for None values (e.g., row 14 for non-grand-total columns)
                if col_type != "GRAND_TOTAL" and label == "14.กำไร(ขาดทุน) สุทธิ (12) - (13)":
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            elif isinstance(value, float) and "สัดส่วน" in label:
                cell.value = value
                cell.number_format = '0.00%'
            else:
                cell.value = value
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'

            cell.font = Font(name=font_name, size=font_size, bold=is_bold)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Apply section header color to all data cells in the row (only if not already gray)
            if section_header_color and not (value is None and col_type != "GRAND_TOTAL" and label == "14.กำไร(ขาดทุน) สุทธิ (12) - (13)"):
                cell.fill = PatternFill(start_color=section_header_color, end_color=section_header_color, fill_type="solid")

        current_row += 1

    # Freeze panes
    ws.freeze_panes = f"{get_column_letter(grand_total_col + 2)}{start_row + header_rows + 1}"

    # Add remarks - แยกบรรทัด (แต่ละบรรทัดเป็น row ใหม่)
    remark_file = csv_path.parent / f"remark_{csv_path.stem.split('_')[-1]}.txt"
    if remark_file.exists():
        try:
            for enc in ['utf-8', 'tis-620', 'cp874']:
                try:
                    with open(remark_file, 'r', encoding=enc) as f:
                        remark_content = f.read()
                    break
                except:
                    continue

            current_row += 2

            # Title
            remark_title_cell = ws.cell(row=current_row + 1, column=start_col + 1)
            remark_title_cell.value = "หมายเหตุ:"
            remark_title_cell.font = Font(name=font_name, size=14, bold=True)
            remark_title_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
            current_row += 1

            # Content - แยกแต่ละบรรทัดเป็น row ใหม่
            remark_lines = remark_content.strip().split('\n')
            for line in remark_lines:
                if line.strip():  # Skip empty lines
                    remark_cell = ws.cell(row=current_row + 1, column=start_col + 1)
                    remark_cell.value = line.strip()
                    remark_cell.font = Font(name=font_name, size=14)
                    remark_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)  # No wrap
                    current_row += 1

        except Exception as e:
            logger.warning(f"Could not read remark file: {e}")

    # Save
    wb.save(output_path)
    logger.info(f"Report saved to: {output_path}")


if __name__ == "__main__":
    csv_path = Path("../data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    output_path = Path("./output/correct_report.xlsx")

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)

    output_path.parent.mkdir(exist_ok=True)

    generate_correct_report(csv_path, output_path)
    logger.info("Done!")
