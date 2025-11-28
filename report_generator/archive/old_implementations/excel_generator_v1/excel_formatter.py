"""
Excel Formatter - Handle all Excel formatting (colors, fonts, borders, number formats)
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
import logging

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Handle Excel cell formatting"""

    def __init__(
        self,
        font_name: str = "TH Sarabun New",
        font_size: int = 18,
        header_font_size: int = 18,
        remark_font_size: int = 14
    ):
        """
        Initialize Excel Formatter

        Args:
            font_name: Font name to use
            font_size: Default font size
            header_font_size: Font size for headers
            remark_font_size: Font size for remarks
        """
        self.font_name = font_name
        self.font_size = font_size
        self.header_font_size = header_font_size
        self.remark_font_size = remark_font_size

    def create_font(
        self,
        bold: bool = False,
        size: Optional[int] = None,
        color: str = "000000"
    ) -> Font:
        """
        Create font object

        Args:
            bold: Whether font is bold
            size: Font size (None = use default)
            color: Font color in hex (without #)

        Returns:
            Font object
        """
        return Font(
            name=self.font_name,
            size=size or self.font_size,
            bold=bold,
            color=color
        )

    def create_fill(self, color: str) -> PatternFill:
        """
        Create fill object

        Args:
            color: Fill color in hex (with or without #)

        Returns:
            PatternFill object
        """
        # Remove # if present
        color = color.lstrip('#')
        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    def create_border(
        self,
        style: str = 'thin',
        all_sides: bool = True
    ) -> Border:
        """
        Create border object

        Args:
            style: Border style (thin, medium, thick)
            all_sides: Apply to all sides

        Returns:
            Border object
        """
        side = Side(style=style)
        if all_sides:
            return Border(left=side, right=side, top=side, bottom=side)
        else:
            return Border()

    def create_alignment(
        self,
        horizontal: str = 'left',
        vertical: str = 'center',
        wrap_text: bool = False
    ) -> Alignment:
        """
        Create alignment object

        Args:
            horizontal: Horizontal alignment (left, center, right)
            vertical: Vertical alignment (top, center, bottom)
            wrap_text: Whether to wrap text

        Returns:
            Alignment object
        """
        return Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=wrap_text
        )

    def apply_header_style(
        self,
        cell,
        bg_color: str = "#F8CBAD",
        bold: bool = True,
        font_color: str = "000000"
    ):
        """
        Apply header style to a cell

        Args:
            cell: Cell object to style
            bg_color: Background color
            bold: Whether text is bold
            font_color: Font color
        """
        cell.font = self.create_font(bold=bold, size=self.header_font_size, color=font_color)
        cell.fill = self.create_fill(bg_color)
        cell.border = self.create_border()
        cell.alignment = self.create_alignment(horizontal='left', vertical='center')

    def apply_data_cell_style(
        self,
        cell,
        bg_color: Optional[str] = None,
        bold: bool = False,
        is_number: bool = False
    ):
        """
        Apply data cell style

        Args:
            cell: Cell object to style
            bg_color: Background color (optional)
            bold: Whether text is bold
            is_number: Whether cell contains number
        """
        cell.font = self.create_font(bold=bold)
        if bg_color:
            cell.fill = self.create_fill(bg_color)
        cell.border = self.create_border()

        if is_number:
            # Number format: positive as 1,234.00, negative as (1,234.00) in red, zero as empty
            cell.number_format = '#,##0.00;[Red](#,##0.00);""'
            cell.alignment = self.create_alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = self.create_alignment(horizontal='left', vertical='center')

    def apply_column_style(
        self,
        cell,
        bu_name: str,
        bu_colors: dict,
        is_header: bool = False,
        bold: bool = False
    ):
        """
        Apply column style based on business unit

        Args:
            cell: Cell object to style
            bu_name: Business unit name
            bu_colors: Dictionary mapping BU names to colors
            is_header: Whether this is a header cell
            bold: Whether text is bold
        """
        bg_color = bu_colors.get(bu_name, "#FFFFFF")

        if is_header:
            self.apply_header_style(cell, bg_color=bg_color, bold=True)
        else:
            self.apply_data_cell_style(cell, bg_color=bg_color, bold=bold, is_number=True)

    def apply_info_box_style(self, ws, start_row: int, start_col: int, content: str, bg_color: str = "#F8CBAD"):
        """
        Apply style to info box

        Args:
            ws: Worksheet object
            start_row: Starting row (0-indexed)
            start_col: Starting column (0-indexed)
            content: Content to display
            bg_color: Background color
        """
        # Split content into lines
        lines = content.strip().split('\n')

        for i, line in enumerate(lines):
            cell = ws.cell(row=start_row + i + 1, column=start_col + 1)
            cell.value = line
            cell.font = self.create_font(size=self.remark_font_size)
            cell.fill = self.create_fill(bg_color)
            cell.border = self.create_border()
            cell.alignment = self.create_alignment(horizontal='left', vertical='top', wrap_text=True)

    def apply_remark_style(self, ws, start_row: int, content: str):
        """
        Apply style to remarks section

        Args:
            ws: Worksheet object
            start_row: Starting row (0-indexed)
            content: Remark content
        """
        # Column B (index 1 in 0-indexed)
        col_idx = 1

        # Header "หมายเหตุ"
        cell = ws.cell(row=start_row + 1, column=col_idx + 1)
        cell.value = "หมายเหตุ"
        cell.font = self.create_font(bold=True, size=self.remark_font_size)
        cell.alignment = self.create_alignment(horizontal='left', vertical='top')

        # Parse and format remark content
        lines = content.strip().split('\n')
        current_row = start_row + 2

        for line in lines:
            line = line.strip()
            if not line:
                continue

            cell = ws.cell(row=current_row, column=col_idx + 1)
            cell.value = line
            cell.font = self.create_font(size=self.remark_font_size)
            cell.alignment = self.create_alignment(horizontal='left', vertical='top', wrap_text=True)

            # Increase indent for sub-items (e.g., 6.1, 6.2)
            if line and (line[0].isdigit() and '.' in line[:5]):
                # Check if it's a sub-item (e.g., "6.1")
                parts = line.split('.', 2)
                if len(parts) >= 2 and len(parts[1]) > 0 and parts[1][0].isdigit():
                    cell.alignment = self.create_alignment(horizontal='left', vertical='top', wrap_text=True)

            current_row += 1

    def set_column_width(self, ws, column_index: int, width: float):
        """
        Set column width

        Args:
            ws: Worksheet object
            column_index: Column index (0-indexed)
            width: Width in characters
        """
        col_letter = get_column_letter(column_index + 1)
        ws.column_dimensions[col_letter].width = width

    def set_row_height(self, ws, row_index: int, height: float):
        """
        Set row height

        Args:
            ws: Worksheet object
            row_index: Row index (0-indexed)
            height: Height in points
        """
        ws.row_dimensions[row_index + 1].height = height

    def apply_freeze_panes(self, ws, row: int, col: int):
        """
        Apply freeze panes

        Args:
            ws: Worksheet object
            row: Row to freeze at (0-indexed)
            col: Column to freeze at (0-indexed)
        """
        # Convert to Excel cell reference (1-indexed)
        cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
        ws.freeze_panes = cell_ref
        logger.info(f"Freeze panes set at: {cell_ref}")
