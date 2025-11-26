#!/usr/bin/env python3
"""
Generate Excel Report with Debugging and Validation
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.data_loader import CSVLoader, DataProcessor, DataAggregator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.row_order import ROW_ORDER
from config.data_mapping import get_group_sub_group, is_calculated_row
from config.settings import settings
import logging

logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# BU Colors
BU_COLORS = {
    '1.กลุ่มธุรกิจ HARD INFRASTRUCTURE': 'E2EFDA',
    '2.กลุ่มธุรกิจ INTERNATIONAL': 'DDEBF7',
    '3.กลุ่มธุรกิจ MOBILE': 'DBD3E5',
    '4.กลุ่มธุรกิจ FIXED LINE & BROADBAND': 'FCE4D6',
    '5.กลุ่มธุรกิจ DIGITAL': 'D9E1F2',
    '6.กลุ่มธุรกิจ ICT SOLUTION': 'C6E0B4',
    '7.กลุ่มบริการอื่นไม่ใช่โทรคมนาคม': 'BDD7EE',
    '8.รายได้อื่น/ค่าใช้จ่ายอื่น': 'EAC1C0',
}


class MergeTracker:
    """Track all merged cell ranges to detect overlaps"""
    def __init__(self):
        self.merges = []

    def add_merge(self, start_row, start_col, end_row, end_col, description=""):
        """Add a merge and check for overlaps"""
        new_merge = (start_row, start_col, end_row, end_col)

        # Check for overlaps
        for existing_merge in self.merges:
            if self._ranges_overlap(new_merge, existing_merge):
                logger.error(f"OVERLAP DETECTED!")
                logger.error(f"  New merge: {description}")
                logger.error(f"    Rows {start_row}-{end_row}, Cols {start_col}-{end_col}")
                logger.error(f"  Existing merge: Rows {existing_merge[0]}-{existing_merge[2]}, Cols {existing_merge[1]}-{existing_merge[3]}")
                return False

        self.merges.append(new_merge)
        logger.debug(f"Merge OK: {description} | R{start_row}:R{end_row} x C{start_col}:C{end_col}")
        return True

    def _ranges_overlap(self, range1, range2):
        """Check if two ranges overlap"""
        r1_start, c1_start, r1_end, c1_end = range1
        r2_start, c2_start, r2_end, c2_end = range2

        # Check if ranges overlap
        rows_overlap = not (r1_end < r2_start or r2_end < r1_start)
        cols_overlap = not (c1_end < c2_start or c2_end < c1_start)

        return rows_overlap and cols_overlap


def generate_debug_report(csv_path: Path, output_path: Path):
    """Generate P&L Excel report with debugging"""
    logger.info(f"Loading CSV: {csv_path}")

    # Load CSV
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)
    logger.info(f"Loaded {len(df)} rows")

    # Process data
    data_processor = DataProcessor()
    df = data_processor.process_data(df)

    # Create aggregator
    aggregator = DataAggregator(df)

    # Get unique BUs
    bu_list = data_processor.get_unique_business_units(df)
    logger.info(f"BUs: {bu_list}")

    # Build service group dict
    service_group_dict = {}
    for bu in bu_list:
        service_group_dict[bu] = data_processor.get_unique_service_groups(df, bu)

    # Build product dict
    product_dict = {}
    for bu in bu_list:
        product_dict[bu] = {}
        for sg in service_group_dict[bu]:
            products = df[(df['BU'] == bu) & (df['SERVICE_GROUP'] == sg)][['PRODUCT_KEY', 'PRODUCT_NAME']].drop_duplicates()
            product_dict[bu][sg] = list(products.itertuples(index=False, name=None))

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    font_name = settings.excel_font_name
    font_size = settings.excel_font_size

    start_row = 5  # Row 6 - data headers start here (info box ends at row 5)
    start_col = 1
    header_rows = 4

    # Merge tracker
    merge_tracker = MergeTracker()

    # Header rows
    cell = ws.cell(row=2, column=2)
    cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=3, column=2)
    period_desc, months = data_processor.get_period_description(df, "MTH" if "MTH" in csv_path.name else "YTD")
    report_type = "มิติประเภทต้นทุน" if "COSTTYPE" in csv_path.name else "มิติหมวดบัญชี"
    cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {report_type}"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=4, column=2)
    cell.value = period_desc
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Info box - separate rows
    info_lines = [
        "วัตถุประสงค์ของรายงาน :",
        "เพื่อทราบผลดำเนินงานรายเดือนหลังหักค่าใช้จ่ายรวมทั้งหมด ประกอบด้วย ประเภทต้นทุน ขายและบริหาร",
        "ใช้รายงานเพื่อการบริหารค่าใช้จ่ายแต่ละประเภทให้มีประสิทธิภาพ",
        "ไม่เหมาะสมที่จะใช้เพื่อวัด Performance ส่วนงาน เนื่องจากเป็นการแสดงมิติค่าใช้จ่ายแบบรวม ไม่ได้แสดงเป็นค่าใช้จ่ายที่ส่วนงานมีอำนาจควบคุมได้หรือไม่ได้",
        "ไม่เหมาะสมที่จะใช้เพื่อพิจารณาการยกเลิกบริการ เนื่องจากไม่ได้แสดงมิติต้นทุนผันแปรและคงที่ หากยกเลิกบริการแล้ว ควรทราบว่าต้นทุนใดบ้างลดลงได้ ต้นทุนใดบ้างยังคงอยู่ เพื่อเปรียบกับรายได้ที่จะหายไป"
    ]

    logger.info("Creating info box...")
    for idx, line in enumerate(info_lines):
        info_cell = ws.cell(row=1 + idx, column=7)
        info_cell.value = line
        info_cell.font = Font(name=font_name, size=14, bold=(idx == 0))
        info_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        info_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        info_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Merge columns G-J for each line (rows 1-5, no overlap with data headers at row 6)
        if merge_tracker.add_merge(1 + idx, 7, 1 + idx, 10, f"Info line {idx+1}"):
            ws.merge_cells(start_row=1 + idx, start_column=7, end_row=1 + idx, end_column=10)
        ws.row_dimensions[1 + idx].height = 20

    # Build column structure
    current_col = start_col
    column_mapping = {}

    logger.info("Creating column headers...")

    # Column B: รายละเอียด
    detail_cell = ws.cell(row=start_row + 1, column=current_col + 1)
    detail_cell.value = "รายละเอียด"
    detail_cell.font = Font(name=font_name, size=font_size, bold=True)
    detail_cell.fill = PatternFill(start_color="F4DEDC", end_color="F4DEDC", fill_type="solid")
    detail_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    if merge_tracker.add_merge(start_row + 1, current_col + 1, start_row + header_rows, current_col + 1, "รายละเอียด"):
        ws.merge_cells(start_row=start_row + 1, start_column=current_col + 1,
                       end_row=start_row + header_rows, end_column=current_col + 1)
    ws.column_dimensions[get_column_letter(current_col + 1)].width = 50
    current_col += 1

    # Column C: รวมทั้งสิ้น
    grand_total_col = current_col
    gt_cell = ws.cell(row=start_row + 1, column=grand_total_col + 1)
    gt_cell.value = "รวมทั้งสิ้น"
    gt_cell.font = Font(name=font_name, size=font_size, bold=True)
    gt_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    gt_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gt_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    if merge_tracker.add_merge(start_row + 1, grand_total_col + 1, start_row + header_rows, grand_total_col + 1, "รวมทั้งสิ้น"):
        ws.merge_cells(start_row=start_row + 1, start_column=grand_total_col + 1,
                       end_row=start_row + header_rows, end_column=grand_total_col + 1)
    ws.column_dimensions[get_column_letter(grand_total_col + 1)].width = 20
    column_mapping[grand_total_col] = ("GRAND_TOTAL", None, None, None, None)
    current_col += 1

    # For each BU
    for bu_idx, bu in enumerate(bu_list):
        logger.info(f"Processing BU {bu_idx+1}/{len(bu_list)}: {bu}")
        bu_start_col = current_col
        bu_color = BU_COLORS.get(bu, 'FFFFFF')

        # BU total column
        bu_total_col = current_col
        bu_total_cell = ws.cell(row=start_row + 1, column=bu_total_col + 1)
        bu_total_cell.value = f"รวม {bu}"
        bu_total_cell.font = Font(name=font_name, size=font_size, bold=True)
        bu_total_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
        bu_total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bu_total_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        if merge_tracker.add_merge(start_row + 1, bu_total_col + 1, start_row + header_rows, bu_total_col + 1, f"BU Total: {bu}"):
            ws.merge_cells(start_row=start_row + 1, start_column=bu_total_col + 1,
                           end_row=start_row + header_rows, end_column=bu_total_col + 1)
        ws.column_dimensions[get_column_letter(bu_total_col + 1)].width = 18
        column_mapping[bu_total_col] = ("BU_TOTAL", bu, None, None, None)
        current_col += 1

        # Track first SG column for BU header merge
        bu_first_sg_col = current_col

        # Service groups
        for sg_idx, sg in enumerate(service_group_dict.get(bu, [])):
            logger.debug(f"  SG {sg_idx+1}: {sg}")
            sg_start_col = current_col

            # Service Group total column
            sg_total_col = current_col
            sg_total_cell = ws.cell(row=start_row + 2, column=sg_total_col + 1)
            sg_total_cell.value = f"รวม {sg}"
            sg_total_cell.font = Font(name=font_name, size=font_size, bold=True)
            sg_total_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
            sg_total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sg_total_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # Merge rows 7-9 (start_row + 2 to start_row + 4)
            if merge_tracker.add_merge(start_row + 2, sg_total_col + 1, start_row + 4, sg_total_col + 1, f"SG Total: {sg}"):
                ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 1,
                               end_row=start_row + 4, end_column=sg_total_col + 1)
            ws.column_dimensions[get_column_letter(sg_total_col + 1)].width = 18
            column_mapping[sg_total_col] = ("SG_TOTAL", bu, sg, None, None)
            current_col += 1

            products = product_dict[bu][sg]
            logger.debug(f"    Products: {len(products)}")

            if products:
                # Write products
                for prod_idx, (product_key, product_name) in enumerate(products):
                    col = current_col

                    # Row 8: PRODUCT_KEY
                    pk_cell = ws.cell(row=start_row + 3, column=col + 1)
                    pk_cell.value = product_key
                    pk_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pk_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pk_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    # Row 9: PRODUCT_NAME
                    pn_cell = ws.cell(row=start_row + 4, column=col + 1)
                    pn_cell.value = product_name
                    pn_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pn_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pn_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pn_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    ws.column_dimensions[get_column_letter(col + 1)].width = 18
                    column_mapping[col] = ("PRODUCT", bu, sg, product_key, product_name)
                    current_col += 1

                sg_end_col = current_col - 1

                # Row 7: SERVICE_GROUP header (merge across products only, not including SG total)
                if sg_end_col > sg_total_col:
                    sg_header_cell = ws.cell(row=start_row + 2, column=sg_total_col + 2)
                    sg_header_cell.value = sg
                    sg_header_cell.font = Font(name=font_name, size=font_size, bold=True)
                    sg_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    sg_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sg_header_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Merge only row 7
                    if merge_tracker.add_merge(start_row + 2, sg_total_col + 2, start_row + 2, sg_end_col + 1, f"SG Header: {sg}"):
                        ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 2,
                                       end_row=start_row + 2, end_column=sg_end_col + 1)

        bu_end_col = current_col - 1

        # Row 6: BU header (merge from first SG to last column of BU)
        if bu_end_col >= bu_first_sg_col:
            bu_header_cell = ws.cell(row=start_row + 1, column=bu_first_sg_col + 1)
            bu_header_cell.value = bu
            bu_header_cell.font = Font(name=font_name, size=font_size, bold=True)
            bu_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
            bu_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bu_header_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # Merge only row 6
            if merge_tracker.add_merge(start_row + 1, bu_first_sg_col + 1, start_row + 1, bu_end_col + 1, f"BU Header: {bu}"):
                ws.merge_cells(start_row=start_row + 1, start_column=bu_first_sg_col + 1,
                               end_row=start_row + 1, end_column=bu_end_col + 1)

    logger.info("Writing data rows...")
    # Write data rows
    current_row = start_row + header_rows
    all_row_data = {}

    for level, label, is_calc, formula, is_bold in ROW_ORDER:
        if not label:
            current_row += 1
            continue

        # Write label
        cell = ws.cell(row=current_row + 1, column=start_col + 1)
        cell.value = label
        cell.font = Font(name=font_name, size=font_size, bold=is_bold)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        if is_bold and level == 0:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        # Get row data
        if is_calculated_row(label):
            row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
        else:
            row_data = aggregator.get_row_data(label, bu_list, service_group_dict)

        all_row_data[label] = row_data
        group, sub_group = get_group_sub_group(label)

        # Write values
        for col_idx, (col_type, bu, sg, pk, pn) in column_mapping.items():
            if col_type == "GRAND_TOTAL":
                value = row_data.get("GRAND_TOTAL", 0)
            elif col_type == "BU_TOTAL":
                value = row_data.get(f"BU_TOTAL_{bu}", 0)
            elif col_type == "SG_TOTAL":
                value = row_data.get(f"{bu}_{sg}", 0)
            elif col_type == "PRODUCT":
                if group and not is_calculated_row(label):
                    value = aggregator.get_value_by_product(group, sub_group, bu, sg, pk)
                else:
                    value = 0
            else:
                continue

            cell = ws.cell(row=current_row + 1, column=col_idx + 1)

            if value is None:
                cell.value = ""
            elif isinstance(value, float) and "สัดส่วน" in label:
                cell.value = value
                cell.number_format = '0.00%'
            else:
                cell.value = value
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'

            cell.font = Font(name=font_name, size=font_size, bold=is_bold)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        current_row += 1

    # Freeze panes
    ws.freeze_panes = f"{get_column_letter(grand_total_col + 2)}{start_row + header_rows + 1}"

    # Add remarks
    logger.info("Adding remarks...")
    remark_file = csv_path.parent / f"remark_{csv_path.stem.split('_')[-1]}.txt"
    if remark_file.exists():
        try:
            for enc in ['utf-8', 'tis-620', 'cp874']:
                try:
                    with open(remark_file, 'r', encoding=enc) as f:
                        remark_content = f.read()
                    break
                except:
                    continue

            current_row += 2

            # Title
            remark_title_cell = ws.cell(row=current_row + 1, column=start_col + 1)
            remark_title_cell.value = "หมายเหตุ:"
            remark_title_cell.font = Font(name=font_name, size=14, bold=True)
            remark_title_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
            current_row += 1

            # Content - separate rows
            remark_lines = remark_content.strip().split('\n')
            for line in remark_lines:
                if line.strip():
                    remark_cell = ws.cell(row=current_row + 1, column=start_col + 1)
                    remark_cell.value = line.strip()
                    remark_cell.font = Font(name=font_name, size=14)
                    remark_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
                    current_row += 1

        except Exception as e:
            logger.warning(f"Could not read remark file: {e}")

    # Save
    logger.info(f"Saving to: {output_path}")
    wb.save(output_path)
    logger.info(f"Report saved successfully!")
    logger.info(f"Total merges: {len(merge_tracker.merges)}")


if __name__ == "__main__":
    csv_path = Path("../data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    output_path = Path("./output/debug_report.xlsx")

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)

    output_path.parent.mkdir(exist_ok=True)

    try:
        generate_debug_report(csv_path, output_path)
        logger.info("✅ Done!")
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
