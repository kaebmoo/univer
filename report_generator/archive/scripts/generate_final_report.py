#!/usr/bin/env python3
"""
Generate Final Excel Report with Multi-Level Headers and Correct Colors
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.data_loader import CSVLoader, DataProcessor, DataAggregator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.row_order import ROW_ORDER
from config.data_mapping import get_group_sub_group, is_calculated_row
from config.settings import settings
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# BU Colors
BU_COLORS = {
    '1.กลุ่มธุรกิจ HARD INFRASTRUCTURE': 'E2EFDA',
    '2.กลุ่มธุรกิจ INTERNATIONAL': 'DDEBF7',
    '3.กลุ่มธุรกิจ MOBILE': 'DBD3E5',
    '4.กลุ่มธุรกิจ FIXED LINE & BROADBAND': 'FCE4D6',
    '5.กลุ่มธุรกิจ DIGITAL': 'D9E1F2',
    '6.กลุ่มธุรกิจ ICT SOLUTION': 'C6E0B4',
    '7.กลุ่มบริการอื่นไม่ใช่โทรคมนาคม': 'BDD7EE',
    '8.รายได้อื่น/ค่าใช้จ่ายอื่น': 'EAC1C0',
}


def generate_final_report(csv_path: Path, output_path: Path):
    """
    Generate P&L Excel report with multi-level headers and correct colors
    """
    logger.info(f"Loading CSV: {csv_path}")

    # Load CSV
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)

    logger.info(f"Loaded {len(df)} rows")

    # Process data
    data_processor = DataProcessor()
    df = data_processor.process_data(df)

    # Create aggregator
    aggregator = DataAggregator(df)

    # Get unique BUs
    bu_list = data_processor.get_unique_business_units(df)
    logger.info(f"BUs: {bu_list}")

    # Build service group dict
    service_group_dict = {}
    for bu in bu_list:
        service_group_dict[bu] = data_processor.get_unique_service_groups(df, bu)

    # Build product dict
    product_dict = {}
    for bu in bu_list:
        product_dict[bu] = {}
        for sg in service_group_dict[bu]:
            products = df[(df['BU'] == bu) & (df['SERVICE_GROUP'] == sg)][['PRODUCT_KEY', 'PRODUCT_NAME']].drop_duplicates()
            product_dict[bu][sg] = list(products.itertuples(index=False, name=None))

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    # Font settings
    font_name = settings.excel_font_name
    font_size = settings.excel_font_size

    # Start positions
    start_row = 5
    start_col = 1
    header_rows = 4

    # Header rows (rows 1-5)
    cell = ws.cell(row=2, column=2)
    cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=3, column=2)
    period_desc, months = data_processor.get_period_description(df, "MTH" if "MTH" in csv_path.name else "YTD")
    report_type = "มิติประเภทต้นทุน" if "COSTTYPE" in csv_path.name else "มิติหมวดบัญชี"
    cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {report_type}"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    cell = ws.cell(row=4, column=2)
    cell.value = period_desc
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Info box (Column G onwards, row 2-5)
    info_box_content = """วัตถุประสงค์ของรายงาน :
เพื่อทราบผลดำเนินงานรายเดือนหลังหักค่าใช้จ่ายรวมทั้งหมด ประกอบด้วย ประเภทต้นทุน ขายและบริหาร
ใช้รายงานเพื่อการบริหารค่าใช้จ่ายแต่ละประเภทให้มีประสิทธิภาพ
ไม่เหมาะสมที่จะใช้เพื่อวัด Performance ส่วนงาน เนื่องจากเป็นการแสดงมิติค่าใช้จ่ายแบบรวม ไม่ได้แสดงเป็นค่าใช้จ่ายที่ส่วนงานมีอำนาจควบคุมได้หรือไม่ได้
ไม่เหมาะสมที่จะใช้เพื่อพิจารณาการยกเลิกบริการ เนื่องจากไม่ได้แสดงมิติต้นทุนผันแปรและคงที่ หากยกเลิกบริการแล้ว ควรทราบว่าต้นทุนใดบ้างลดลงได้ ต้นทุนใดบ้างยังคงอยู่ เพื่อเปรียบกับรายได้ที่จะหายไป"""

    info_cell = ws.cell(row=2, column=7)
    info_cell.value = info_box_content
    info_cell.font = Font(name=font_name, size=14)
    info_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    info_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
    info_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=2, start_column=7, end_row=5, end_column=10)

    # Build column structure
    current_col = start_col
    column_mapping = {}

    # Column B: รายละเอียด
    detail_cell = ws.cell(row=start_row + 1, column=current_col + 1)
    detail_cell.value = "รายละเอียด"
    detail_cell.font = Font(name=font_name, size=font_size, bold=True)
    detail_cell.fill = PatternFill(start_color="F4DEDC", end_color="F4DEDC", fill_type="solid")
    detail_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=start_row + 1, start_column=current_col + 1,
                   end_row=start_row + header_rows, end_column=current_col + 1)
    ws.column_dimensions[get_column_letter(current_col + 1)].width = 50
    current_col += 1

    # Column C: รวมทั้งสิ้น
    grand_total_col = current_col
    gt_cell = ws.cell(row=start_row + 1, column=grand_total_col + 1)
    gt_cell.value = "รวมทั้งสิ้น"
    gt_cell.font = Font(name=font_name, size=font_size, bold=True)
    gt_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    gt_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gt_cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=start_row + 1, start_column=grand_total_col + 1,
                   end_row=start_row + header_rows, end_column=grand_total_col + 1)
    ws.column_dimensions[get_column_letter(grand_total_col + 1)].width = 20
    column_mapping[grand_total_col] = ("GRAND_TOTAL", None, None, None, None)
    current_col += 1

    # For each BU
    for bu in bu_list:
        bu_start_col = current_col
        bu_color = BU_COLORS.get(bu, 'FFFFFF')

        # BU total column (รวม + BU name)
        bu_total_col = current_col
        bu_total_cell = ws.cell(row=start_row + 1, column=bu_total_col + 1)
        bu_total_cell.value = f"รวม {bu}"
        bu_total_cell.font = Font(name=font_name, size=font_size, bold=True)
        bu_total_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
        bu_total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bu_total_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        ws.merge_cells(start_row=start_row + 1, start_column=bu_total_col + 1,
                       end_row=start_row + header_rows, end_column=bu_total_col + 1)
        ws.column_dimensions[get_column_letter(bu_total_col + 1)].width = 18
        column_mapping[bu_total_col] = ("BU_TOTAL", bu, None, None, None)
        current_col += 1

        # Service groups
        for sg in service_group_dict.get(bu, []):
            sg_start_col = current_col
            products = product_dict[bu][sg]

            if products:
                # Write products
                for product_key, product_name in products:
                    col = current_col

                    # Row 3: PRODUCT_KEY
                    pk_cell = ws.cell(row=start_row + 3, column=col + 1)
                    pk_cell.value = product_key
                    pk_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pk_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pk_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    # Row 4: PRODUCT_NAME
                    pn_cell = ws.cell(row=start_row + 4, column=col + 1)
                    pn_cell.value = product_name
                    pn_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pn_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    pn_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pn_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    ws.column_dimensions[get_column_letter(col + 1)].width = 18
                    column_mapping[col] = ("PRODUCT", bu, sg, product_key, product_name)
                    current_col += 1

                sg_end_col = current_col - 1

                # Row 2: SERVICE_GROUP (merge across all products)
                if sg_start_col <= sg_end_col:
                    sg_cell = ws.cell(row=start_row + 2, column=sg_start_col + 1)
                    sg_cell.value = sg
                    sg_cell.font = Font(name=font_name, size=font_size, bold=True)
                    sg_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                    sg_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sg_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if sg_end_col > sg_start_col:
                        ws.merge_cells(start_row=start_row + 2, start_column=sg_start_col + 1,
                                       end_row=start_row + 2, end_column=sg_end_col + 1)

        bu_end_col = current_col - 1

        # Row 1: BU (merge across all service groups)
        if bu_start_col <= bu_end_col:
            bu_cell = ws.cell(row=start_row + 1, column=bu_start_col + 1)
            bu_cell.value = bu
            bu_cell.font = Font(name=font_name, size=font_size, bold=True)
            bu_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
            bu_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bu_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if bu_end_col > bu_start_col:
                ws.merge_cells(start_row=start_row + 1, start_column=bu_start_col + 1,
                               end_row=start_row + 1, end_column=bu_end_col + 1)

    # Write data rows
    current_row = start_row + header_rows
    all_row_data = {}

    for level, label, is_calc, formula, is_bold in ROW_ORDER:
        if not label:
            current_row += 1
            continue

        # Write label
        cell = ws.cell(row=current_row + 1, column=start_col + 1)
        cell.value = label
        cell.font = Font(name=font_name, size=font_size, bold=is_bold)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Color for section headers
        if is_bold and level == 0:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        # Get row data
        if is_calculated_row(label):
            row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
        else:
            row_data = aggregator.get_row_data(label, bu_list, service_group_dict)

        all_row_data[label] = row_data

        # Get GROUP and SUB_GROUP for this row
        group, sub_group = get_group_sub_group(label)

        # Write values
        for col_idx, (col_type, bu, sg, pk, pn) in column_mapping.items():
            if col_type == "GRAND_TOTAL":
                value = row_data.get("GRAND_TOTAL", 0)
            elif col_type == "BU_TOTAL":
                value = row_data.get(f"BU_TOTAL_{bu}", 0)
            elif col_type == "PRODUCT":
                if group and not is_calculated_row(label):
                    value = aggregator.get_value_by_product(group, sub_group, bu, sg, pk)
                else:
                    value = 0
            else:
                continue

            cell = ws.cell(row=current_row + 1, column=col_idx + 1)

            if value is None:
                cell.value = ""
            elif isinstance(value, float) and "สัดส่วน" in label:
                cell.value = value
                cell.number_format = '0.00%'
            else:
                cell.value = value
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'

            cell.font = Font(name=font_name, size=font_size, bold=is_bold)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        current_row += 1

    # Freeze panes
    ws.freeze_panes = f"{get_column_letter(grand_total_col + 2)}{start_row + header_rows + 1}"

    # Add remarks at bottom
    remark_file = csv_path.parent / f"remark_{csv_path.stem.split('_')[-1]}.txt"
    if remark_file.exists():
        try:
            # Try multiple encodings
            for enc in ['utf-8', 'tis-620', 'cp874']:
                try:
                    with open(remark_file, 'r', encoding=enc) as f:
                        remark_content = f.read()
                    break
                except:
                    continue

            current_row += 2
            remark_title_cell = ws.cell(row=current_row + 1, column=start_col + 1)
            remark_title_cell.value = "หมายเหตุ"
            remark_title_cell.font = Font(name=font_name, size=14, bold=True)
            remark_title_cell.alignment = Alignment(horizontal='left', vertical='top')

            current_row += 1
            remark_cell = ws.cell(row=current_row + 1, column=start_col + 1)
            remark_cell.value = remark_content
            remark_cell.font = Font(name=font_name, size=14)
            remark_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        except Exception as e:
            logger.warning(f"Could not read remark file: {e}")

    # Save
    wb.save(output_path)
    logger.info(f"Report saved to: {output_path}")


if __name__ == "__main__":
    csv_path = Path("../data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    output_path = Path("./output/final_report.xlsx")

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)

    output_path.parent.mkdir(exist_ok=True)

    generate_final_report(csv_path, output_path)
    logger.info("Done!")
