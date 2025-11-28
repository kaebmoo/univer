#!/usr/bin/env python3
"""
Test Script 3: Comparison Test
Generate reports with both old and new methods and compare
"""
import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent))

print("="*60)
print("Test 3: Compare Old vs New")
print("="*60)

try:
    print("\n1. Checking for data file...")
    csv_path = Path("data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    
    if not csv_path.exists():
        print(f"   ❌ Data file not found: {csv_path}")
        sys.exit(1)
    
    print(f"   ✅ Data file found")
    
    # ============================================
    # Generate with OLD method (main_generator.py)
    # ============================================
    print("\n2. Generating with OLD method (main_generator.py)...")
    
    from src.data_loader import CSVLoader, DataProcessor
    from openpyxl import Workbook
    
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)
    
    data_processor = DataProcessor()
    df = data_processor.process_data(df)
    
    # Import and run old function
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "main_generator",
        "main_generator.py"
    )
    main_gen = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(main_gen)
    
    old_output = Path("output/old_method.xlsx")
    old_output.parent.mkdir(exist_ok=True)
    
    main_gen.generate_correct_report(csv_path, old_output)
    
    old_size = old_output.stat().st_size / 1024
    print(f"   ✅ OLD method done: {old_size:.1f} KB")
    
    # ============================================
    # Generate with NEW method (report_generator)
    # ============================================
    print("\n3. Generating with NEW method (report_generator)...")
    
    from src.report_generator import ReportBuilder, ReportConfig
    
    # Reload data
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)
    df = data_processor.process_data(df)
    
    config = ReportConfig(
        report_type="COSTTYPE",
        period_type="MTH",
        detail_level="BU_SG_PRODUCT"
    )
    
    builder = ReportBuilder(config)
    
    new_output = Path("output/new_method.xlsx")
    
    # Load remark
    remark_file = csv_path.parent / "remark_MTH.txt"
    remark_content = ""
    if remark_file.exists():
        try:
            for enc in ['utf-8', 'tis-620', 'cp874']:
                try:
                    with open(remark_file, 'r', encoding=enc) as f:
                        remark_content = f.read()
                    break
                except:
                    continue
        except:
            pass
    
    builder.generate_report(df, new_output, remark_content)
    
    new_size = new_output.stat().st_size / 1024
    print(f"   ✅ NEW method done: {new_size:.1f} KB")
    
    # ============================================
    # Compare file sizes
    # ============================================
    print("\n4. Comparing outputs...")
    print(f"\n   OLD method: {old_output.name} ({old_size:.1f} KB)")
    print(f"   NEW method: {new_output.name} ({new_size:.1f} KB)")
    print(f"   Difference: {abs(old_size - new_size):.1f} KB")
    
    if abs(old_size - new_size) < 5:
        print("\n   ✅ File sizes similar (within 5 KB)")
    else:
        print("\n   ⚠️  File sizes differ significantly")
    
    # ============================================
    # Compare content (basic)
    # ============================================
    print("\n5. Loading workbooks for comparison...")
    from openpyxl import load_workbook
    
    wb_old = load_workbook(old_output)
    wb_new = load_workbook(new_output)
    
    ws_old = wb_old.active
    ws_new = wb_new.active
    
    # Compare dimensions
    old_dims = f"{ws_old.max_row}x{ws_old.max_column}"
    new_dims = f"{ws_new.max_row}x{ws_new.max_column}"
    
    print(f"\n   OLD dimensions: {old_dims}")
    print(f"   NEW dimensions: {new_dims}")
    
    if old_dims == new_dims:
        print("   ✅ Dimensions match")
    else:
        print("   ⚠️  Dimensions differ")
    
    # Sample cell comparison
    print("\n6. Comparing sample cells...")
    test_cells = [
        ('B2', 'Company name'),
        ('B3', 'Report title'),
        ('B6', 'Label column header'),
        ('C6', 'Grand total header'),
    ]
    
    matches = 0
    for cell_ref, desc in test_cells:
        old_val = ws_old[cell_ref].value
        new_val = ws_new[cell_ref].value
        
        if old_val == new_val:
            print(f"   ✅ {cell_ref} ({desc}): Match")
            matches += 1
        else:
            print(f"   ⚠️  {cell_ref} ({desc}): Different")
            print(f"      OLD: {old_val}")
            print(f"      NEW: {new_val}")
    
    print(f"\n   Matched: {matches}/{len(test_cells)}")
    
    print("\n" + "="*60)
    print("✅ COMPARISON TEST COMPLETE!")
    print("="*60)
    print("\nGenerated files:")
    print(f"  OLD: {old_output.absolute()}")
    print(f"  NEW: {new_output.absolute()}")
    print("\nManual verification needed:")
    print("  1. Open both files in Excel")
    print("  2. Compare columns side-by-side")
    print("  3. Verify calculations")
    print("  4. Check formatting")
    
except Exception as e:
    print("\n" + "="*60)
    print("❌ COMPARISON TEST FAILED!")
    print("="*60)
    print(f"\nError: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
