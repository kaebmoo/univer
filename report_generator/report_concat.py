import os
import logging
import argparse
import pandas as pd
import openpyxl
from copy import copy

# Configure basic logging for this script
logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')

# 1. หาตำแหน่งที่ตั้งจริงของไฟล์ script (report_concat.py) ในเครื่อง
# ผลลัพธ์จะเป็น .../univer/report_generator
script_dir = os.path.dirname(os.path.abspath(__file__))

# 2. กำหนด input/output โดยอ้างอิงจากตำแหน่ง script
# โครงสร้างโฟลเดอร์ที่ code นี้คาดหวังคือ:
# univer/
#   └── report_generator/
#       ├── report_concat.py  (ไฟล์นี้)
#       ├── data/
#       └── output/           (ต้องมีไฟล์ excel อยู่ในนี้) (ไฟล์ผลลัพธ์จะมาโผล่ที่นี่)

input_dir = os.path.join(script_dir, 'output')
output_dir = os.path.join(script_dir, 'output')

# Parse command line arguments
parser = argparse.ArgumentParser(description='Concatenate monthly reports')
parser.add_argument('--month', type=str, help='Specific month to process (YYYYMM format, e.g., 202509)')
args = parser.parse_args()

# Debug: ปริ้นท์ออกมาดูว่า path ถูกต้องไหม
logging.info(f"Script Location: {script_dir}")
logging.info(f"Reading from:    {input_dir}")
logging.info(f"Writing to:      {output_dir}")
if args.month:
    logging.info(f"Processing month: {args.month}")
else:
    logging.info("Processing all months found")
logging.info("-" * 30)

# ตรวจสอบว่ามี folder output หรือไม่ ถ้าไม่มีให้สร้าง
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 1. Get a list of all files in the input_dir (ดึงรายชื่อไฟล์ครั้งเดียว ใช้ได้ทั้ง 2 งาน)
all_files = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]

def extract_months_from_files(files, report_type):
    """Extract unique YYYYMM values from file names"""
    months = set()
    for f in files:
        if report_type in f and f.endswith('.xlsx'):
            # Extract date part from filename (last part before .xlsx)
            date_part = f.split('_')[-1].replace('.xlsx', '')
            if len(date_part) >= 6 and date_part[:6].isdigit():
                months.add(date_part[:6])  # YYYYMM
    return sorted(list(months))


def copy_sheet_with_formatting(source_file_path, target_workbook, sheet_name):
    """Copy a sheet from source file to target workbook with formatting"""
    source_wb = openpyxl.load_workbook(source_file_path)
    source_sheet = source_wb.active

    new_sheet = target_workbook.create_sheet(title=sheet_name)

    # Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copy dimensions
    for row_idx, rd in source_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row_idx] = copy(rd)
    for col_idx, cd in source_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col_idx] = copy(cd)

    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_cell_range))

    source_wb.close()


def create_report(files_list, date_part, report_type, patterns, sheet_names_map):
    """Create a combined report for a specific month and report type"""
    if report_type == 'MTH':
        output_filename = f"Report_NT_{date_part}.xlsx"
    else:  # YTD
        output_filename = f"Report_NT_YTD_{date_part}.xlsx"

    output_filepath = os.path.join(output_dir, output_filename)
    logging.info(f"\nCreating {report_type} report for {date_part}: {output_filename}")

    # Create new workbook
    output_wb = openpyxl.Workbook()
    default_sheet = output_wb['Sheet']
    output_wb.remove(default_sheet)

    # Iterate through patterns
    for pattern in patterns:
        found_file = None
        for f in files_list:
            if f.startswith(pattern) and date_part in f and f.endswith(f'_{date_part}.xlsx'):
                found_file = f
                break

        if found_file:
            full_file_path = os.path.join(input_dir, found_file)
            sheet_name = sheet_names_map.get(pattern, pattern)
            copy_sheet_with_formatting(full_file_path, output_wb, sheet_name)
            logging.info(f"  - Copied '{found_file}' to sheet '{sheet_name}'")
        else:
            logging.warning(f"  - Warning: No file found for pattern '{pattern}' with date {date_part}")

    # Save
    output_wb.save(output_filepath)
    logging.info(f"Successfully created: {output_filepath}")
    return output_filepath


logging.info("Starting Process...\n")

# Define filename patterns and sheet names
file_patterns_mth = [
    "PL_COSTTYPE_MTH_BU_ONLY",
    "PL_COSTTYPE_MTH_BU_SG",
    "PL_COSTTYPE_MTH_BU_SG_PRODUCT",
    "PL_GLGROUP_MTH_BU_ONLY",
    "PL_GLGROUP_MTH_BU_SG",
    "PL_GLGROUP_MTH_BU_SG_PRODUCT"
]

file_patterns_ytd = [
    "PL_COSTTYPE_YTD_BU_ONLY",
    "PL_COSTTYPE_YTD_BU_SG",
    "PL_COSTTYPE_YTD_BU_SG_PRODUCT",
    "PL_GLGROUP_YTD_BU_ONLY",
    "PL_GLGROUP_YTD_BU_SG",
    "PL_GLGROUP_YTD_BU_SG_PRODUCT"
]

thai_sheet_names_map = {
    "PL_COSTTYPE_MTH_BU_ONLY": "ต้นทุน_กลุ่มธุรกิจ",
    "PL_COSTTYPE_MTH_BU_SG": "ต้นทุน_กลุ่มบริการ",
    "PL_COSTTYPE_MTH_BU_SG_PRODUCT": "ต้นทุน_บริการ",
    "PL_GLGROUP_MTH_BU_ONLY": "หมวดบัญชี_กลุ่มธุรกิจ",
    "PL_GLGROUP_MTH_BU_SG": "หมวดบัญชี_กลุ่มบริการ",
    "PL_GLGROUP_MTH_BU_SG_PRODUCT": "หมวดบัญชี_บริการ",
    "PL_COSTTYPE_YTD_BU_ONLY": "ต้นทุน_กลุ่มธุรกิจ",
    "PL_COSTTYPE_YTD_BU_SG": "ต้นทุน_กลุ่มบริการ",
    "PL_COSTTYPE_YTD_BU_SG_PRODUCT": "ต้นทุน_บริการ",
    "PL_GLGROUP_YTD_BU_ONLY": "หมวดบัญชี_กลุ่มธุรกิจ",
    "PL_GLGROUP_YTD_BU_SG": "หมวดบัญชี_กลุ่มบริการ",
    "PL_GLGROUP_YTD_BU_SG_PRODUCT": "หมวดบัญชี_บริการ"
}

# Filter files
mth_files = sorted([f for f in all_files if 'MTH' in f and f.endswith('.xlsx')])
ytd_files = sorted([f for f in all_files if 'YTD' in f and f.endswith('.xlsx')])

# Extract available months
mth_months = extract_months_from_files(mth_files, 'MTH')
ytd_months = extract_months_from_files(ytd_files, 'YTD')

logging.info(f"Found MTH months: {mth_months}")
logging.info(f"Found YTD months: {ytd_months}")

# Determine which months to process
if args.month:
    # Process only specified month
    months_to_process_mth = [args.month] if args.month in mth_months else []
    months_to_process_ytd = [args.month] if args.month in ytd_months else []

    if not months_to_process_mth and not months_to_process_ytd:
        logging.error(f"Month {args.month} not found in any files")
        exit(1)
else:
    # Process all months found
    months_to_process_mth = mth_months
    months_to_process_ytd = ytd_months

# ==========================================
# PART 1: สร้างรายงานเดือน (MTH)
# ==========================================
logging.info("\n" + "="*40)
logging.info("--- Processing MTH Reports ---")
logging.info("="*40)

if not months_to_process_mth:
    logging.warning("No MTH months to process")
else:
    for month in months_to_process_mth:
        create_report(mth_files, month, 'MTH', file_patterns_mth, thai_sheet_names_map)

# ==========================================
# PART 2: สร้างรายงานปี (YTD)
# ==========================================
logging.info("\n" + "="*40)
logging.info("--- Processing YTD Reports ---")
logging.info("="*40)

if not months_to_process_ytd:
    logging.warning("No YTD months to process")
else:
    for month in months_to_process_ytd:
        create_report(ytd_files, month, 'YTD', file_patterns_ytd, thai_sheet_names_map)

logging.info("\n" + "="*40)
logging.info("All processes completed.")
logging.info("="*40)