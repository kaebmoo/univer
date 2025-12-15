#!/usr/bin/env python3
"""Check specific cost calculation rows in Excel"""
from pathlib import Path
from openpyxl import load_workbook

report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')
wb = load_workbook(report_path)
ws = wb.active

# Column 25 is the satellite summary column
satellite_col = 25

# Find specific rows
target_rows = {
    "     1. ต้นทุนบริการรวม": None,
    "     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ": 32326181.81,
    "     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ": 283920980.50,
}

print("Searching for cost calculation rows...")
print("="*80)

found_rows = {}

for row_idx in range(1, 100):
    label_cell = ws.cell(row_idx, 1)  # Column A
    if label_cell.value:
        label = str(label_cell.value).strip()
        if label in target_rows:
            satellite_value = ws.cell(row_idx, satellite_col).value
            found_rows[label] = {
                'row': row_idx,
                'value': satellite_value,
                'expected': target_rows[label]
            }
            print(f"\nFound row {row_idx}: {label}")
            print(f"  Satellite value: {satellite_value}")
            if target_rows[label] is not None:
                print(f"  Expected: {target_rows[label]:,.2f}")
                if satellite_value is not None:
                    diff = abs(float(satellite_value) - target_rows[label])
                    if diff < 1:
                        print(f"  ✓ MATCH")
                    else:
                        print(f"  ❌ MISMATCH (diff: {diff:,.2f})")
                else:
                    print(f"  ❌ Got None/empty")

print("\n" + "="*80)
print("SUMMARY:")
print("="*80)

all_pass = True
for label, expected in target_rows.items():
    if label not in found_rows:
        print(f"❌ '{label}' - NOT FOUND")
        all_pass = False
    elif expected is not None:
        info = found_rows[label]
        if info['value'] is None:
            print(f"❌ '{label}' - Empty value")
            all_pass = False
        else:
            diff = abs(float(info['value']) - expected)
            if diff < 1:
                print(f"✓ '{label}' - PASS")
            else:
                print(f"❌ '{label}' - FAIL (got {info['value']:,.2f}, expected {expected:,.2f})")
                all_pass = False

if all_pass:
    print("\n✅ ALL CHECKS PASSED")
else:
    print("\n❌ SOME CHECKS FAILED")
