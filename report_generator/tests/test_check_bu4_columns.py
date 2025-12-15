#!/usr/bin/env python3
"""Check BU 4 columns in detail"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Load the generated report
report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')

if not report_path.exists():
    print(f"Report not found: {report_path}")
    exit(1)

wb = load_workbook(report_path)
ws = wb.active

print("BU 4 (FIXED LINE & BROADBAND) Columns:")
print("="*80)

# Row 6 has BU headers
# Row 7 has SG headers
bu4_start = None
bu4_end = None

# Find BU 4 columns
for col_idx in range(1, 50):
    cell = ws.cell(6, col_idx)
    if cell.value and '4.กลุ่มธุรกิจ FIXED LINE & BROADBAND' in str(cell.value):
        if bu4_start is None:
            bu4_start = col_idx
        bu4_end = col_idx

if bu4_start:
    print(f"\nBU 4 columns span: {bu4_start} to {bu4_end}")
    print("\nRow 6 (BU header):")
    for col_idx in range(bu4_start, bu4_end + 1):
        cell = ws.cell(6, col_idx)
        if cell.value:
            print(f"  Col {col_idx} ({cell.column_letter}): {cell.value}")

    print("\nRow 7 (SG header):")
    for col_idx in range(bu4_start, bu4_end + 1):
        cell = ws.cell(7, col_idx)
        if cell.value:
            print(f"  Col {col_idx} ({cell.column_letter}): {cell.value}")

    # Check if there are SATELLITE columns
    has_satellite = False
    for col_idx in range(bu4_start, bu4_end + 1):
        cell_row6 = ws.cell(6, col_idx)
        cell_row7 = ws.cell(7, col_idx)
        if (cell_row6.value and 'SATELLITE' in str(cell_row6.value)) or \
           (cell_row7.value and 'SATELLITE' in str(cell_row7.value)):
            has_satellite = True
            print(f"\n✓ Found SATELLITE at column {col_idx} ({cell.column_letter})")

    if not has_satellite:
        print("\n❌ NO SATELLITE columns found in BU 4")
        print("\nThis suggests SATELLITE split is not enabled for COSTTYPE BU_SG report")
else:
    print("❌ BU 4 not found")
