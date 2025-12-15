#!/usr/bin/env python3
"""Count actual columns in Excel file"""
from pathlib import Path
from openpyxl import load_workbook

report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')
wb = load_workbook(report_path)
ws = wb.active

# Count non-empty columns in header row 6
non_empty_cols = 0
for col_idx in range(1, 100):
    cell = ws.cell(6, col_idx)
    if cell.value:
        non_empty_cols = col_idx

print(f"Non-empty columns in row 6 (BU header): {non_empty_cols}")

# Also check row 7 (SG header)
non_empty_cols_7 = 0
for col_idx in range(1, 100):
    cell = ws.cell(7, col_idx)
    if cell.value:
        non_empty_cols_7 = col_idx

print(f"Non-empty columns in row 7 (SG header): {non_empty_cols_7}")

# Count columns with data in row 10 (first data row)
non_empty_data = 0
for col_idx in range(1, 100):
    cell = ws.cell(10, col_idx)
    if cell.value is not None:
        non_empty_data = col_idx

print(f"Non-empty columns in row 10 (first data): {non_empty_data}")

# Check specifically around BU 4 columns
print("\nBU 4 region (columns 20-30):")
for col_idx in range(20, 31):
    bu_cell = ws.cell(6, col_idx)
    sg_cell = ws.cell(7, col_idx)
    data_cell = ws.cell(10, col_idx)

    if bu_cell.value or sg_cell.value or data_cell.value is not None:
        print(f"  Col {col_idx} ({ws.cell(1, col_idx).column_letter}):")
        if bu_cell.value:
            print(f"    BU: {bu_cell.value}")
        if sg_cell.value:
            print(f"    SG: {sg_cell.value}")
        if data_cell.value is not None:
            print(f"    Data: {data_cell.value}")
