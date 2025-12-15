#!/usr/bin/env python3
"""Find cost-related rows in Excel"""
from pathlib import Path
from openpyxl import load_workbook

report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')
wb = load_workbook(report_path)
ws = wb.active

print("Searching for rows containing 'ต้นทุนบริการ'...")
print("="*80)

for row_idx in range(1, 100):
    label_cell = ws.cell(row_idx, 1)
    if label_cell.value:
        label = str(label_cell.value).strip()
        if 'ต้นทุนบริการ' in label:
            print(f"Row {row_idx}: {label}")
