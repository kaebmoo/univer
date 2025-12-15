#!/usr/bin/env python3
"""Inspect Excel structure"""
import openpyxl
from pathlib import Path

filepath = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')
wb = openpyxl.load_workbook(filepath)
ws = wb.active

print("Excel Structure Inspection")
print("="*80)

# Show first 15 rows and columns
print("\nFirst 15 rows x first 30 columns:")
for row_idx in range(1, 16):
    print(f"\nRow {row_idx:2d}:")
    for col_idx in range(1, min(31, ws.max_column + 1)):
        value = ws.cell(row_idx, col_idx).value
        if value:
            print(f"  Col {col_idx:2d}: {str(value)[:70]}")

# Look for SATELLITE in entire first 20 rows
print("\n" + "="*80)
print("Searching for 'SATELLITE' or 'Satellite' in rows 1-20:")
for row_idx in range(1, 21):
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row_idx, col_idx).value
        if value and ('SATELLITE' in str(value).upper() or 'Satellite' in str(value)):
            print(f"  Row {row_idx:2d}, Col {col_idx:2d}: {value}")

wb.close()
