#!/usr/bin/env python3
"""Comprehensive verification of all SATELLITE reports"""
import openpyxl
from pathlib import Path

def verify_report(filepath):
    """Verify a single report"""
    print(f"\n{'='*80}")
    print(f"📊 {filepath.name}")
    print('='*80)

    if not filepath.exists():
        print(f"❌ File not found!")
        return False

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Find SATELLITE columns in header rows (rows 6-10)
    satellite_cols = {}
    for row_idx in range(6, 11):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value
            if value and ('SATELLITE' in str(value).upper() or 'Satellite' in str(value)):
                if col_idx not in satellite_cols:
                    satellite_cols[col_idx] = str(value)

    if not satellite_cols:
        print("❌ NO SATELLITE columns found!")
        wb.close()
        return False

    print(f"\n✅ Found {len(satellite_cols)} SATELLITE columns:")
    for col_idx in sorted(satellite_cols.keys()):
        print(f"   Col {col_idx:2d}: {satellite_cols[col_idx]}")

    # Identify column types
    summary_col = None
    nt_col = None
    thaicom_col = None

    for col_idx, name in satellite_cols.items():
        if 'รวม' in name or 'SUMMARY' in name.upper():
            summary_col = col_idx
        elif '4.5.1' in name or 'NT' in name.upper():
            nt_col = col_idx
        elif '4.5.2' in name or 'ไทยคม' in name:
            thaicom_col = col_idx

    # Verify data row (Row 10 = รายได้)
    data_row = 10
    print(f"\nData verification (Row {data_row} - รายได้):")

    all_pass = True

    if summary_col:
        summary_val = ws.cell(data_row, summary_col).value or 0
        print(f"   Summary: {summary_val:>15,.2f}")

    if nt_col:
        nt_val = ws.cell(data_row, nt_col).value or 0
        print(f"   4.5.1 NT: {nt_val:>15,.2f}")

    if thaicom_col:
        thaicom_val = ws.cell(data_row, thaicom_col).value or 0
        print(f"   4.5.2 ไทยคม: {thaicom_val:>15,.2f}")

    # Verify calculation
    if summary_col and nt_col and thaicom_col:
        summary_val = ws.cell(data_row, summary_col).value or 0
        detail_sum = (ws.cell(data_row, nt_col).value or 0) + (ws.cell(data_row, thaicom_col).value or 0)

        print(f"\nCalculation check:")
        print(f"   4.5.1 + 4.5.2 = {detail_sum:,.2f}")
        print(f"   Summary value = {summary_val:,.2f}")

        if abs(detail_sum - summary_val) < 0.01:
            print(f"   ✅ PASS: Summary equals sum of details")
        else:
            print(f"   ❌ FAIL: Difference = {abs(detail_sum - summary_val):,.2f}")
            all_pass = False
    else:
        print(f"\n⚠️  WARNING: Not all columns found for calculation check")
        print(f"   Summary: {'✓' if summary_col else '✗'}")
        print(f"   4.5.1: {'✓' if nt_col else '✗'}")
        print(f"   4.5.2: {'✓' if thaicom_col else '✗'}")

    # Check column ordering
    if summary_col and nt_col and thaicom_col:
        print(f"\nColumn ordering check:")
        if summary_col < nt_col and nt_col < thaicom_col:
            print(f"   ✅ PASS: Summary ({summary_col}) < NT ({nt_col}) < ไทยคม ({thaicom_col})")
        else:
            print(f"   ❌ FAIL: Wrong order - Summary:{summary_col}, NT:{nt_col}, ไทยคม:{thaicom_col}")
            all_pass = False

    wb.close()
    return all_pass

# Test all reports
print("\n" + "="*80)
print("🔍 SATELLITE FEATURE VERIFICATION")
print("="*80)

output_dir = Path('output')
reports = [
    ('COSTTYPE', 'BU_SG', 'PL_COSTTYPE_YTD_BU_SG_202510.xlsx'),
    ('COSTTYPE', 'BU_SG_PRODUCT', 'PL_COSTTYPE_YTD_BU_SG_PRODUCT_202510.xlsx'),
    ('GLGROUP', 'BU_SG', 'PL_GLGROUP_YTD_BU_SG_202510.xlsx'),
    ('GLGROUP', 'BU_SG_PRODUCT', 'PL_GLGROUP_YTD_BU_SG_PRODUCT_202510.xlsx'),
]

results = {}
for report_type, detail_level, filename in reports:
    report_key = f"{report_type}_{detail_level}"
    filepath = output_dir / filename
    results[report_key] = verify_report(filepath)

# Summary
print("\n" + "="*80)
print("📋 SUMMARY")
print("="*80)

all_passed = True
for report_type, detail_level, filename in reports:
    report_key = f"{report_type}_{detail_level}"
    status = "✅ PASS" if results.get(report_key, False) else "❌ FAIL"
    print(f"   {status}  {report_type:10s} {detail_level:15s}")
    if not results.get(report_key, False):
        all_passed = False

print("="*80)
if all_passed:
    print("🎉 ALL TESTS PASSED!")
else:
    print("⚠️  SOME TESTS FAILED - Review output above")
print("="*80)
