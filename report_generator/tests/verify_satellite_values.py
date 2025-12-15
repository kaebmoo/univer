#!/usr/bin/env python3
"""Verify SATELLITE values in generated reports"""
import openpyxl
from pathlib import Path

def check_report(filepath):
    """Check a single report file"""
    print(f"\n{'='*80}")
    print(f"Checking: {filepath.name}")
    print('='*80)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Find column headers in row 10 (0-indexed = row 9)
    # Look for SATELLITE-related columns
    header_row = 10
    satellite_cols = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col_idx).value
        if cell_value and 'SATELLITE' in str(cell_value).upper():
            satellite_cols[col_idx] = cell_value

    print(f"\nFound {len(satellite_cols)} SATELLITE columns:")
    for col_idx, name in sorted(satellite_cols.items()):
        print(f"  Column {col_idx:3d}: {name}")

    # Find a sample data row (e.g., row 15 - "รายได้จากการให้บริการ")
    # Check values in SATELLITE columns
    print(f"\nSample values (Row 12 - รายได้จากการให้บริการ):")
    for col_idx in sorted(satellite_cols.keys()):
        value = ws.cell(12, col_idx).value
        col_name = satellite_cols[col_idx]
        print(f"  {col_name:40s}: {value}")

    # Verify summary = sum of details
    if len(satellite_cols) >= 3:
        cols_list = sorted(satellite_cols.keys())
        # Assume: first is summary, rest are details
        # But we need to identify which is which by name
        summary_col = None
        detail_cols = []

        for col_idx, name in satellite_cols.items():
            if 'รวม' in name or 'SUMMARY' in name.upper():
                summary_col = col_idx
            elif '4.5.1' in name or '4.5.2' in name:
                detail_cols.append(col_idx)

        if summary_col and len(detail_cols) >= 2:
            print(f"\nVerifying calculation (Row 12):")
            summary_val = ws.cell(12, summary_col).value or 0
            detail_sum = sum(ws.cell(12, col).value or 0 for col in detail_cols)

            print(f"  Summary column value: {summary_val:,.2f}")
            print(f"  Sum of details: {detail_sum:,.2f}")

            if abs(summary_val - detail_sum) < 0.01:
                print(f"  ✅ PASS: Values match!")
            else:
                print(f"  ❌ FAIL: Values don't match! Difference: {abs(summary_val - detail_sum):,.2f}")

    wb.close()

# Check all generated reports
output_dir = Path('output')
reports = [
    'PL_COSTTYPE_YTD_BU_SG_202510.xlsx',
    'PL_COSTTYPE_YTD_BU_SG_PRODUCT_202510.xlsx',
    'PL_GLGROUP_YTD_BU_SG_202510.xlsx',
    'PL_GLGROUP_YTD_BU_SG_PRODUCT_202510.xlsx',
]

for report_name in reports:
    report_path = output_dir / report_name
    if report_path.exists():
        check_report(report_path)
    else:
        print(f"\n❌ Not found: {report_name}")

print(f"\n{'='*80}")
print("Verification complete!")
print('='*80)
