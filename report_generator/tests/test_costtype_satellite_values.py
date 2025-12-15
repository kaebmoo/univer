#!/usr/bin/env python3
"""Verify COSTTYPE satellite summary values for cost calculation rows"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from config.satellite_config import SATELLITE_SUMMARY_ID

# Load the generated report
report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')

if not report_path.exists():
    print(f"Report not found: {report_path}")
    exit(1)

wb = load_workbook(report_path)
ws = wb.active

# Find the SATELLITE summary column
# Headers start at row 8 (after freeze panes at row 11, so header should be row 8-10)
satellite_col = None
for col_idx, cell in enumerate(ws[10], start=1):  # Row 10 is the third header row
    if cell.value and '4.5_SUMMARY' in str(cell.value):
        satellite_col = col_idx
        break

if not satellite_col:
    print("❌ SATELLITE summary column not found!")
    exit(1)

print(f"✓ Found SATELLITE summary column at index {satellite_col}")
print(f"  Column letter: {ws.cell(10, satellite_col).column_letter}")
print()

# Check specific rows
test_rows = {
    "     1. ต้นทุนบริการรวม": {"expected": None, "row": None},
    "     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ": {"expected": 32326181.81, "row": None},
    "         สัดส่วนต่อรายได้": {"expected": None, "row": None},  # After row 2
    "     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ": {"expected": 283920980.50, "row": None},
}

# Find row numbers
for row_idx in range(11, 100):  # Start after headers
    cell_value = ws.cell(row_idx, 1).value
    if cell_value:
        label = str(cell_value).strip()
        if label in test_rows:
            test_rows[label]["row"] = row_idx
            print(f"Found '{label}' at row {row_idx}")

print()
print("="*80)
print("VERIFICATION RESULTS")
print("="*80)

all_pass = True
for label, info in test_rows.items():
    row_num = info["row"]
    expected = info["expected"]

    if not row_num:
        print(f"❌ '{label}' - Row not found")
        all_pass = False
        continue

    # Get the satellite value
    cell = ws.cell(row_num, satellite_col)
    actual = cell.value

    # Convert to float if numeric
    if actual is not None and not isinstance(actual, str):
        actual_num = float(actual)
    else:
        actual_num = 0 if actual is None else actual

    print(f"\nRow {row_num}: {label}")

    if expected is not None:
        # Check against expected value
        if isinstance(actual_num, (int, float)):
            diff = abs(actual_num - expected)
            if diff < 1:  # Allow small rounding difference
                print(f"  ✓ Satellite: {actual_num:,.2f} (expected: {expected:,.2f})")
            else:
                print(f"  ❌ Satellite: {actual_num:,.2f} (expected: {expected:,.2f}, diff: {diff:,.2f})")
                all_pass = False
        else:
            print(f"  ❌ Satellite: {actual} (expected numeric: {expected:,.2f})")
            all_pass = False
    else:
        # Just report the value
        if isinstance(actual_num, (int, float)):
            print(f"  Satellite: {actual_num:,.2f}")
        else:
            print(f"  Satellite: {actual}")

print()
print("="*80)
if all_pass:
    print("✅ ALL CHECKS PASSED")
else:
    print("❌ SOME CHECKS FAILED")
print("="*80)
