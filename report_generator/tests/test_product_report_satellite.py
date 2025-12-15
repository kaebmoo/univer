#!/usr/bin/env python3
"""Check satellite values in COSTTYPE BU_SG_PRODUCT report"""
from pathlib import Path
from openpyxl import load_workbook

report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_PRODUCT_202510.xlsx')
wb = load_workbook(report_path)
ws = wb.active

# First, find the satellite summary column for BU 4
print("Finding SATELLITE summary column...")
print("="*80)

satellite_col = None
for col_idx in range(1, 200):
    # Check row 8 (SG header)
    cell = ws.cell(8, col_idx)
    if cell.value and ('SATELLITE' in str(cell.value) or '4.5_SUMMARY' in str(cell.value)) and 'รวม' in str(cell.value):
        satellite_col = col_idx
        print(f"Found SATELLITE summary column at index {col_idx} ({ws.cell(1, col_idx).column_letter})")
        print(f"  Header value: {cell.value}")
        break

if not satellite_col:
    print("❌ SATELLITE summary column not found")
    print("\nChecking all BU 4 columns:")
    # Find BU 4 columns
    for col_idx in range(1, 200):
        bu_cell = ws.cell(6, col_idx)
        sg_cell = ws.cell(8, col_idx)
        if bu_cell.value and '4.กลุ่มธุรกิจ FIXED LINE' in str(bu_cell.value):
            if sg_cell.value:
                print(f"  Col {col_idx}: {sg_cell.value}")
    exit(1)

print()

# Check the three rows
target_rows = {
    "     1. ต้นทุนบริการรวม": {"row": 78, "expected": None},
    "     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ": {"row": 80, "expected": 32326181.81},
    "     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ": {"row": 82, "expected": 283920980.50},
}

print("Checking satellite values:")
print("="*80)

all_pass = True
for label, info in target_rows.items():
    row_num = info['row']
    expected = info['expected']

    # Verify label
    actual_label = ws.cell(row_num, 2).value
    if actual_label != label:
        print(f"\n❌ Row {row_num} label mismatch:")
        print(f"   Expected: '{label}'")
        print(f"   Got: '{actual_label}'")
        all_pass = False
        continue

    # Get satellite value
    satellite_value = ws.cell(row_num, satellite_col).value

    print(f"\nRow {row_num}: {label}")
    print(f"  Satellite: {satellite_value}")

    if expected is not None:
        if satellite_value is None or satellite_value == 0:
            print(f"  ❌ Expected: {expected:,.2f} but got {satellite_value}")
            all_pass = False
        else:
            diff = abs(float(satellite_value) - expected)
            if diff < 1:
                print(f"  ✓ MATCH (expected: {expected:,.2f})")
            else:
                print(f"  ❌ MISMATCH: expected {expected:,.2f}, diff = {diff:,.2f}")
                all_pass = False

print()
print("="*80)
if all_pass:
    print("✅ ALL CHECKS PASSED")
else:
    print("❌ SOME CHECKS FAILED")
print("="*80)
