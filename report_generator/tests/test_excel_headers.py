#!/usr/bin/env python3
"""Check Excel file header structure"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Load the generated report
report_path = Path('output/PL_COSTTYPE_YTD_BU_SG_202510.xlsx')

if not report_path.exists():
    print(f"Report not found: {report_path}")
    exit(1)

wb = load_workbook(report_path)
ws = wb.active

print("Excel Headers:")
print("="*80)

# Check multiple rows for headers
for row_idx in range(1, 15):
    print(f"\nRow {row_idx}:")
    row_values = []
    for col_idx in range(1, 30):  # Check first 30 columns
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            row_values.append(f"  Col {col_idx} ({cell.column_letter}): {cell.value}")

    if row_values:
        for val in row_values[:10]:  # Show first 10
            print(val)
        if len(row_values) > 10:
            print(f"  ... and {len(row_values) - 10} more columns")
