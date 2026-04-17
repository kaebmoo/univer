#!/usr/bin/env python3
"""
verify_report_completeness.py

ตรวจสอบความสมบูรณ์ของ Excel report ที่โปรแกรมสร้างขึ้น
เทียบกับ text-report CSV (TIS-620) ที่ export จากระบบต้นทาง

จับกรณี:
  1. MISSING rows — text-report มีค่า (non-zero) แต่ Excel ไม่มี row นั้น
     (มักเกิดจาก mapping ใน config/row_order.py หรือ config/data_mapping.py ขาด)
  2. VALUE MISMATCH — row ตรงกันแต่ค่าต่างกันเกิน tolerance

Current scope: ระดับ BU (BU_ONLY variants) ทั้ง COSTTYPE และ GLGROUP
Future work: PD Group และ Product level (BU_SG, BU_SG_PRODUCT)

Usage:
    python verify_report_completeness.py --month 202603
    python verify_report_completeness.py --month 202603 --tolerance 1.0
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


TOLERANCE_DEFAULT = 1.0  # บาท — text-report ปัดทศนิยม 2 ตำแหน่ง


# ---------- Normalization helpers ----------

_MAIN_RE = re.compile(r"^\s*0?(\d+)[\.\s]\s*(.*)$")
_SUB_TEXT_RE = re.compile(r"^#\s*\d+\.\s*")
_SUB_EXCEL_RE = re.compile(r"^\s*-\s*")


def canonical_main(label: str) -> str:
    """
    Normalize main group labels from either text-report or Excel.
    Handles:
      '01.รายได้'          → '1.รายได้'   (strip leading zero)
      '1.รายได้'           → '1.รายได้'
      '1 รวมรายได้'        → '1.รายได้'   (GLGROUP Excel display)
      '2 รวมค่าใช้จ่าย'      → '2.ค่าใช้จ่าย'  (GLGROUP Excel display)
    """
    s = (label or "").strip()
    m = _MAIN_RE.match(s)
    if m:
        desc = m.group(2).strip()
        if desc.startswith("รวม"):
            desc = desc[3:].strip()
        return f"{m.group(1)}.{desc}"
    return s


def canonical_sub(label: str) -> str:
    """
    Text-report: '# 09.ค่าใช้จ่ายเกี่ยวกับการกำกับดูแลของ กสทช.' → core
    Excel:       '- ค่าใช้จ่ายเกี่ยวกับการกำกับดูแลของ กสทช.'       → core
    Excel sub-sub: '     - ค่าใช้จ่ายตอบแทนแรงงาน-MSP' → core (with ER convert)
    Case-folded so 'BUSINESS' vs 'Business' match.
    """
    s = (label or "").strip()
    s = _SUB_TEXT_RE.sub("", s)   # strip '# 09.'
    s = _SUB_EXCEL_RE.sub("", s)  # strip '- ' or '   - '
    # Excel แสดง '-MSP' แต่ text-report ใช้ '-ER' สำหรับรายการเดียวกัน
    s = s.replace("-MSP", "-ER").replace("-msp", "-ER")
    return s.strip().casefold()


def parse_thai_number(s: str) -> Optional[float]:
    """Text-report ใช้ trailing minus: '123.45-' = -123.45"""
    if s is None:
        return None
    s = s.strip().replace(",", "")
    if not s:
        return None
    if s.endswith("-"):
        s = "-" + s[:-1]
    try:
        return float(s)
    except ValueError:
        return None


# ---------- Text-report parser ----------

def parse_text_report(path: Path) -> Tuple[List[str], Dict[Tuple[str, str, str], float]]:
    """
    Return:
        bu_names: list of BU column names
        data: {(main_core, sub_core, bu_name): value}
              - sub_core = '' สำหรับ main-group row (total)
    """
    with open(path, encoding="tis-620", errors="replace") as f:
        rows = list(csv.reader(f, delimiter="\t"))

    if len(rows) < 5:
        raise ValueError(f"text-report too short: {path}")

    # Header row 5 (index 4): col 0=รายละเอียด, col 1=รวมทั้งสิ้น, col 2+ = BU names
    header = rows[4]
    bu_cols: List[Tuple[int, str]] = []
    for i in range(2, len(header)):
        name = (header[i] or "").strip()
        if name and name != "N/A":
            bu_cols.append((i, name))

    data: Dict[Tuple[str, str, str], float] = {}
    current_main = ""

    for r in rows[5:]:
        if not r or not (r[0] or "").strip():
            continue
        label = r[0]
        stripped = label.strip()
        is_sub = stripped.startswith("#")
        if is_sub:
            sub_core = canonical_sub(label)
            main_core = current_main
        else:
            main_core = canonical_main(label)
            current_main = main_core
            sub_core = ""  # main-group level

        for col_idx, bu_name in bu_cols:
            if col_idx >= len(r):
                continue
            v = parse_thai_number(r[col_idx])
            if v is None:
                continue
            key = (main_core, sub_core, bu_name)
            data[key] = data.get(key, 0.0) + v

    return [n for _, n in bu_cols], data


# ---------- Excel parser (BU_ONLY variant) ----------

def parse_excel_bu_only(path: Path) -> Tuple[List[str], Dict[Tuple[str, str, str], float]]:
    """
    Layout: col A=label, col B=รวมทั้งสิ้น amount, col C=รวมทั้งสิ้น common size,
            col D,E = BU1 amount+cs, col F,G = BU2 amount+cs, ...
    Row 6 = BU names (merged), row 7 = จำนวนเงิน/Common Size subheader.

    Preserves main-group totals (sub_core=='') and sub-item values, same key shape
    as text-report. For sub-items that have both level-1 parent and level-2 children
    (combined total + detail), level-2 values win (level-1 would double-count).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Layout: col A=blank, col B=label, col C=grand total amount, col D=grand total CS,
    #         col E=BU1 amount, col F=BU1 CS, col G=BU2 amount, col H=BU2 CS, ...
    LABEL_COL = 2
    # Identify amount columns per BU: col with row7='จำนวนเงิน' and BU name from row6.
    # Skip the grand-total amount column (c=3, BU name = "รวมทั้งสิ้น").
    bu_amount_cols: List[Tuple[int, str]] = []
    current_bu_name = None
    for c in range(4, ws.max_column + 1):
        name = ws.cell(row=6, column=c).value
        if name is not None and str(name).strip():
            current_bu_name = str(name).strip()
        sub = ws.cell(row=7, column=c).value
        if current_bu_name and sub and "จำนวนเงิน" in str(sub):
            bu_amount_cols.append((c, current_bu_name))

    # Pass 1: collect all sub-item rows with indent level
    entries: List[Tuple[str, str, str, float, int]] = []  # (main, sub, bu, v, level)
    main_totals: Dict[Tuple[str, str], float] = {}  # (main_core, bu) -> total (main row)
    current_main = ""

    for row_idx in range(8, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=LABEL_COL).value
        if raw is None:
            continue
        label = str(raw)
        if not label.strip():
            continue

        stripped_ls = label.lstrip()
        indent = len(label) - len(stripped_ls)
        is_sub = stripped_ls.startswith("-")

        if not is_sub:
            # Main group row
            main_core = canonical_main(label)
            current_main = main_core
            # Record main-group totals per BU
            for c, bu_name in bu_amount_cols:
                v = ws.cell(row=row_idx, column=c).value
                if isinstance(v, (int, float)) and v != 0:
                    main_totals[(main_core, bu_name)] = main_totals.get((main_core, bu_name), 0.0) + float(v)
            continue

        sub_core = canonical_sub(label)
        level = 2 if indent >= 3 else 1  # '     - xxx' has 5 leading spaces

        for c, bu_name in bu_amount_cols:
            v = ws.cell(row=row_idx, column=c).value
            if not isinstance(v, (int, float)) or v == 0:
                continue
            entries.append((current_main, sub_core, bu_name, float(v), level))

    # Pass 2: drop level-1 row if a level-2 exists for same (main, sub_core)
    has_level2 = {(m, s) for m, s, _, _, lvl in entries if lvl == 2}
    data: Dict[Tuple[str, str, str], float] = {}
    for m, s, bu, v, lvl in entries:
        if lvl == 1 and (m, s) in has_level2:
            continue
        key = (m, s, bu)
        data[key] = data.get(key, 0.0) + v

    # Merge in main-group totals with sub=''
    for (m, bu), v in main_totals.items():
        data[(m, "", bu)] = data.get((m, "", bu), 0.0) + v

    return [n for _, n in bu_amount_cols], data


# ---------- Comparison ----------

# BU ที่ skip การเทียบค่า — text-report ยังไม่ได้ net สุทธิ ค่าจะไม่ตรงกับ Excel ที่ net แล้ว
_SKIP_BU_MATCH = {"8.รายได้อื่น/ค่าใช้จ่ายอื่น"}


def compare(
    text_data: Dict[Tuple[str, str, str], float],
    excel_data: Dict[Tuple[str, str, str], float],
    tolerance: float,
) -> Tuple[List, List]:
    """
    Returns:
        missing: keys where text has non-zero value but Excel is absent/zero
        mismatch: keys where both sides have values but differ beyond tolerance
    """
    missing: List[Tuple[Tuple[str, str, str], float]] = []
    mismatch: List[Tuple[Tuple[str, str, str], float, float, float]] = []

    for k, tv in text_data.items():
        if abs(tv) < 1e-9:
            continue
        if k[2] in _SKIP_BU_MATCH:
            continue
        ev = excel_data.get(k, 0.0)
        diff = tv - ev
        if k not in excel_data or abs(ev) < 1e-9:
            missing.append((k, tv))
        elif abs(diff) > tolerance:
            mismatch.append((k, tv, ev, diff))

    return missing, mismatch


# ---------- File discovery ----------

def find_text_report(data_dir: Path, kind: str, month: str) -> Optional[Path]:
    """
    kind: 'ต้นทุน' (COSTTYPE) or 'บัญชี' (GLGROUP)
    month: 'YYYYMM' e.g. '202603'
    filename code: 'MMYY' e.g. '0326' (month=03, year=26)
    If print-date suffix differs, take latest lex-sorted match.
    """
    month_code = month[4:6] + month[2:4]
    patterns = [f"001_{kind}_BU_{month_code}*.csv" if kind == "ต้นทุน"
                else f"002_{kind}_BU_{month_code}*.csv",
                f"*{kind}_BU_{month_code}*.csv"]
    for pat in patterns:
        matches = sorted(data_dir.glob(pat))
        if matches:
            return matches[-1]
    return None


# ---------- Report writer ----------

def run_one(out, label: str, text_path: Optional[Path], excel_path: Path, tolerance: float):
    out.write(f"\n\n=== Dimension: {label} ===\n")
    if not text_path or not text_path.exists():
        out.write(f"  SKIP: text-report not found\n")
        return 0, 0
    if not excel_path.exists():
        out.write(f"  SKIP: Excel not found at {excel_path}\n")
        return 0, 0
    out.write(f"  Text-report: {text_path.name}\n  Excel:       {excel_path.name}\n")
    try:
        _, text_data = parse_text_report(text_path)
        _, excel_data = parse_excel_bu_only(excel_path)
    except Exception as e:
        out.write(f"  ERROR parsing: {e}\n")
        return 0, 0

    missing, mismatch = compare(text_data, excel_data, tolerance)
    missing_subs = [(k, v) for k, v in missing if k[1]]       # sub-item level
    missing_mains = [(k, v) for k, v in missing if not k[1]]  # main group totals

    out.write(f"\n  [MISSING sub-items] row มีข้อมูลแต่ Excel ไม่มีบรรทัด — น่าจะเป็น mapping bug ({len(missing_subs)}):\n")
    if not missing_subs:
        out.write("    (none)\n")
    else:
        for (m, s, bu), v in sorted(missing_subs):
            out.write(f"    {m:<45s} | {s:<50s} | {bu:<40s} | {v:>18,.2f}\n")

    out.write(f"\n  [MISSING main-group totals] Excel ไม่มีค่าในคอลัมน์ BU ระดับ group (อาจเป็น calculated row ที่ไม่ได้ break down per-BU) ({len(missing_mains)}):\n")
    if not missing_mains:
        out.write("    (none)\n")
    else:
        for (m, _, bu), v in sorted(missing_mains):
            out.write(f"    {m:<50s} | {bu:<40s} | {v:>18,.2f}\n")

    out.write(f"\n  [VALUE MISMATCH] ค่าต่างกันเกิน tolerance ({len(mismatch)}):\n")
    if not mismatch:
        out.write("    (none)\n")
    else:
        for (m, s, bu), tv, ev, diff in sorted(mismatch, key=lambda x: -abs(x[3])):
            desc = s if s else "(main total)"
            out.write(f"    {m:<40s} | {desc:<40s} | {bu:<35s} | txt={tv:>15,.2f} xls={ev:>15,.2f} diff={diff:>+12,.2f}\n")

    return len(missing_subs), len(missing_mains), len(mismatch)


def main():
    ap = argparse.ArgumentParser(description="Verify report completeness vs text-report CSV")
    ap.add_argument("--month", required=True, help="YYYYMM (e.g. 202603)")
    ap.add_argument("--tolerance", type=float, default=TOLERANCE_DEFAULT,
                    help=f"absolute tolerance in บาท (default {TOLERANCE_DEFAULT})")
    ap.add_argument("--period", default="YTD", choices=["YTD", "MTH"],
                    help="which Excel period to verify (default YTD — matches text-report)")
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent
    proj_dir = script_dir.parent
    text_dir = proj_dir / "data" / "text-report"
    excel_dir = proj_dir / "output"

    text_cost = find_text_report(text_dir, "ต้นทุน", args.month)
    text_gl = find_text_report(text_dir, "บัญชี", args.month)
    excel_cost = excel_dir / f"PL_COSTTYPE_{args.period}_BU_ONLY_{args.month}.xlsx"
    excel_gl = excel_dir / f"PL_GLGROUP_{args.period}_BU_ONLY_{args.month}.xlsx"

    report_path = script_dir / f"verification_report_{args.month}.txt"

    with open(report_path, "w", encoding="utf-8") as out:
        out.write(f"=== Report Completeness Verification ===\n")
        out.write(f"Month: {args.month} | Period: {args.period} | Tolerance: ±{args.tolerance:.2f}\n")

        total_sub = 0
        total_main = 0
        total_mismatch = 0

        for label, tp, ep in [
            ("COSTTYPE (มิติประเภทต้นทุน)", text_cost, excel_cost),
            ("GLGROUP (มิติหมวดบัญชี)", text_gl, excel_gl),
        ]:
            s, m, mm = run_one(out, label, tp, ep, args.tolerance)
            total_sub += s
            total_main += m
            total_mismatch += mm

        out.write(f"\n\n=== SUMMARY ===\n")
        out.write(f"Missing sub-items (mapping bugs): {total_sub}\n")
        out.write(f"Missing main totals (calc/design):{total_main}\n")
        out.write(f"Value mismatches:                  {total_mismatch}\n")

    print(f"Verification report saved: {report_path}")
    print(f"Summary: {total_sub} sub-item gaps, {total_main} main-total gaps, {total_mismatch} value mismatches")

    return 1 if total_sub else 0


if __name__ == "__main__":
    sys.exit(main())
