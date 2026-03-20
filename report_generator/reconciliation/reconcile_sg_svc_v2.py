"""
NT P&L Reconciliation - Service Group & Service (Product) Level v2
===================================================================
Fixed version: proper GROUP matching and dynamic product key row detection
"""

import re
import pandas as pd
import openpyxl
import argparse
from openpyxl.utils import get_column_letter
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from datetime import datetime

# ==========================================
# PATH CONFIGURATION
# ==========================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / 'data'
REPORT_DIR = SCRIPT_DIR / 'report'
OUTPUT_DIR = SCRIPT_DIR / 'output'

TOLERANCE = 0.001  # ยอมรับผลต่างไม่เกิน 0.001 บาท (floating point)


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='NT P&L Reconciliation - Service Group & Product level')
    parser.add_argument('--date', type=str, required=True,
                        help='วันที่ข้อมูล YYYYMMDD (เช่น 20260131)')
    parser.add_argument('--company', type=str, default='NT',
                        help='รหัสบริษัท (default: NT)')
    parser.add_argument('--excel-mth', type=str, default=None,
                        help='ชื่อไฟล์ Excel report MTH (ถ้าไม่ระบุจะใช้ Report_{company}_{YYYYMM}.xlsx)')
    parser.add_argument('--excel-ytd', type=str, default=None,
                        help='ชื่อไฟล์ Excel report YTD (ถ้าไม่ระบุจะใช้ Report_{company}_YTD_{YYYYMM}.xlsx)')
    return parser.parse_args()


def build_file_config(args):
    """Build file paths from --date argument"""
    try:
        date_obj = datetime.strptime(args.date, '%Y%m%d')
    except ValueError:
        print("❌ รูปแบบวันที่ไม่ถูกต้อง: {} (ใช้ YYYYMMDD)".format(args.date))
        return None

    month = date_obj.strftime('%Y%m')
    company = args.company

    csv_filenames = {
        'COSTTYPE_MTH': 'TRN_PL_COSTTYPE_{}_MTH_TABLE_{}.csv'.format(company, args.date),
        'COSTTYPE_YTD': 'TRN_PL_COSTTYPE_{}_YTD_TABLE_{}.csv'.format(company, args.date),
        'GLGROUP_MTH':  'TRN_PL_GLGROUP_{}_MTH_TABLE_{}.csv'.format(company, args.date),
        'GLGROUP_YTD':  'TRN_PL_GLGROUP_{}_YTD_TABLE_{}.csv'.format(company, args.date),
    }

    excel_mth = args.excel_mth or 'Report_{}_{}.xlsx'.format(company, month)
    excel_ytd = args.excel_ytd or 'Report_{}_YTD_{}.xlsx'.format(company, month)
    excel_filenames = {'MTH': excel_mth, 'YTD': excel_ytd}

    output_filename = 'reconciliation_sg_svc_v2_{}.xlsx'.format(month)

    return {
        'date_obj': date_obj,
        'month': month,
        'company': company,
        'csv_filenames': csv_filenames,
        'excel_filenames': excel_filenames,
        'output_filename': output_filename,
    }


def validate_period(csv_data, excel_files, expected_month):
    """Validate that CSV and Excel are for the same period"""
    warnings = []

    for key, df in csv_data.items():
        if 'TIME_KEY' in df.columns:
            time_keys = df['TIME_KEY'].astype(str).unique()
            for tk in time_keys:
                if not tk.startswith(expected_month):
                    warnings.append("CSV {}: TIME_KEY={} ไม่ตรงกับงวด {}".format(
                        key, tk, expected_month))
                    break

    for period, path in excel_files.items():
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row_idx in range(1, 10):
                # Scan first few columns for period header
                val = None
                for col_idx in range(1, 6):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val and 'ประจำเดือน' in str(cell_val):
                        val = cell_val
                        break
                if val and 'ประจำเดือน' in str(val):
                    header_text = str(val)
                    if expected_month not in header_text:
                        warnings.append("Excel {} header: '{}' - ตรวจสอบว่าตรงกับงวด {} หรือไม่".format(
                            period, header_text.strip(), expected_month))
                    break
            wb.close()
        except Exception:
            pass

    return warnings

@dataclass
class CheckResult:
    category: str
    check_name: str
    source_label: str
    source_value: float
    target_label: str
    target_value: float
    tolerance: float = TOLERANCE

    @property
    def diff(self):
        return self.source_value - self.target_value
    @property
    def passed(self):
        return abs(self.diff) <= self.tolerance
    @property
    def status(self):
        return "PASS" if self.passed else "FAIL"


def read_csv_auto_encoding(filepath):
    for enc in ['utf-8', 'cp874', 'tis-620', 'latin-1']:
        try:
            df = pd.read_csv(filepath, encoding=enc)
            if 'TIME_KEY' in df.columns:
                return df
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("Cannot read {}".format(filepath))


# ==========================================
# Row/Column definitions with GROUP mapping
# ==========================================

# Each entry: (excel_keywords, csv_group_pattern, display_label)
# csv_group_pattern is used to filter CSV GROUP column

COSTTYPE_CHECKS = [
    (['1', 'รวมรายได้'], '01.รายได้', 'รายได้'),
    (['2.', 'ต้นทุนบริการ'], '02.ต้นทุนบริการ', 'ต้นทุนบริการ'),
    (['3.', 'กำไร', 'ขั้นต้น'], '03.กำไร', 'กำไรขั้นต้น'),
    (['5.', 'กำไร', 'หลังหัก'], '05.กำไร', 'กำไรหลังหักค่าใช้จ่ายขาย'),
    (['8.', 'กำไร', 'จัดหาเงิน'], '08.กำไร', 'กำไรก่อนต้นทุนจัดหาเงิน'),
    (['12.', 'กำไร', 'EBT'], '12.กำไร', 'กำไรก่อนภาษี'),
    (['14.', 'กำไร', 'สุทธิ'], '14.กำไร', 'กำไรสุทธิ'),
]

GLGROUP_CHECKS = [
    (['1', 'รวมรายได้'], '01.รายได้', 'รายได้'),
    (['2', 'รวมค่าใช้จ่าย'], '02.ค่าใช้จ่าย', 'ค่าใช้จ่าย'),
    (['3.', 'กำไร', 'EBT'], '03.กำไร', 'กำไรก่อนภาษี'),
    (['5.', 'กำไร', 'สุทธิ'], '05.กำไร', 'กำไรสุทธิ'),
]


@dataclass
class SheetLayout:
    """Dynamic layout detection result for an Excel sheet"""
    label_col: int       # Column containing row labels (e.g., "รายละเอียด")
    sg_row: int          # Row containing service group names
    data_start: int      # First data row (e.g., "1.รายได้")
    header_row: int      # Row containing "รายละเอียด"


def detect_sheet_layout(ws):
    """Detect sheet layout dynamically by finding 'รายละเอียด' cell

    Returns SheetLayout with all key positions.
    """
    # Scan first 15 rows × 10 cols for "รายละเอียด"
    label_col = 2     # default
    header_row = 6    # default
    for row_idx in range(1, 16):
        for col_idx in range(1, 11):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and 'รายละเอียด' in str(val):
                label_col = col_idx
                header_row = row_idx
                break
        else:
            continue
        break

    sg_row = header_row + 1

    # Find data start: first row with "รายได้" in label_col
    data_start = header_row + 4  # default
    for row_idx in range(header_row + 1, header_row + 10):
        val = ws.cell(row=row_idx, column=label_col).value
        if val and 'รายได้' in str(val):
            data_start = row_idx
            break

    return SheetLayout(
        label_col=label_col,
        sg_row=sg_row,
        data_start=data_start,
        header_row=header_row,
    )


def find_data_start_row(ws, layout=None):
    if layout:
        return layout.data_start
    # Legacy fallback
    for row_idx in range(1, 15):
        val = ws.cell(row=row_idx, column=2).value
        if val and str(val).strip().startswith('1') and 'รายได้' in str(val):
            return row_idx
    return 10


def find_row_index(ws, data_start, keywords, label_col=2):
    for row_idx in range(data_start, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=label_col).value
        if val is None:
            continue
        label = str(val).strip()
        if all(kw in label for kw in keywords):
            return row_idx
    return None


# ==========================================
# Service Group Mapping
# ==========================================

def _extract_sg_number(text):
    """Extract service group number prefix e.g. '4.5' from '4.5 กลุ่มบริการ SATELLITE'
    or '4.5.1' from '4.5.1 กลุ่มบริการ SATELLITE-NT'"""
    import re
    m = re.match(r'(\d+(?:\.\d+)*)', text)
    return m.group(1) if m else None


def build_sg_column_map(ws, layout=None):
    """Build SERVICE_GROUP name → column index from SG row

    รวม columns (เช่น "รวม 4.5 SATELLITE") จะถูกเก็บแยก
    เพื่อใช้เป็น fallback เมื่อไม่มี column ย่อยที่ตรง
    """
    sg_row = layout.sg_row if layout else 7
    header_row = layout.header_row if layout else 6

    col_map = {}       # columns ปกติ
    summary_map = {}   # columns ที่ขึ้นต้นด้วย "รวม"

    # Scan both header row and SG row for column names
    for scan_row in [sg_row, header_row]:
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=scan_row, column=col_idx).value
            if val is None:
                continue
            val_str = str(val).strip()
            val_upper = val_str.upper()
            # Skip label column header
            if 'รายละเอียด' in val_str or 'รวมทั้งสิ้น' in val_str:
                continue
            if val_str.startswith('รวม'):
                sg_num = _extract_sg_number(val_str[len('รวม'):].strip())
                if sg_num:
                    summary_map[sg_num] = col_idx
                summary_map[val_upper] = col_idx
            else:
                if val_upper not in col_map:  # SG row takes priority
                    col_map[val_upper] = col_idx
    return col_map, summary_map


def match_sg(csv_sg, sg_col_map, summary_map=None):
    """Find matching Excel column for a CSV SERVICE_GROUP

    ลำดับการ match:
    1. Exact match (CSV uppercase == Excel uppercase)
    2. Exact number match (เช่น CSV "4.5" ตรงกับ Excel "4.5" ไม่ใช่ "4.5.1")
    3. Summary column fallback (ถ้ามี "รวม 4.5 ..." ให้ใช้)
    """
    if summary_map is None:
        summary_map = {}

    csv_norm = csv_sg.strip().upper()
    csv_num = _extract_sg_number(csv_norm)

    # 1. Exact match
    if csv_norm in sg_col_map:
        return sg_col_map[csv_norm]

    # 2. Exact number match (e.g. "4.5" must match "4.5" not "4.5.1")
    if csv_num:
        for excel_sg, col_idx in sg_col_map.items():
            excel_num = _extract_sg_number(excel_sg)
            if excel_num == csv_num:
                return col_idx

    # 3. Summary column fallback (เช่น CSV "4.5 SATELLITE" → Excel "รวม 4.5 SATELLITE")
    if csv_num and csv_num in summary_map:
        return summary_map[csv_num]

    return None


# ==========================================
# Product Key Mapping
# ==========================================

def build_product_column_map(ws, layout=None):
    """Build PRODUCT_KEY → column index, auto-detecting which row has keys"""
    sg_row = layout.sg_row if layout else 7
    # Scan rows after SG row for numeric product keys
    for pk_row in range(sg_row + 1, sg_row + 5):
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=pk_row, column=col_idx).value
            if val is None:
                continue
            val_str = str(val).strip()
            try:
                int(val_str)
                col_map[val_str] = col_idx
            except ValueError:
                continue
        if len(col_map) > 10:  # Found product keys
            return col_map, pk_row

    return {}, sg_row + 1


# ==========================================
# Reconcile Service Group Level
# ==========================================

def reconcile_service_group(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    category = "CSV vs Excel by ServiceGroup ({}) - {}".format(period_label, sheet_name)
    checks = COSTTYPE_CHECKS if csv_type == 'COSTTYPE' else GLGROUP_CHECKS

    if layout is None:
        layout = detect_sheet_layout(ws)
    data_start = layout.data_start
    sg_col_map, sg_summary_map = build_sg_column_map(ws, layout)

    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)

    unique_sgs = csv_df['SERVICE_GROUP'].unique()
    total_checks = 0

    for csv_sg in sorted(unique_sgs):
        excel_col = match_sg(csv_sg, sg_col_map, sg_summary_map)
        if excel_col is None:
            continue

        for excel_kw, csv_group_prefix, desc in checks:
            # Find Excel row
            excel_row = find_row_index(ws, data_start, excel_kw, layout.label_col)
            if excel_row is None:
                continue

            # Get Excel value
            excel_val = ws.cell(row=excel_row, column=excel_col).value
            if excel_val is None:
                excel_val = 0.0
            try:
                excel_val = float(excel_val)
            except (ValueError, TypeError):
                continue

            # Skip if Excel is 0 and this is a row that may not have SG breakdown
            # (e.g., กำไรสุทธิ is often only at BU total level)
            if excel_val == 0.0 and 'สุทธิ' in desc:
                # Check if ALL service group columns are 0 for this row
                all_zero = True
                for _, c in sg_col_map.items():
                    v = ws.cell(row=excel_row, column=c).value
                    if v is not None and v != 0:
                        all_zero = False
                        break
                if all_zero:
                    continue  # Skip - this row doesn't have SG breakdown

            # Get CSV value (filter by GROUP prefix)
            csv_filtered = csv_df[
                (csv_df['SERVICE_GROUP'] == csv_sg) &
                (csv_df['GROUP'].str.startswith(csv_group_prefix))
            ]
            csv_val = csv_filtered['VALUE'].sum() if len(csv_filtered) > 0 else 0.0

            sg_label = csv_sg[:45]
            results.append(CheckResult(
                category=category,
                check_name="{} - {}".format(desc, sg_label),
                source_label="CSV",
                source_value=csv_val,
                target_label="Excel col {}".format(get_column_letter(excel_col)),
                target_value=excel_val,
            ))
            total_checks += 1

    return total_checks


# ==========================================
# Reconcile Product Level
# ==========================================

def reconcile_product(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    category = "CSV vs Excel by Product ({}) - {}".format(period_label, sheet_name)
    checks = COSTTYPE_CHECKS if csv_type == 'COSTTYPE' else GLGROUP_CHECKS

    if layout is None:
        layout = detect_sheet_layout(ws)
    data_start = layout.data_start
    pk_col_map, pk_row = build_product_column_map(ws, layout)

    if len(pk_col_map) == 0:
        print("    Warning: No product keys found in {} {}".format(period_label, sheet_name))
        return 0

    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)
    csv_df['PRODUCT_KEY'] = csv_df['PRODUCT_KEY'].astype(str).str.strip()

    # Get product names for labeling
    pk_names = csv_df.drop_duplicates('PRODUCT_KEY').set_index('PRODUCT_KEY')['PRODUCT_NAME'].to_dict()

    unique_pks = csv_df['PRODUCT_KEY'].unique()
    total_checks = 0
    unmatched = []

    for csv_pk in sorted(unique_pks):
        excel_col = pk_col_map.get(csv_pk)
        if excel_col is None:
            unmatched.append(csv_pk)
            continue

        for excel_kw, csv_group_prefix, desc in checks:
            # Find Excel row
            excel_row = find_row_index(ws, data_start, excel_kw, layout.label_col)
            if excel_row is None:
                continue

            # Get Excel value
            excel_val = ws.cell(row=excel_row, column=excel_col).value
            if excel_val is None:
                excel_val = 0.0
            try:
                excel_val = float(excel_val)
            except (ValueError, TypeError):
                continue

            # Skip rows where product level isn't populated
            if excel_val == 0.0:
                # Check if this is a row that might not have product breakdown
                sample_vals = [ws.cell(row=excel_row, column=c).value for c in list(pk_col_map.values())[:10]]
                if all(v is None or v == 0 for v in sample_vals):
                    continue

            # Get CSV value
            csv_filtered = csv_df[
                (csv_df['PRODUCT_KEY'] == csv_pk) &
                (csv_df['GROUP'].str.startswith(csv_group_prefix))
            ]
            csv_val = csv_filtered['VALUE'].sum() if len(csv_filtered) > 0 else 0.0

            product_name = pk_names.get(csv_pk, csv_pk)
            if len(product_name) > 25:
                product_name = product_name[:25] + '..'

            results.append(CheckResult(
                category=category,
                check_name="{} - {} ({})".format(desc, product_name, csv_pk),
                source_label="CSV",
                source_value=csv_val,
                target_label="Excel",
                target_value=excel_val,
            ))
            total_checks += 1

    if unmatched:
        print("    Warning: {} unmatched PKs".format(len(unmatched)))

    return total_checks


# ==========================================
# Reconcile EBITDA (calculated row)
# ==========================================

# --- COSTTYPE GROUP structure ---
# 01.รายได้, 02.ต้นทุนบริการ, 04.ค่าใช้จ่ายขาย, 06.ค่าใช้จ่ายบริหาร,
# 07.ต้นทุนทางการเงิน-ดำเนินงาน, 08.EBIT, 09.ผลตอบแทนฯ, 10.ค่าใช้จ่ายอื่น, 12.EBT
#
# --- GLGROUP GROUP structure ---
# 01.รายได้ (รวมผลตอบแทนฯแล้ว), 02.ค่าใช้จ่าย (รวมทุกอย่าง), 03.EBT, 05.กำไรสุทธิ
# SUB_GROUP 19.ต้นทุนทางการเงิน-ดำเนินงาน
# SUB_GROUP 12.ค่าเสื่อมราคาฯ, 13.ค่าตัดจำหน่ายฯ

EBITDA_KEYWORDS_COSTTYPE = ['EBITDA']
EBITDA_KEYWORDS_GLGROUP = ['EBITDA']

EBIT_KEYWORDS_COSTTYPE = ['8.', 'กำไร', 'จัดหาเงิน']
# GLGROUP: ใช้ row "ค่าใช้จ่ายรวม (ไม่รวมต้นทุนทางการเงิน)" เป็น reference
# แต่ EBIT = Revenue - row นั้น ต้องคำนวณใน reconcile function
EBIT_KEYWORDS_GLGROUP = ['ค่าใช้จ่ายรวม', 'ไม่รวมต้นทุนทางการเงิน']

# COSTTYPE config
COSTTYPE_EBIT_REVENUE = ['01.']
COSTTYPE_EBIT_EXPENSE = ['02.', '04.', '06.', '07.']
COSTTYPE_EBITDA_REVENUE = ['01.', '09.']
COSTTYPE_EBITDA_EXPENSE = ['02.', '04.', '06.', '10.']
COSTTYPE_EBITDA_DEP_GROUPS = ['02.', '04.', '06.']
COSTTYPE_EBITDA_DEP_SUBGROUPS = [
    '12.ค่าเสื่อมราคาและรายจ่ายตัดบัญชีสินทรัพย์',
    '13.ค่าตัดจำหน่ายสิทธิการใช้ตามสัญญาเช่า',
]

# GLGROUP config
GLGROUP_FINANCE_OPERATING_SUBGROUP = '19.ต้นทุนทางการเงิน-ด้านการดำเนินงาน'
GLGROUP_DEP_SUBGROUPS = [
    '12.ค่าเสื่อมราคาและรายจ่ายตัดบัญชีสินทรัพย์',
    '13.ค่าตัดจำหน่ายสิทธิการใช้ตามสัญญาเช่า',
]


def _filter_df(csv_df, sg_filter=None, pk_filter=None):
    """Helper: filter DataFrame by SERVICE_GROUP and/or PRODUCT_KEY"""
    df = csv_df.copy()
    if sg_filter:
        df = df[df['SERVICE_GROUP'] == sg_filter]
    if pk_filter:
        df = df[df['PRODUCT_KEY'].astype(str).str.strip() == pk_filter]
    return df


def calc_ebit_from_csv(csv_df, csv_type, sg_filter=None, pk_filter=None):
    """Calculate EBIT from CSV data

    COSTTYPE: GROUP 01 - GROUP 02 - GROUP 04 - GROUP 06 - GROUP 07
    GLGROUP:  GROUP 01 - GROUP 02 + SUB_GROUP 19 (ต้นทุนทางการเงิน-ดำเนินงาน)
              = Revenue - Expense + FinanceOperating (add back finance from expense)
    """
    df = _filter_df(csv_df, sg_filter, pk_filter)

    if csv_type == 'COSTTYPE':
        revenue = df[df['GROUP'].apply(
            lambda g: any(g.startswith(p) for p in COSTTYPE_EBIT_REVENUE)
        )]['VALUE'].sum()
        expense = df[df['GROUP'].apply(
            lambda g: any(g.startswith(p) for p in COSTTYPE_EBIT_EXPENSE)
        )]['VALUE'].sum()
        return revenue - expense
    else:
        # GLGROUP: EBIT = Revenue(01) - Expense(02) + FinanceOperating(SUB_GROUP 19)
        revenue = df[df['GROUP'].str.startswith('01.')]['VALUE'].sum()
        expense = df[df['GROUP'].str.startswith('02.')]['VALUE'].sum()
        finance_op = df[
            df['GROUP'].str.startswith('02.') &
            (df['SUB_GROUP'] == GLGROUP_FINANCE_OPERATING_SUBGROUP)
        ]['VALUE'].sum()
        return revenue - expense + finance_op


def calc_ebitda_from_csv(csv_df, csv_type, sg_filter=None, pk_filter=None):
    """Calculate EBITDA from CSV data

    COSTTYPE: (GROUP 01 + 09) - (GROUP 02 + 04 + 06 + 10) + Depreciation(SUB_GROUP 12,13)
    GLGROUP:  EBIT + Depreciation(SUB_GROUP 12,13)
    """
    df = _filter_df(csv_df, sg_filter, pk_filter)

    if csv_type == 'COSTTYPE':
        revenue = df[df['GROUP'].apply(
            lambda g: any(g.startswith(p) for p in COSTTYPE_EBITDA_REVENUE)
        )]['VALUE'].sum()
        expense = df[df['GROUP'].apply(
            lambda g: any(g.startswith(p) for p in COSTTYPE_EBITDA_EXPENSE)
        )]['VALUE'].sum()
        depreciation = df[
            df['GROUP'].apply(lambda g: any(g.startswith(p) for p in COSTTYPE_EBITDA_DEP_GROUPS)) &
            df['SUB_GROUP'].isin(COSTTYPE_EBITDA_DEP_SUBGROUPS)
        ]['VALUE'].sum()
        return revenue - expense + depreciation
    else:
        # GLGROUP: EBITDA = EBIT + Depreciation(SUB_GROUP 12,13)
        ebit = calc_ebit_from_csv(csv_df, csv_type, sg_filter, pk_filter)
        depreciation = df[
            df['GROUP'].str.startswith('02.') &
            df['SUB_GROUP'].isin(GLGROUP_DEP_SUBGROUPS)
        ]['VALUE'].sum()
        return ebit + depreciation


def reconcile_ebitda_sg(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    """Reconcile EBITDA at Service Group level"""
    category = "EBITDA CSV vs Excel by ServiceGroup ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)
    ebitda_kw = EBITDA_KEYWORDS_COSTTYPE if csv_type == 'COSTTYPE' else EBITDA_KEYWORDS_GLGROUP
    ebitda_row = find_row_index(ws, layout.data_start, ebitda_kw, layout.label_col)
    if ebitda_row is None:
        print("    Warning: EBITDA row not found in {}".format(sheet_name))
        return 0

    sg_col_map, sg_summary_map = build_sg_column_map(ws, layout)
    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)

    unique_sgs = csv_df['SERVICE_GROUP'].unique()
    total_checks = 0

    for csv_sg in sorted(unique_sgs):
        excel_col = match_sg(csv_sg, sg_col_map, sg_summary_map)
        if excel_col is None:
            continue

        excel_val = ws.cell(row=ebitda_row, column=excel_col).value
        if excel_val is None:
            excel_val = 0.0
        try:
            excel_val = float(excel_val)
        except (ValueError, TypeError):
            continue

        csv_ebitda = calc_ebitda_from_csv(csv_df, csv_type, sg_filter=csv_sg)
        sg_label = csv_sg[:45]
        results.append(CheckResult(
            category=category,
            check_name="EBITDA - {}".format(sg_label),
            source_label="CSV(calc)",
            source_value=csv_ebitda,
            target_label="Excel col {}".format(get_column_letter(excel_col)),
            target_value=excel_val,
        ))
        total_checks += 1

    return total_checks


def reconcile_ebitda_product(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    """Reconcile EBITDA at Product level"""
    category = "EBITDA CSV vs Excel by Product ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)
    ebitda_kw = EBITDA_KEYWORDS_COSTTYPE if csv_type == 'COSTTYPE' else EBITDA_KEYWORDS_GLGROUP
    ebitda_row = find_row_index(ws, layout.data_start, ebitda_kw, layout.label_col)
    if ebitda_row is None:
        return 0

    pk_col_map, _ = build_product_column_map(ws, layout)
    if len(pk_col_map) == 0:
        return 0

    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)
    csv_df['PRODUCT_KEY'] = csv_df['PRODUCT_KEY'].astype(str).str.strip()
    pk_names = csv_df.drop_duplicates('PRODUCT_KEY').set_index('PRODUCT_KEY')['PRODUCT_NAME'].to_dict()

    total_checks = 0
    for csv_pk in sorted(csv_df['PRODUCT_KEY'].unique()):
        excel_col = pk_col_map.get(csv_pk)
        if excel_col is None:
            continue

        excel_val = ws.cell(row=ebitda_row, column=excel_col).value
        if excel_val is None:
            excel_val = 0.0
        try:
            excel_val = float(excel_val)
        except (ValueError, TypeError):
            continue

        csv_ebitda = calc_ebitda_from_csv(csv_df, csv_type, pk_filter=csv_pk)
        product_name = pk_names.get(csv_pk, csv_pk)
        if len(product_name) > 25:
            product_name = product_name[:25] + '..'

        results.append(CheckResult(
            category=category,
            check_name="EBITDA - {} ({})".format(product_name, csv_pk),
            source_label="CSV(calc)",
            source_value=csv_ebitda,
            target_label="Excel",
            target_value=excel_val,
        ))
        total_checks += 1

    return total_checks


# ==========================================
# Reconcile EBIT (calculated row)
# ==========================================

def _read_excel_cell_float(ws, row, col):
    """Read Excel cell and return as float (0.0 if None)"""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def reconcile_ebit_sg(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    """Reconcile EBIT at Service Group level"""
    category = "EBIT CSV vs Excel by ServiceGroup ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)
    lc = layout.label_col
    ds = layout.data_start

    if csv_type == 'COSTTYPE':
        ebit_row = find_row_index(ws, ds, EBIT_KEYWORDS_COSTTYPE, lc)
        if ebit_row is None:
            print("    Warning: EBIT row not found in {}".format(sheet_name))
            return 0
        revenue_row = None
        expense_no_fin_row = None
    else:
        ebit_row = None
        revenue_row = find_row_index(ws, ds, ['รายได้รวม'], lc)
        expense_no_fin_row = find_row_index(ws, ds, EBIT_KEYWORDS_GLGROUP, lc)
        if revenue_row is None or expense_no_fin_row is None:
            print("    Warning: Revenue/Expense row not found for EBIT calc in {}".format(sheet_name))
            return 0

    sg_col_map, sg_summary_map = build_sg_column_map(ws, layout)
    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)

    unique_sgs = csv_df['SERVICE_GROUP'].unique()
    total_checks = 0

    for csv_sg in sorted(unique_sgs):
        excel_col = match_sg(csv_sg, sg_col_map, sg_summary_map)
        if excel_col is None:
            continue

        if csv_type == 'COSTTYPE':
            excel_val = _read_excel_cell_float(ws, ebit_row, excel_col)
        else:
            rev = _read_excel_cell_float(ws, revenue_row, excel_col)
            exp = _read_excel_cell_float(ws, expense_no_fin_row, excel_col)
            if rev is None or exp is None:
                continue
            excel_val = rev - exp

        if excel_val is None:
            continue

        csv_ebit = calc_ebit_from_csv(csv_df, csv_type, sg_filter=csv_sg)
        sg_label = csv_sg[:45]
        results.append(CheckResult(
            category=category,
            check_name="EBIT - {}".format(sg_label),
            source_label="CSV(calc)",
            source_value=csv_ebit,
            target_label="Excel col {}".format(get_column_letter(excel_col)),
            target_value=excel_val,
        ))
        total_checks += 1

    return total_checks


def reconcile_ebit_product(csv_df, ws, csv_type, period_label, sheet_name, results, layout=None):
    """Reconcile EBIT at Product level"""
    category = "EBIT CSV vs Excel by Product ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)
    lc = layout.label_col
    ds = layout.data_start

    if csv_type == 'COSTTYPE':
        ebit_row = find_row_index(ws, ds, EBIT_KEYWORDS_COSTTYPE, lc)
        if ebit_row is None:
            return 0
        revenue_row = None
        expense_no_fin_row = None
    else:
        ebit_row = None
        revenue_row = find_row_index(ws, ds, ['รายได้รวม'], lc)
        expense_no_fin_row = find_row_index(ws, ds, EBIT_KEYWORDS_GLGROUP, lc)
        if revenue_row is None or expense_no_fin_row is None:
            return 0

    pk_col_map, _ = build_product_column_map(ws, layout)
    if len(pk_col_map) == 0:
        return 0

    csv_df = csv_df.copy()
    csv_df['VALUE'] = csv_df['VALUE'].astype(float)
    csv_df['PRODUCT_KEY'] = csv_df['PRODUCT_KEY'].astype(str).str.strip()
    pk_names = csv_df.drop_duplicates('PRODUCT_KEY').set_index('PRODUCT_KEY')['PRODUCT_NAME'].to_dict()

    total_checks = 0
    for csv_pk in sorted(csv_df['PRODUCT_KEY'].unique()):
        excel_col = pk_col_map.get(csv_pk)
        if excel_col is None:
            continue

        if csv_type == 'COSTTYPE':
            excel_val = _read_excel_cell_float(ws, ebit_row, excel_col)
        else:
            rev = _read_excel_cell_float(ws, revenue_row, excel_col)
            exp = _read_excel_cell_float(ws, expense_no_fin_row, excel_col)
            if rev is None or exp is None:
                continue
            excel_val = rev - exp

        if excel_val is None:
            continue

        csv_ebit = calc_ebit_from_csv(csv_df, csv_type, pk_filter=csv_pk)
        product_name = pk_names.get(csv_pk, csv_pk)
        if len(product_name) > 25:
            product_name = product_name[:25] + '..'

        results.append(CheckResult(
            category=category,
            check_name="EBIT - {} ({})".format(product_name, csv_pk),
            source_label="CSV(calc)",
            source_value=csv_ebit,
            target_label="Excel",
            target_value=excel_val,
        ))
        total_checks += 1

    return total_checks


# ==========================================
# Cross-Column Consistency Check
# ==========================================

def find_summary_columns(ws, layout=None):
    """Find summary columns (รวม X.X ...) and their child columns

    In กลุ่มบริการ sheets: "รวม 4.5" → children = "4.5.1 ...", "4.5.2 ..."
    In บริการ sheets: "รวม 4.5" → children = "รวม 4.5.1 ...", "รวม 4.5.2 ..."
                      "รวม 4.5.1" → children = individual product columns "4.5.1 ..."
    """
    sg_row = layout.sg_row if layout else 7
    header_row = layout.header_row if layout else 6
    summaries = []

    all_cols = {}
    for scan_row in [sg_row, header_row]:
        for col_idx in range(1, ws.max_column + 1):
            if col_idx in all_cols:
                continue
            val = ws.cell(row=scan_row, column=col_idx).value
            if val is None:
                continue
            val_str = str(val).strip()
            if 'รายละเอียด' in val_str or 'รวมทั้งสิ้น' in val_str:
                continue
            all_cols[col_idx] = val_str

    for col_idx, label in all_cols.items():
        if not label.startswith('รวม'):
            continue

        parts = label.split()
        if len(parts) < 2:
            continue
        prefix = parts[1]  # e.g., "4.5"

        # First, look for "รวม" sub-summary children (e.g., "รวม 4.5.1", "รวม 4.5.2")
        summary_children = []
        for c_idx, c_label in all_cols.items():
            if c_idx == col_idx:
                continue
            if c_label.startswith('รวม'):
                c_parts = c_label.split()
                if len(c_parts) >= 2 and c_parts[1].startswith(prefix + '.'):
                    summary_children.append((c_idx, c_label))

        if summary_children:
            # Use sub-summary columns as children (e.g., "รวม 4.5.1" + "รวม 4.5.2")
            summaries.append((col_idx, label, summary_children))
        else:
            # Fallback: use non-รวม columns (e.g., "4.5.1 ..." + "4.5.2 ...")
            child_cols = []
            for c_idx, c_label in all_cols.items():
                if c_idx == col_idx:
                    continue
                if c_label.startswith(prefix + '.') and not c_label.startswith('รวม'):
                    child_cols.append((c_idx, c_label))
            if child_cols:
                summaries.append((col_idx, label, child_cols))

    return summaries


def reconcile_cross_column(ws, period_label, sheet_name, results, layout=None):
    """Check that summary columns = sum of child columns for every data row"""
    category = "Cross-Column ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)

    summaries = find_summary_columns(ws, layout)
    if not summaries:
        return 0

    data_start = find_data_start_row(ws, layout)
    label_col = layout.label_col
    total_checks = 0

    # Ratio rows should not be summed across columns
    RATIO_KEYWORDS = ['สัดส่วน', 'อัตรา', '%', 'ร้อยละ']

    for row_idx in range(data_start, ws.max_row + 1):
        row_label_cell = ws.cell(row=row_idx, column=label_col).value
        if row_label_cell is None:
            continue
        row_label = str(row_label_cell).strip()
        if not row_label:
            continue

        # Skip ratio/percentage rows — ratios are not additive
        if any(kw in row_label for kw in RATIO_KEYWORDS):
            continue

        for summary_col, summary_label, child_cols in summaries:
            summary_val = ws.cell(row=row_idx, column=summary_col).value
            if summary_val is None:
                continue
            try:
                summary_val = float(summary_val)
            except (ValueError, TypeError):
                continue

            # Sum child columns
            child_sum = 0.0
            skip = False
            for c_idx, _ in child_cols:
                c_val = ws.cell(row=row_idx, column=c_idx).value
                if c_val is None:
                    c_val = 0.0
                try:
                    child_sum += float(c_val)
                except (ValueError, TypeError):
                    skip = True
                    break
            if skip:
                continue

            # Only check rows where either value is non-zero
            if summary_val == 0.0 and child_sum == 0.0:
                continue

            child_labels = '+'.join(cl[:15] for _, cl in child_cols)
            display_row = row_label[:40]

            results.append(CheckResult(
                category=category,
                check_name="{} | {} vs sum({})".format(
                    display_row, summary_label[:20], child_labels[:40]),
                source_label=summary_label[:20],
                source_value=summary_val,
                target_label="sum(children)",
                target_value=child_sum,
            ))
            total_checks += 1

    return total_checks


# ==========================================
# Column-Total: sum(columns) = รวมทั้งสิ้น
# ==========================================
# ตรวจว่า ผลรวม column ย่อย = column รวมทั้งสิ้น
# ถ้าตรง → มั่นใจได้ว่าทุก column ถูก aggregate มาถูกต้อง

# Key rows to check for column-total consistency
COLUMN_TOTAL_CHECKS = [
    (['รายได้รวม'], 'รายได้รวม'),
    (['EBITDA'], 'EBITDA'),
    (['กำไร', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
    (['กำไร', 'สุทธิ'], 'กำไรสุทธิ'),
]


def _find_top_level_data_cols(ws, layout):
    """Find top-level data columns (not sub-totals, not Common Size, not รวมทั้งสิ้น)

    กลุ่มธุรกิจ: BU columns ที่ row 7 = "จำนวนเงิน" (skip Common Size)
    กลุ่มบริการ/บริการ: "รวม X." columns ใน row 6 (BU-level summaries)
    """
    header_row = layout.header_row  # row 6
    sg_row = layout.sg_row          # row 7
    total_col = None
    bu_summary_cols = []  # "รวม X." columns (BU-level)
    bu_data_cols = []     # BU columns with "จำนวนเงิน" in row 7

    for col_idx in range(1, ws.max_column + 1):
        val6 = ws.cell(row=header_row, column=col_idx).value
        if val6 is None:
            continue
        val6_str = str(val6).strip()

        if val6_str == 'รวมทั้งสิ้น':
            total_col = col_idx
            continue

        if 'รายละเอียด' in val6_str:
            continue

        # Check if this is a "รวม X." BU-level summary (กลุ่มบริการ/บริการ sheets)
        if val6_str.startswith('รวม') and re.match(r'รวม\s+\d+\.', val6_str):
            bu_summary_cols.append(col_idx)
            continue

        # Check if row 7 has "จำนวนเงิน" (กลุ่มธุรกิจ sheet)
        val7 = ws.cell(row=sg_row, column=col_idx).value
        if val7 and str(val7).strip() == 'จำนวนเงิน':
            bu_data_cols.append(col_idx)

    # Prefer BU summary cols (กลุ่มบริการ/บริการ), fallback to จำนวนเงิน cols (กลุ่มธุรกิจ)
    data_cols = bu_summary_cols if bu_summary_cols else bu_data_cols
    return total_col, data_cols


def reconcile_column_total(ws, period_label, sheet_name, results, layout=None):
    """Check: sum of top-level columns = รวมทั้งสิ้น for key rows"""
    category = "Column-Total ({}) - {}".format(period_label, sheet_name)

    if layout is None:
        layout = detect_sheet_layout(ws)

    total_col, data_cols = _find_top_level_data_cols(ws, layout)
    if total_col is None or len(data_cols) == 0:
        print("    Warning: Cannot find รวมทั้งสิ้น or data columns in {}".format(sheet_name))
        return 0

    total_checks = 0

    for keywords, desc in COLUMN_TOTAL_CHECKS:
        row_idx = find_row_index(ws, layout.data_start, keywords, layout.label_col)
        if row_idx is None:
            continue

        total_val = _read_excel_cell_float(ws, row_idx, total_col)
        if total_val is None:
            continue

        col_sum = 0.0
        skip = False
        for col_idx in data_cols:
            val = _read_excel_cell_float(ws, row_idx, col_idx)
            if val is None:
                skip = True
                break
            col_sum += val
        if skip:
            continue

        if total_val == 0.0 and col_sum == 0.0:
            continue

        results.append(CheckResult(
            category=category,
            check_name="{}: sum(columns) vs รวมทั้งสิ้น".format(desc),
            source_label="sum({} cols)".format(len(data_cols)),
            source_value=col_sum,
            target_label="รวมทั้งสิ้น",
            target_value=total_val,
        ))
        total_checks += 1

    return total_checks


# ==========================================
# Cross-Sheet Total: รวมทั้งสิ้น ต้นทุน vs หมวดบัญชี
# ==========================================

def reconcile_cross_sheet_total(ws_cost, ws_gl, period_label, level, results,
                                 layout_cost=None, layout_gl=None):
    """Cross-sheet: compare รวมทั้งสิ้น column between ต้นทุน and หมวดบัญชี"""
    category = "Cross-sheet Total ({}) - {}".format(period_label, level)

    if layout_cost is None:
        layout_cost = detect_sheet_layout(ws_cost)
    if layout_gl is None:
        layout_gl = detect_sheet_layout(ws_gl)

    # Find รวมทั้งสิ้น column in both sheets
    cost_total_col, _ = _find_top_level_data_cols(ws_cost, layout_cost)
    gl_total_col, _ = _find_top_level_data_cols(ws_gl, layout_gl)

    if cost_total_col is None or gl_total_col is None:
        return 0

    # Rows to compare: EBT and Net Profit (same value in both sheets)
    cross_checks = [
        (['รายได้รวม'], ['รายได้รวม'], 'รายได้รวม'),
        (['EBITDA'], ['EBITDA'], 'EBITDA'),
        (['12.', 'กำไร', 'EBT'], ['3.', 'กำไร', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
        (['14.', 'กำไร', 'สุทธิ'], ['5.', 'กำไร', 'สุทธิ'], 'กำไรสุทธิ'),
    ]

    total_checks = 0
    for cost_kw, gl_kw, desc in cross_checks:
        cost_row = find_row_index(ws_cost, layout_cost.data_start, cost_kw, layout_cost.label_col)
        gl_row = find_row_index(ws_gl, layout_gl.data_start, gl_kw, layout_gl.label_col)
        if cost_row is None or gl_row is None:
            continue

        cost_val = _read_excel_cell_float(ws_cost, cost_row, cost_total_col)
        gl_val = _read_excel_cell_float(ws_gl, gl_row, gl_total_col)
        if cost_val is None or gl_val is None:
            continue

        results.append(CheckResult(
            category=category,
            check_name="{} (รวมทั้งสิ้น)".format(desc),
            source_label="ต้นทุน_{}".format(level),
            source_value=cost_val,
            target_label="หมวดบัญชี_{}".format(level),
            target_value=gl_val,
        ))
        total_checks += 1

    return total_checks


# ==========================================
# Cross-Sheet: ต้นทุน vs หมวดบัญชี (net profit)
# ==========================================
# หมายเหตุ:
#   ต้นทุน (COSTTYPE) มี GROUP 01-14, หมวดบัญชี (GLGROUP) มี GROUP 01-05
#   เทียบตรงได้เฉพาะยอดสุทธิ เช่น EBT, กำไรสุทธิ
#   การเทียบ per-row อื่น (เช่น revenue, expense) ต้อง aggregate SUB_GROUP
#   ของ COSTTYPE ให้ตรงกับ GLGROUP ก่อน ซึ่งยังไม่ได้ทำ
#
# สิ่งที่ยังไม่ได้ตรวจ (TODO):
#   - Cross-sheet revenue per-SG/Product (ต้อง aggregate COSTTYPE GROUP 01+09 = GLGROUP GROUP 01)
#   - Cross-sheet expense per-SG/Product (ต้อง aggregate COSTTYPE GROUP 02+04+06+07+10 = GLGROUP GROUP 02)
#   - Cross-sheet EBIT/EBITDA per-SG/Product

# Rows that have the SAME value between ต้นทุน and หมวดบัญชี
CROSS_SHEET_CHECKS_SG = [
    # (costtype_keywords, glgroup_keywords, description)
    (['12.', 'กำไร', 'EBT'], ['3.', 'กำไร', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
    (['14.', 'กำไร', 'สุทธิ'], ['5.', 'กำไร', 'สุทธิ'], 'กำไรสุทธิ'),
]


def reconcile_cross_sheet_sg(ws_cost, ws_gl, period_label, results,
                              layout_cost=None, layout_gl=None):
    """Cross-sheet: ต้นทุน_กลุ่มบริการ vs หมวดบัญชี_กลุ่มบริการ (EBT + Net Profit per SG)"""
    category = "Cross-sheet SG ({}) - ต้นทุน vs หมวดบัญชี".format(period_label)

    if layout_cost is None:
        layout_cost = detect_sheet_layout(ws_cost)
    if layout_gl is None:
        layout_gl = detect_sheet_layout(ws_gl)

    sg_map_cost, summary_map_cost = build_sg_column_map(ws_cost, layout_cost)
    sg_map_gl, summary_map_gl = build_sg_column_map(ws_gl, layout_gl)

    total_checks = 0

    for cost_kw, gl_kw, desc in CROSS_SHEET_CHECKS_SG:
        cost_row = find_row_index(ws_cost, layout_cost.data_start, cost_kw, layout_cost.label_col)
        gl_row = find_row_index(ws_gl, layout_gl.data_start, gl_kw, layout_gl.label_col)
        if cost_row is None or gl_row is None:
            continue

        # Match SG columns between sheets
        for sg_name, cost_col in sg_map_cost.items():
            gl_col = sg_map_gl.get(sg_name)
            if gl_col is None:
                continue

            cost_val = _read_excel_cell_float(ws_cost, cost_row, cost_col)
            gl_val = _read_excel_cell_float(ws_gl, gl_row, gl_col)
            if cost_val is None or gl_val is None:
                continue
            if cost_val == 0.0 and gl_val == 0.0:
                continue

            sg_label = sg_name[:45]
            results.append(CheckResult(
                category=category,
                check_name="{} - {}".format(desc, sg_label),
                source_label="ต้นทุน",
                source_value=cost_val,
                target_label="หมวดบัญชี",
                target_value=gl_val,
            ))
            total_checks += 1

    return total_checks


def reconcile_cross_sheet_product(ws_cost, ws_gl, period_label, results,
                                   layout_cost=None, layout_gl=None):
    """Cross-sheet: ต้นทุน_บริการ vs หมวดบัญชี_บริการ (EBT + Net Profit per Product)"""
    category = "Cross-sheet Product ({}) - ต้นทุน vs หมวดบัญชี".format(period_label)

    if layout_cost is None:
        layout_cost = detect_sheet_layout(ws_cost)
    if layout_gl is None:
        layout_gl = detect_sheet_layout(ws_gl)

    pk_map_cost, _ = build_product_column_map(ws_cost, layout_cost)
    pk_map_gl, _ = build_product_column_map(ws_gl, layout_gl)

    if len(pk_map_cost) == 0 or len(pk_map_gl) == 0:
        return 0

    total_checks = 0

    for cost_kw, gl_kw, desc in CROSS_SHEET_CHECKS_SG:
        cost_row = find_row_index(ws_cost, layout_cost.data_start, cost_kw, layout_cost.label_col)
        gl_row = find_row_index(ws_gl, layout_gl.data_start, gl_kw, layout_gl.label_col)
        if cost_row is None or gl_row is None:
            continue

        for pk, cost_col in pk_map_cost.items():
            gl_col = pk_map_gl.get(pk)
            if gl_col is None:
                continue

            cost_val = _read_excel_cell_float(ws_cost, cost_row, cost_col)
            gl_val = _read_excel_cell_float(ws_gl, gl_row, gl_col)
            if cost_val is None or gl_val is None:
                continue
            if cost_val == 0.0 and gl_val == 0.0:
                continue

            results.append(CheckResult(
                category=category,
                check_name="{} - PK {}".format(desc, pk),
                source_label="ต้นทุน",
                source_value=cost_val,
                target_label="หมวดบัญชี",
                target_value=gl_val,
            ))
            total_checks += 1

    return total_checks


# ==========================================
# Main
# ==========================================

def main():
    args = parse_args()
    config = build_file_config(args)
    if config is None:
        return

    date_obj = config['date_obj']
    month = config['month']

    print("\n📅 วันที่ข้อมูล: {}".format(date_obj.strftime('%d/%m/%Y')))
    print("🏢 บริษัท: {}".format(config['company']))

    csv_files = {key: DATA_DIR / fname for key, fname in config['csv_filenames'].items()}
    excel_files = {key: REPORT_DIR / fname for key, fname in config['excel_filenames'].items()}

    # ตรวจว่าไฟล์มีอยู่จริง
    for key, path in {**csv_files, **excel_files}.items():
        if not path.exists():
            print("❌ ไม่พบไฟล์: {}".format(path))
            if 'Report' in str(path):
                print("   กรุณา copy จาก report_generator/output/ มาไว้ใน report/")
                print("   หรือระบุชื่อไฟล์ด้วย --excel-mth / --excel-ytd")
            return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / config['output_filename']

    sheet_sg = {'COSTTYPE': 'ต้นทุน_กลุ่มบริการ', 'GLGROUP': 'หมวดบัญชี_กลุ่มบริการ'}
    sheet_svc = {'COSTTYPE': 'ต้นทุน_บริการ', 'GLGROUP': 'หมวดบัญชี_บริการ'}

    # Load CSVs
    print("\nLoading CSV files...")
    csv_data = {}
    for key, path in csv_files.items():
        csv_data[key] = read_csv_auto_encoding(str(path))
        csv_data[key]['VALUE'] = csv_data[key]['VALUE'].astype(float)
        print("  {} -> {} rows".format(key, len(csv_data[key])))

    # Validate period
    period_warnings = validate_period(csv_data, excel_files, month)
    if period_warnings:
        print("\n⚠️  คำเตือนเรื่องงวดข้อมูล:")
        for w in period_warnings:
            print("   - {}".format(w))
        print("   กรุณาตรวจสอบว่าไฟล์ CSV และ Excel เป็นงวดเดียวกัน")
        print()

    results = []

    for period in ['MTH', 'YTD']:
        print("\n{}".format('=' * 60))
        print("Reconciling {} - Service Group & Product level".format(period))
        print('=' * 60)

        wb = openpyxl.load_workbook(str(excel_files[period]), data_only=True)

        for csv_type in ['COSTTYPE', 'GLGROUP']:
            csv_key = '{}_{}'.format(csv_type, period)
            csv_df = csv_data[csv_key]

            # Service Group
            sn = sheet_sg[csv_type]
            ws = wb[sn]
            layout = detect_sheet_layout(ws)
            print("\n  SG: {} vs {}".format(csv_key, sn))
            n = reconcile_service_group(csv_df, ws, csv_type, period, sn, results, layout)
            print("    -> {} checks".format(n))

            # EBIT at Service Group level
            print("\n  EBIT SG: {} vs {}".format(csv_key, sn))
            n = reconcile_ebit_sg(csv_df, ws, csv_type, period, sn, results, layout)
            print("    -> {} checks".format(n))

            # EBITDA at Service Group level
            print("\n  EBITDA SG: {} vs {}".format(csv_key, sn))
            n = reconcile_ebitda_sg(csv_df, ws, csv_type, period, sn, results, layout)
            print("    -> {} checks".format(n))

            # Cross-column consistency (SG sheet)
            print("\n  Cross-Column: {}".format(sn))
            n = reconcile_cross_column(ws, period, sn, results, layout)
            print("    -> {} checks".format(n))

            # Product
            sn = sheet_svc[csv_type]
            ws = wb[sn]
            layout_svc = detect_sheet_layout(ws)
            print("\n  Product: {} vs {}".format(csv_key, sn))
            n = reconcile_product(csv_df, ws, csv_type, period, sn, results, layout_svc)
            print("    -> {} checks".format(n))

            # EBIT at Product level
            print("\n  EBIT Product: {} vs {}".format(csv_key, sn))
            n = reconcile_ebit_product(csv_df, ws, csv_type, period, sn, results, layout_svc)
            print("    -> {} checks".format(n))

            # EBITDA at Product level
            print("\n  EBITDA Product: {} vs {}".format(csv_key, sn))
            n = reconcile_ebitda_product(csv_df, ws, csv_type, period, sn, results, layout_svc)
            print("    -> {} checks".format(n))

            # Cross-column consistency (Product sheet)
            print("\n  Cross-Column: {}".format(sn))
            n = reconcile_cross_column(ws, period, sn, results, layout_svc)
            print("    -> {} checks".format(n))

        # Column-Total: sum(columns) = รวมทั้งสิ้น for each sheet
        for csv_type in ['COSTTYPE', 'GLGROUP']:
            for sheet_map in [sheet_sg, sheet_svc]:
                sn = sheet_map[csv_type]
                ws = wb[sn]
                ly = detect_sheet_layout(ws)
                print("\n  Column-Total: {}".format(sn))
                n = reconcile_column_total(ws, period, sn, results, ly)
                print("    -> {} checks".format(n))

        # Cross-sheet Total: รวมทั้งสิ้น ต้นทุน vs หมวดบัญชี
        for sheet_map, level in [(sheet_sg, 'กลุ่มบริการ'), (sheet_svc, 'บริการ')]:
            ws_c = wb[sheet_map['COSTTYPE']]
            ws_g = wb[sheet_map['GLGROUP']]
            ly_c = detect_sheet_layout(ws_c)
            ly_g = detect_sheet_layout(ws_g)
            print("\n  Cross-sheet Total: ต้นทุน vs หมวดบัญชี {}".format(level))
            n = reconcile_cross_sheet_total(ws_c, ws_g, period, level, results, ly_c, ly_g)
            print("    -> {} checks".format(n))

        # Cross-sheet: ต้นทุน vs หมวดบัญชี (EBT + กำไรสุทธิ per SG/Product)
        ws_cost_sg = wb[sheet_sg['COSTTYPE']]
        ws_gl_sg = wb[sheet_sg['GLGROUP']]
        layout_cost_sg = detect_sheet_layout(ws_cost_sg)
        layout_gl_sg = detect_sheet_layout(ws_gl_sg)
        print("\n  Cross-sheet SG: ต้นทุน vs หมวดบัญชี กลุ่มบริการ")
        n = reconcile_cross_sheet_sg(ws_cost_sg, ws_gl_sg, period, results,
                                      layout_cost_sg, layout_gl_sg)
        print("    -> {} checks".format(n))

        ws_cost_svc = wb[sheet_svc['COSTTYPE']]
        ws_gl_svc = wb[sheet_svc['GLGROUP']]
        layout_cost_svc = detect_sheet_layout(ws_cost_svc)
        layout_gl_svc = detect_sheet_layout(ws_gl_svc)
        print("\n  Cross-sheet Product: ต้นทุน vs หมวดบัญชี บริการ")
        n = reconcile_cross_sheet_product(ws_cost_svc, ws_gl_svc, period, results,
                                           layout_cost_svc, layout_gl_svc)
        print("    -> {} checks".format(n))

        wb.close()

    # Print results
    passed = sum(1 for r in results if r.passed)
    failed = sum(1 for r in results if not r.passed)
    total = len(results)

    # Print FAIL details
    fail_results = [r for r in results if not r.passed]
    if fail_results:
        print("\n{}".format('=' * 130))
        print("{:^130}".format("FAILED CHECKS DETAIL"))
        print('=' * 130)

        current_cat = None
        for r in fail_results:
            if r.category != current_cat:
                current_cat = r.category
                print("\n--- {} ---".format(current_cat))
                print("{:<60} | {:>18} | {:>18} | {:>15}".format(
                    'Check', 'CSV', 'Excel', 'Diff'))
                print("{}-+-{}-+-{}-+-{}".format('-'*60, '-'*18, '-'*18, '-'*15))

            print("{:<60} | {:>18,.2f} | {:>18,.2f} | {:>15,.2f}".format(
                r.check_name[:60], r.source_value, r.target_value, r.diff))

    print("\n{}".format('=' * 130))
    print("Summary: {} PASSED / {} FAILED / {} Total checks".format(passed, failed, total))

    # Pass rate by category
    print("\n  Pass rate by category:")
    cat_stats = {}
    for r in results:
        if r.category not in cat_stats:
            cat_stats[r.category] = {'passed': 0, 'failed': 0}
        if r.passed:
            cat_stats[r.category]['passed'] += 1
        else:
            cat_stats[r.category]['failed'] += 1

    for cat, stats in cat_stats.items():
        total_cat = stats['passed'] + stats['failed']
        pct = stats['passed'] / total_cat * 100 if total_cat > 0 else 0
        print("    {} -> {}/{} ({:.1f}%)".format(cat[:70], stats['passed'], total_cat, pct))

    print('=' * 130)

    # Export
    records = [{
        'Category': r.category, 'Check': r.check_name,
        'Source Label': r.source_label, 'Source Value': r.source_value,
        'Target Label': r.target_label, 'Target Value': r.target_value,
        'Difference': r.diff, 'Status': r.status,
    } for r in results]

    df_results = pd.DataFrame(records)
    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        pd.DataFrame([{
            'Total': total, 'Passed': passed, 'Failed': failed,
            'Rate': "{:.1f}%".format(passed/total*100) if total > 0 else 'N/A',
            'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }]).to_excel(writer, sheet_name='Summary', index=False)

        df_results.to_excel(writer, sheet_name='All Checks', index=False)
        df_failed = df_results[df_results['Status'] == 'FAIL']
        if len(df_failed) > 0:
            df_failed.to_excel(writer, sheet_name='Failed Checks', index=False)

    print("\nExported to: {}".format(str(output_path)))


if __name__ == '__main__':
    main()
