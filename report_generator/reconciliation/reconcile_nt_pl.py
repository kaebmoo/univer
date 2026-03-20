"""
NT P&L Reconciliation Script
==============================
ตรวจกระทบยอดระหว่าง CSV source data กับ Excel report
รองรับทั้ง MTH (รายเดือน) และ YTD (สะสม)

การตรวจสอบ:
1. CSV vs Excel: เทียบยอด "รวมทั้งสิ้น" ทุก GROUP row
2. CSV vs Excel: เทียบยอดรายกลุ่มธุรกิจ (per-BU)
3. CSV vs Excel: เทียบยอดรายกลุ่มบริการ (per-SERVICE_GROUP)
4. Cross-sheet: ต้นทุน vs หมวดบัญชี consistency
5. Internal: พันธมิตร + ไม่รวมพันธมิตร = รวมทั้งสิ้น
"""

import pandas as pd
import openpyxl
import json
import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from datetime import datetime

# ==========================================
# PATH CONFIGURATION
# ==========================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / 'data'
REPORT_DIR = SCRIPT_DIR / 'report'
OUTPUT_DIR = SCRIPT_DIR / 'output'

TOLERANCE = 0.001  # ยอมรับผลต่างไม่เกิน 0.001 บาท (floating point)


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='NT P&L Reconciliation - BU level + Cross-sheet + Alliance')
    parser.add_argument('--date', type=str, required=True,
                        help='วันที่ข้อมูล YYYYMMDD (เช่น 20260131)')
    parser.add_argument('--company', type=str, default='NT',
                        help='รหัสบริษัท (default: NT)')
    parser.add_argument('--excel-mth', type=str, default=None,
                        help='ชื่อไฟล์ Excel report MTH (ถ้าไม่ระบุจะใช้ Report_{company}_{YYYYMM}.xlsx)')
    parser.add_argument('--excel-ytd', type=str, default=None,
                        help='ชื่อไฟล์ Excel report YTD (ถ้าไม่ระบุจะใช้ Report_{company}_YTD_{YYYYMM}.xlsx)')
    return parser.parse_args()


def build_file_config(args):
    """Build file paths from --date argument"""
    try:
        date_obj = datetime.strptime(args.date, '%Y%m%d')
    except ValueError:
        print("❌ รูปแบบวันที่ไม่ถูกต้อง: {} (ใช้ YYYYMMDD)".format(args.date))
        return None

    month = date_obj.strftime('%Y%m')
    company = args.company

    # CSV filenames (auto-generate from date)
    csv_filenames = {
        'COSTTYPE_MTH': 'TRN_PL_COSTTYPE_{}_MTH_TABLE_{}.csv'.format(company, args.date),
        'COSTTYPE_YTD': 'TRN_PL_COSTTYPE_{}_YTD_TABLE_{}.csv'.format(company, args.date),
        'GLGROUP_MTH':  'TRN_PL_GLGROUP_{}_MTH_TABLE_{}.csv'.format(company, args.date),
        'GLGROUP_YTD':  'TRN_PL_GLGROUP_{}_YTD_TABLE_{}.csv'.format(company, args.date),
    }

    # Excel filenames (auto or user-specified)
    excel_mth = args.excel_mth or 'Report_{}_{}.xlsx'.format(company, month)
    excel_ytd = args.excel_ytd or 'Report_{}_YTD_{}.xlsx'.format(company, month)
    excel_filenames = {'MTH': excel_mth, 'YTD': excel_ytd}

    output_filename = 'reconciliation_report_{}.xlsx'.format(month)

    return {
        'date_obj': date_obj,
        'month': month,
        'company': company,
        'csv_filenames': csv_filenames,
        'excel_filenames': excel_filenames,
        'output_filename': output_filename,
    }


def validate_period(csv_data, excel_files, expected_month):
    """Validate that CSV and Excel are for the same period"""
    warnings = []

    # Check CSV TIME_KEY
    for key, df in csv_data.items():
        if 'TIME_KEY' in df.columns:
            time_keys = df['TIME_KEY'].astype(str).unique()
            for tk in time_keys:
                if not tk.startswith(expected_month):
                    warnings.append("CSV {}: TIME_KEY={} ไม่ตรงกับงวด {}".format(
                        key, tk, expected_month))
                    break

    # Check Excel header for period text
    for period, path in excel_files.items():
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            found = False
            for row_idx in range(1, 10):
                for col_idx in range(1, 6):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val and 'ประจำเดือน' in str(val):
                        header_text = str(val)
                        if expected_month not in header_text:
                            warnings.append("Excel {} header: '{}' - ตรวจสอบว่าตรงกับงวด {} หรือไม่".format(
                                period, header_text.strip(), expected_month))
                        found = True
                        break
                if found:
                    break
            wb.close()
        except Exception:
            pass

    return warnings

# ==========================================
# Data Classes
# ==========================================

@dataclass
class CheckResult:
    category: str
    check_name: str
    source_label: str
    source_value: float
    target_label: str
    target_value: float
    tolerance: float = TOLERANCE

    @property
    def diff(self) -> float:
        return self.source_value - self.target_value

    @property
    def passed(self) -> bool:
        return abs(self.diff) <= self.tolerance

    @property
    def status(self) -> str:
        return "PASS" if self.passed else "FAIL"

# ==========================================
# CSV Reader
# ==========================================

def read_csv_auto_encoding(filepath: str) -> pd.DataFrame:
    """Read CSV with auto-detection of encoding (TIS-620/cp874/utf-8)"""
    for enc in ['utf-8', 'cp874', 'tis-620', 'latin-1']:
        try:
            df = pd.read_csv(filepath, encoding=enc)
            # Verify by checking that column names are readable
            if 'TIME_KEY' in df.columns or 'GROUP' in df.columns:
                return df
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"Cannot read {filepath} with any known encoding")

def aggregate_csv_totals(df: pd.DataFrame) -> Dict[str, float]:
    """Aggregate CSV by GROUP → total VALUE"""
    result = {}
    for group, grp_df in df.groupby('GROUP'):
        result[group] = grp_df['VALUE'].astype(float).sum()
    return result

def aggregate_csv_by_bu(df: pd.DataFrame) -> Dict[Tuple[str, str], float]:
    """Aggregate CSV by (GROUP, BU) → VALUE"""
    result = {}
    for (group, bu), grp_df in df.groupby(['GROUP', 'BU']):
        result[(group, bu)] = grp_df['VALUE'].astype(float).sum()
    return result

def aggregate_csv_by_service_group(df: pd.DataFrame) -> Dict[Tuple[str, str, str], float]:
    """Aggregate CSV by (GROUP, BU, SERVICE_GROUP) → VALUE"""
    result = {}
    for (group, bu, sg), grp_df in df.groupby(['GROUP', 'BU', 'SERVICE_GROUP']):
        result[(group, bu, sg)] = grp_df['VALUE'].astype(float).sum()
    return result

def aggregate_csv_by_product(df: pd.DataFrame) -> Dict[Tuple[str, str], float]:
    """Aggregate CSV by (GROUP, PRODUCT_KEY) → VALUE"""
    result = {}
    for (group, pk), grp_df in df.groupby(['GROUP', 'PRODUCT_KEY']):
        result[(group, str(pk))] = grp_df['VALUE'].astype(float).sum()
    return result

# ==========================================
# Excel Reader - Dynamic Structure Detection
# ==========================================

class ExcelSheetReader:
    """Read and parse NT P&L Excel report sheet with dynamic structure detection"""

    def __init__(self, ws, sheet_name: str):
        self.ws = ws
        self.sheet_name = sheet_name
        self.max_row = ws.max_row
        self.max_col = ws.max_column

        # Detect layout: find "รายละเอียด" cell to determine label column
        self.label_col = 2  # default
        self._detect_label_column()

        # Detect structure
        self.data_start_row = self._find_data_start_row()
        self.header_info = self._parse_headers()
        self.row_data = self._parse_rows()

    def _detect_label_column(self):
        """Find the column containing 'รายละเอียด' header"""
        for row_idx in range(1, 16):
            for col_idx in range(1, 11):
                val = self.ws.cell(row=row_idx, column=col_idx).value
                if val and 'รายละเอียด' in str(val):
                    self.label_col = col_idx
                    return

    def _find_data_start_row(self) -> int:
        """Find the first data row (looking for '1' + 'รายได้' in label column)"""
        for row_idx in range(1, min(15, self.max_row + 1)):
            val = self.ws.cell(row=row_idx, column=self.label_col).value
            if val is not None:
                val_str = str(val).strip()
                if val_str.startswith('1') and ('รายได้' in val_str or 'รวมรายได้' in val_str):
                    return row_idx
        return 10  # default

    def _parse_headers(self) -> Dict:
        """Parse header structure to find column positions"""
        info = {
            'total_col': None,       # รวมทั้งสิ้น column (จำนวนเงิน)
            'alliance_col': None,    # พันธมิตร column (จำนวนเงิน)
            'non_alliance_col': None, # ไม่รวมพันธมิตร column (จำนวนเงิน)
            'bu_columns': {},        # BU name → column index (จำนวนเงิน)
            'sg_columns': {},        # Service group name → column index
        }

        # Find the header rows (typically row 6-9)
        # Look for "รวมทั้งสิ้น" header
        for row_idx in range(1, self.data_start_row):
            for col_idx in range(1, min(self.max_col + 1, 250)):
                val = self.ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                val_str = str(val).strip()

                # Total column
                if val_str == 'รวมทั้งสิ้น':
                    info['total_col'] = col_idx

                # Alliance columns - look for จำนวนเงิน under พันธมิตร and ไม่รวมพันธมิตร
                if val_str == 'พันธมิตร' and info['alliance_col'] is None:
                    info['alliance_col'] = col_idx
                if 'ไม่รวมพันธมิตร' in val_str and info['non_alliance_col'] is None:
                    info['non_alliance_col'] = col_idx

                # BU columns - try to normalize any header that could be a BU
                if 'กลุ่มธุรกิจ' in val_str or 'กลุ่มบริการอื่น' in val_str or 'รายได้อื่น' in val_str or 'นโยบายภาครัฐ' in val_str or 'ค่าใช้จ่ายอื่น' in val_str:
                    bu_name = self._normalize_bu_name(val_str)
                    if bu_name and bu_name not in info['bu_columns']:
                        info['bu_columns'][bu_name] = col_idx

                # Service group columns
                if 'กลุ่มบริการ' in val_str and 'รวม' not in val_str:
                    info['sg_columns'][val_str] = col_idx

        # Default: data column is right after label column
        if info['total_col'] is None:
            info['total_col'] = self.label_col + 1

        return info

    def _normalize_bu_name(self, text: str) -> Optional[str]:
        """Normalize BU name from Excel header to match CSV BU values

        Excel headers may use different casing or slightly different names.
        This mapping handles case-insensitive matching.
        """
        text = text.strip()
        # Map Excel BU keywords (case-insensitive) → CSV BU identifiers
        mappings = {
            'HARD INFRASTRUCTURE': '1.กลุ่มธุรกิจ HARD INFRASTRUCTURE',
            'INTERNATIONAL': '2.กลุ่มธุรกิจ INTERNATIONAL',
            'MOBILE': '3.กลุ่มธุรกิจ MOBILE',
            'FIXED LINE': '4.กลุ่มธุรกิจ FIXED LINE & BROADBAND',
            'DIGITAL': '5.กลุ่มธุรกิจ DIGITAL',
            'ICT SOLUTION': '6.กลุ่มธุรกิจ ICT SOLUTION',
            'ICT': '6.กลุ่มธุรกิจ ICT SOLUTION',
            'ไม่ใช่โทรคมนาคม': '7.กลุ่มบริการอื่นไม่ใช่โทรคมนาคม',
            'รายได้อื่น/ค่าใช้จ่ายอื่น': '8.รายได้อื่น/ค่าใช้จ่ายอื่น',
            'รายได้อื่น': '8.รายได้อื่น/ค่าใช้จ่ายอื่น',
        }
        text_lower = text.lower()
        for key, csv_bu in mappings.items():
            if key.lower() in text_lower:
                return csv_bu
        return None

    def _parse_rows(self) -> Dict[str, Dict]:
        """Parse all data rows with their labels and values"""
        rows = {}
        data_col_start = self.label_col + 1  # data starts after label column
        for row_idx in range(self.data_start_row, self.max_row + 1):
            label = self.ws.cell(row=row_idx, column=self.label_col).value
            if label is None:
                continue
            label = str(label).strip()
            if not label or label.startswith('หมายเหตุ') or label.startswith('คำนวณ'):
                continue

            row_values = {}
            for col_idx in range(data_col_start, min(self.max_col + 1, 250)):
                val = self.ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    try:
                        row_values[col_idx] = float(val)
                    except (ValueError, TypeError):
                        pass

            rows[label] = {
                'row_idx': row_idx,
                'values': row_values,
                'total': row_values.get(self.header_info['total_col'], None)
            }

        return rows

    def get_total_value(self, label_keywords: List[str]) -> Optional[float]:
        """Get the 'รวมทั้งสิ้น' value for a row matching keywords"""
        for label, data in self.row_data.items():
            if all(kw in label for kw in label_keywords):
                return data['total']
        return None

    def get_bu_value(self, label_keywords: List[str], bu_csv_name: str) -> Optional[float]:
        """Get value for a specific BU column"""
        bu_col = self.header_info['bu_columns'].get(bu_csv_name)
        if bu_col is None:
            return None

        for label, data in self.row_data.items():
            if all(kw in label for kw in label_keywords):
                return data['values'].get(bu_col)
        return None

    def get_alliance_values(self, label_keywords: List[str]) -> Tuple[Optional[float], Optional[float]]:
        """Get (พันธมิตร, ไม่รวมพันธมิตร) values for a row"""
        a_col = self.header_info['alliance_col']
        na_col = self.header_info['non_alliance_col']

        for label, data in self.row_data.items():
            if all(kw in label for kw in label_keywords):
                a_val = data['values'].get(a_col) if a_col else None
                na_val = data['values'].get(na_col) if na_col else None
                return (a_val, na_val)
        return (None, None)

# ==========================================
# Key Row Mappings
# ==========================================

# COSTTYPE CSV GROUP → Excel ต้นทุน row label keywords
COSTTYPE_ROW_MAP = [
    ('01.รายได้', ['1', 'รวมรายได้'], 'รายได้รวมจากการให้บริการ'),
    ('02.ต้นทุนบริการและต้นทุนขาย :', ['2.', 'ต้นทุนบริการ'], 'ต้นทุนบริการและต้นทุนขาย'),
    ('03.กำไร(ขาดทุน)ขั้นต้นจากการดำเนินงาน (1) - (2)', ['3.', 'กำไร', 'ขั้นต้น'], 'กำไรขั้นต้นจากการดำเนินงาน'),
    ('04.ค่าใช้จ่ายขายและการตลาด :', ['4.', 'ค่าใช้จ่ายขาย'], 'ค่าใช้จ่ายขายและการตลาด'),
    ('05.กำไร(ขาดทุน)หลังหักค่าใช้จ่ายขายและการตลาด (3) - (4)', ['5.', 'กำไร', 'หลังหัก'], 'กำไรหลังหักค่าใช้จ่ายขาย'),
    ('06.ค่าใช้จ่ายบริหารและสนับสนุน :', ['6.', 'ค่าใช้จ่ายบริหาร'], 'ค่าใช้จ่ายบริหารและสนับสนุน'),
    ('07.ต้นทุนทางการเงิน-ด้านการดำเนินงาน', ['7.', 'ต้นทุนทางการเงิน', 'ดำเนินงาน'], 'ต้นทุนทางการเงิน-ดำเนินงาน'),
    ('08.กำไร(ขาดทุน)ก่อนต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น (5) - (6) - (7)', ['8.', 'กำไร', 'จัดหาเงิน'], 'กำไรก่อนต้นทุนจัดหาเงิน'),
    ('09.ผลตอบแทนทางการเงินและรายได้อื่น', ['9.', 'ผลตอบแทน', 'รายได้อื่น'], 'ผลตอบแทนทางการเงินและรายได้อื่น'),
    ('10.ค่าใช้จ่ายอื่น', ['10.', 'ค่าใช้จ่ายอื่น'], 'ค่าใช้จ่ายอื่น'),
    ('11.ต้นทุนทางการเงิน-ด้านการจัดหาเงิน', ['11.', 'ต้นทุนทางการเงิน', 'จัดหา'], 'ต้นทุนทางการเงิน-จัดหาเงิน'),
    ('12.กำไร(ขาดทุน)ก่อนหักภาษีเงินได้ (EBT) (8) + (9) - (10) - (11)', ['12.', 'กำไร', 'ภาษี', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
    ('13.ภาษีเงินได้นิติบุคคล', ['13.', 'ภาษีเงินได้'], 'ภาษีเงินได้นิติบุคคล'),
    ('14.กำไร(ขาดทุน) สุทธิ (12) - (13)', ['14.', 'กำไร', 'สุทธิ'], 'กำไร(ขาดทุน) สุทธิ'),
]

# GLGROUP CSV GROUP → Excel หมวดบัญชี row label keywords
GLGROUP_ROW_MAP = [
    ('01.รายได้', ['1', 'รวมรายได้'], 'รวมรายได้'),
    ('02.ค่าใช้จ่าย', ['2', 'รวมค่าใช้จ่าย'], 'รวมค่าใช้จ่าย'),
    ('03.กำไร(ขาดทุน)ก่อนหักภาษีเงินได้ (EBT) (1)-(2)', ['3.', 'กำไร', 'ภาษี', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
    ('04.ภาษีเงินได้นิติบุคคล', ['4.', 'ภาษีเงินได้'], 'ภาษีเงินได้นิติบุคคล'),
    ('05.กำไร(ขาดทุน) สุทธิ (3)-(4)', ['5.', 'กำไร', 'สุทธิ'], 'กำไร(ขาดทุน) สุทธิ'),
]

# ==========================================
# Reconciliation Engine
# ==========================================

class PLReconciler:
    def __init__(self):
        self.results: List[CheckResult] = []

    def reconcile_csv_vs_excel_totals(
        self,
        csv_df: pd.DataFrame,
        excel_reader: ExcelSheetReader,
        csv_type: str,  # 'COSTTYPE' or 'GLGROUP'
        period_label: str,  # 'MTH' or 'YTD'
        sheet_name: str,
    ):
        """Compare CSV aggregated totals vs Excel 'รวมทั้งสิ้น' column"""
        category = f"CSV vs Excel ({period_label}) - {sheet_name}"

        # Choose row mapping
        row_map = COSTTYPE_ROW_MAP if csv_type == 'COSTTYPE' else GLGROUP_ROW_MAP

        # Aggregate CSV totals
        csv_totals = aggregate_csv_totals(csv_df)

        for csv_group, excel_keywords, desc in row_map:
            csv_val = csv_totals.get(csv_group)
            if csv_val is None:
                continue

            excel_val = excel_reader.get_total_value(excel_keywords)
            if excel_val is None:
                # Try alternative keywords
                alt_keywords = [kw for kw in excel_keywords if len(kw) > 2]
                if alt_keywords:
                    excel_val = excel_reader.get_total_value(alt_keywords)

            if excel_val is not None:
                self.results.append(CheckResult(
                    category=category,
                    check_name=f"{desc} (รวมทั้งสิ้น)",
                    source_label=f"CSV {csv_group}",
                    source_value=csv_val,
                    target_label=f"Excel {sheet_name}",
                    target_value=excel_val,
                ))

    def reconcile_csv_vs_excel_by_bu(
        self,
        csv_df: pd.DataFrame,
        excel_reader: ExcelSheetReader,
        csv_type: str,
        period_label: str,
        sheet_name: str,
    ):
        """Compare CSV per-BU totals vs Excel per-BU columns (all GROUP rows)"""
        category = f"CSV vs Excel by BU ({period_label}) - {sheet_name}"

        row_map = COSTTYPE_ROW_MAP if csv_type == 'COSTTYPE' else GLGROUP_ROW_MAP
        csv_by_bu = aggregate_csv_by_bu(csv_df)

        for csv_group, excel_keywords, desc in row_map:
            for bu_name in excel_reader.header_info['bu_columns'].keys():
                csv_val = csv_by_bu.get((csv_group, bu_name))
                if csv_val is None:
                    continue

                excel_val = excel_reader.get_bu_value(excel_keywords, bu_name)
                if excel_val is not None:
                    bu_short = bu_name.split('.')[-1].strip() if '.' in bu_name else bu_name
                    self.results.append(CheckResult(
                        category=category,
                        check_name=f"{desc} - {bu_short}",
                        source_label=f"CSV",
                        source_value=csv_val,
                        target_label=f"Excel",
                        target_value=excel_val,
                    ))

    def reconcile_cross_sheet(
        self,
        cost_reader: ExcelSheetReader,
        gl_reader: ExcelSheetReader,
        period_label: str,
        level: str,  # กลุ่มธุรกิจ / กลุ่มบริการ / บริการ
    ):
        """Compare ต้นทุน vs หมวดบัญชี sheets for consistency

        Checks both รวมทั้งสิ้น (total) and per-BU columns.
        """
        category = f"Cross-sheet ({period_label}) - {level}"

        # Key rows that must match between cost and GL sheets
        cross_checks = [
            (['รายได้รวม'], 'รายได้รวม'),
            (['ค่าใช้จ่ายรวม', 'รวมต้นทุนทางการเงิน'], 'ค่าใช้จ่ายรวม (รวมต้นทุนทางการเงิน)'),
            (['EBITDA'], 'EBITDA'),
        ]

        # กำไรสุทธิ - different keyword patterns between sheets
        cost_net = cost_reader.get_total_value(['14.', 'กำไร', 'สุทธิ'])
        if cost_net is None:
            cost_net = cost_reader.get_total_value(['กำไร', 'สุทธิ'])

        gl_net = gl_reader.get_total_value(['5.', 'กำไร', 'สุทธิ'])
        if gl_net is None:
            gl_net = gl_reader.get_total_value(['กำไร', 'สุทธิ'])

        if cost_net is not None and gl_net is not None:
            self.results.append(CheckResult(
                category=category,
                check_name="กำไร(ขาดทุน) สุทธิ (รวมทั้งสิ้น)",
                source_label=f"ต้นทุน_{level}",
                source_value=cost_net,
                target_label=f"หมวดบัญชี_{level}",
                target_value=gl_net,
            ))

        # Total-level cross checks
        for keywords, desc in cross_checks:
            cost_val = cost_reader.get_total_value(keywords)
            gl_val = gl_reader.get_total_value(keywords)

            if cost_val is not None and gl_val is not None:
                self.results.append(CheckResult(
                    category=category,
                    check_name=f"{desc} (รวมทั้งสิ้น)",
                    source_label=f"ต้นทุน_{level}",
                    source_value=cost_val,
                    target_label=f"หมวดบัญชี_{level}",
                    target_value=gl_val,
                ))

        # Per-BU cross checks
        # Use BU columns from both sheets — check BUs that exist in both
        cost_bus = cost_reader.header_info['bu_columns']
        gl_bus = gl_reader.header_info['bu_columns']
        common_bus = set(cost_bus.keys()) & set(gl_bus.keys())

        # Net profit per BU (different keywords per sheet)
        net_profit_checks = [
            (cost_reader, ['14.', 'กำไร', 'สุทธิ'], gl_reader, ['5.', 'กำไร', 'สุทธิ'], 'กำไรสุทธิ'),
        ]
        for cost_r, cost_kw, gl_r, gl_kw, desc in net_profit_checks:
            for bu_name in sorted(common_bus):
                cost_val = cost_r.get_bu_value(cost_kw, bu_name)
                gl_val = gl_r.get_bu_value(gl_kw, bu_name)
                if cost_val is not None and gl_val is not None:
                    bu_short = bu_name.split('.')[-1].strip() if '.' in bu_name else bu_name
                    self.results.append(CheckResult(
                        category=category,
                        check_name=f"{desc} - {bu_short}",
                        source_label=f"ต้นทุน_{level}",
                        source_value=cost_val,
                        target_label=f"หมวดบัญชี_{level}",
                        target_value=gl_val,
                    ))

        # Same-keyword rows per BU
        for keywords, desc in cross_checks:
            for bu_name in sorted(common_bus):
                cost_val = cost_reader.get_bu_value(keywords, bu_name)
                gl_val = gl_reader.get_bu_value(keywords, bu_name)
                if cost_val is not None and gl_val is not None:
                    bu_short = bu_name.split('.')[-1].strip() if '.' in bu_name else bu_name
                    self.results.append(CheckResult(
                        category=category,
                        check_name=f"{desc} - {bu_short}",
                        source_label=f"ต้นทุน_{level}",
                        source_value=cost_val,
                        target_label=f"หมวดบัญชี_{level}",
                        target_value=gl_val,
                    ))

    def check_column_total(
        self,
        reader: ExcelSheetReader,
        period_label: str,
        sheet_name: str,
    ):
        """Check: sum of BU columns = รวมทั้งสิ้น for key rows

        ตรวจว่าผลรวม column ย่อย (BU) = column รวมทั้งสิ้น
        ข้าม Common Size columns (เป็นสัดส่วน ไม่ใช่จำนวนเงิน)
        """
        category = f"Column-Total ({period_label}) - {sheet_name}"

        total_col = reader.header_info['total_col']
        bu_cols = reader.header_info['bu_columns']

        if total_col is None or len(bu_cols) == 0:
            return

        # Key rows to check
        check_keywords = [
            (['รวมรายได้'], 'รวมรายได้'),
            (['EBITDA'], 'EBITDA'),
            (['กำไร', 'ภาษี', 'EBT'], 'กำไรก่อนภาษี (EBT)'),
            (['กำไร', 'สุทธิ'], 'กำไรสุทธิ'),
        ]

        for keywords, desc in check_keywords:
            total_val = reader.get_total_value(keywords)
            if total_val is None:
                continue

            bu_sum = 0.0
            found_any = False
            for bu_name, bu_col in bu_cols.items():
                bu_val = reader.get_bu_value(keywords, bu_name)
                if bu_val is not None:
                    bu_sum += bu_val
                    found_any = True

            if not found_any:
                continue

            self.results.append(CheckResult(
                category=category,
                check_name=f"{desc}: sum({len(bu_cols)} BU) vs รวมทั้งสิ้น",
                source_label=f"sum(BU cols)",
                source_value=bu_sum,
                target_label="รวมทั้งสิ้น",
                target_value=total_val,
            ))

    def check_alliance_consistency(
        self,
        reader: ExcelSheetReader,
        period_label: str,
        sheet_name: str,
    ):
        """Check: พันธมิตร + ไม่รวมพันธมิตร = รวมทั้งสิ้น"""
        category = f"Alliance check ({period_label}) - {sheet_name}"

        # Key rows to check
        check_keywords = [
            (['รวมรายได้'], 'รวมรายได้'),
            (['กำไร', 'สุทธิ'], 'กำไร(ขาดทุน) สุทธิ'),
        ]

        for keywords, desc in check_keywords:
            total = reader.get_total_value(keywords)
            alliance, non_alliance = reader.get_alliance_values(keywords)

            if total is not None and alliance is not None and non_alliance is not None:
                computed_total = alliance + non_alliance
                self.results.append(CheckResult(
                    category=category,
                    check_name=f"{desc}: พันธมิตร+ไม่รวมพันธมิตร vs รวมทั้งสิ้น",
                    source_label="พันธมิตร + ไม่รวมพันธมิตร",
                    source_value=computed_total,
                    target_label="รวมทั้งสิ้น",
                    target_value=total,
                ))

    def print_results(self):
        """Print all results grouped by category"""
        passed = sum(1 for r in self.results if r.passed)
        failed = sum(1 for r in self.results if not r.passed)
        total = len(self.results)

        print(f"\n{'='*120}")
        print(f"{'NT P&L Reconciliation Report':^120}")
        print(f"{'Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S'):^120}")
        print(f"{'='*120}")

        current_category = None
        for r in self.results:
            if r.category != current_category:
                current_category = r.category
                print(f"\n--- {current_category} ---")
                print(f"{'Check':<55} | {'Source':>18} | {'Target':>18} | {'Diff':>15} | Status")
                print(f"{'-'*55}-+-{'-'*18}-+-{'-'*18}-+-{'-'*15}-+-------")

            status_mark = "PASS" if r.passed else "FAIL <<<"
            print(f"{r.check_name:<55} | {r.source_value:>18,.2f} | {r.target_value:>18,.2f} | {r.diff:>15,.2f} | {status_mark}")
            if not r.passed:
                print(f"  >>> {r.source_label} vs {r.target_label}")

        print(f"\n{'='*120}")
        print(f"Summary: {passed} PASSED / {failed} FAILED / {total} Total checks")
        print(f"{'='*120}\n")

        return {'passed': passed, 'failed': failed, 'total': total}

    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to DataFrame for export"""
        records = []
        for r in self.results:
            records.append({
                'Category': r.category,
                'Check': r.check_name,
                'Source Label': r.source_label,
                'Source Value': r.source_value,
                'Target Label': r.target_label,
                'Target Value': r.target_value,
                'Difference': r.diff,
                'Status': r.status,
            })
        return pd.DataFrame(records)

# ==========================================
# Main Execution
# ==========================================

def main():
    args = parse_args()
    config = build_file_config(args)
    if config is None:
        return

    date_obj = config['date_obj']
    month = config['month']

    print("\n📅 วันที่ข้อมูล: {}".format(date_obj.strftime('%d/%m/%Y')))
    print("🏢 บริษัท: {}".format(config['company']))

    csv_files = {key: DATA_DIR / fname for key, fname in config['csv_filenames'].items()}
    excel_files = {key: REPORT_DIR / fname for key, fname in config['excel_filenames'].items()}

    # ตรวจว่าไฟล์มีอยู่จริง
    for key, path in {**csv_files, **excel_files}.items():
        if not path.exists():
            print("❌ ไม่พบไฟล์: {}".format(path))
            if 'Report' in str(path):
                print("   กรุณา copy จาก report_generator/output/ มาไว้ใน report/")
                print("   หรือระบุชื่อไฟล์ด้วย --excel-mth / --excel-ytd")
            return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / config['output_filename']

    # Sheet names
    SHEETS = {
        'cost_biz': 'ต้นทุน_กลุ่มธุรกิจ',
        'cost_sg': 'ต้นทุน_กลุ่มบริการ',
        'cost_svc': 'ต้นทุน_บริการ',
        'gl_biz': 'หมวดบัญชี_กลุ่มธุรกิจ',
        'gl_sg': 'หมวดบัญชี_กลุ่มบริการ',
        'gl_svc': 'หมวดบัญชี_บริการ',
    }

    # Load CSV files
    print("\nLoading CSV files...")
    csv_data = {}
    for key, path in csv_files.items():
        print(f"  Reading {key}: {path.name}")
        csv_data[key] = read_csv_auto_encoding(str(path))
        csv_data[key]['VALUE'] = csv_data[key]['VALUE'].astype(float)
        print(f"    -> {len(csv_data[key])} rows loaded")

    # Validate period
    period_warnings = validate_period(csv_data, excel_files, month)
    if period_warnings:
        print("\n⚠️  คำเตือนเรื่องงวดข้อมูล:")
        for w in period_warnings:
            print(f"   - {w}")
        print("   กรุณาตรวจสอบว่าไฟล์ CSV และ Excel เป็นงวดเดียวกัน")
        print()

    # Load Excel files
    print("\nLoading Excel files...")
    excel_readers = {}

    for period in ['MTH', 'YTD']:
        excel_path = str(excel_files[period])
        print(f"  Reading {period}: {excel_files[period].name}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        for sheet_key, sheet_name in SHEETS.items():
            ws = wb[sheet_name]
            reader_key = f"{period}_{sheet_key}"
            excel_readers[reader_key] = ExcelSheetReader(ws, sheet_name)
            print(f"    -> Sheet '{sheet_name}': {ws.max_row} rows x {ws.max_column} cols")

        wb.close()

    # Run reconciliation
    reconciler = PLReconciler()

    for period in ['MTH', 'YTD']:
        print(f"\n{'='*60}")
        print(f"Running reconciliation for {period}...")
        print(f"{'='*60}")

        # 1. CSV vs Excel - COSTTYPE → ต้นทุน_กลุ่มธุรกิจ
        cost_csv = csv_data[f'COSTTYPE_{period}']
        cost_biz_reader = excel_readers[f'{period}_cost_biz']

        reconciler.reconcile_csv_vs_excel_totals(
            cost_csv, cost_biz_reader, 'COSTTYPE', period, 'ต้นทุน_กลุ่มธุรกิจ'
        )
        reconciler.reconcile_csv_vs_excel_by_bu(
            cost_csv, cost_biz_reader, 'COSTTYPE', period, 'ต้นทุน_กลุ่มธุรกิจ'
        )

        # 2. CSV vs Excel - GLGROUP → หมวดบัญชี_กลุ่มธุรกิจ
        gl_csv = csv_data[f'GLGROUP_{period}']
        gl_biz_reader = excel_readers[f'{period}_gl_biz']

        reconciler.reconcile_csv_vs_excel_totals(
            gl_csv, gl_biz_reader, 'GLGROUP', period, 'หมวดบัญชี_กลุ่มธุรกิจ'
        )
        reconciler.reconcile_csv_vs_excel_by_bu(
            gl_csv, gl_biz_reader, 'GLGROUP', period, 'หมวดบัญชี_กลุ่มธุรกิจ'
        )

        # 3. Cross-sheet consistency
        levels = [
            ('กลุ่มธุรกิจ', 'cost_biz', 'gl_biz'),
            ('กลุ่มบริการ', 'cost_sg', 'gl_sg'),
            ('บริการ', 'cost_svc', 'gl_svc'),
        ]
        for level_name, cost_key, gl_key in levels:
            cost_r = excel_readers[f'{period}_{cost_key}']
            gl_r = excel_readers[f'{period}_{gl_key}']
            reconciler.reconcile_cross_sheet(cost_r, gl_r, period, level_name)

        # 4. Column-Total: sum(BU columns) = รวมทั้งสิ้น
        for sheet_key in ['cost_biz', 'gl_biz']:
            reader = excel_readers[f'{period}_{sheet_key}']
            reconciler.check_column_total(reader, period, SHEETS[sheet_key])

        # 5. Alliance check on sheets that have it
        for sheet_key in ['cost_biz', 'gl_biz', 'cost_sg', 'gl_sg']:
            reader = excel_readers[f'{period}_{sheet_key}']
            reconciler.check_alliance_consistency(reader, period, SHEETS[sheet_key])

    # Print results
    summary = reconciler.print_results()

    # Export to Excel
    df_results = reconciler.to_dataframe()

    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        # Summary sheet
        summary_df = pd.DataFrame([{
            'Total Checks': summary['total'],
            'Passed': summary['passed'],
            'Failed': summary['failed'],
            'Pass Rate': f"{summary['passed']/summary['total']*100:.1f}%" if summary['total'] > 0 else 'N/A',
            'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # All results
        df_results.to_excel(writer, sheet_name='All Checks', index=False)

        # Failed only
        df_failed = df_results[df_results['Status'] == 'FAIL']
        if len(df_failed) > 0:
            df_failed.to_excel(writer, sheet_name='Failed Checks', index=False)

        # By category
        for cat in df_results['Category'].unique():
            cat_df = df_results[df_results['Category'] == cat]
            safe_name = cat[:31].replace('/', '_').replace('\\', '_')
            cat_df.to_excel(writer, sheet_name=safe_name, index=False)

    print(f"\nReport exported to: {output_path}")
    return reconciler

if __name__ == '__main__':
    main()
