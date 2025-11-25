#!/usr/bin/env python3
"""
Debug ratio calculation in Excel file
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

# Load the generated report
report_path = Path("./output/correct_report.xlsx")

if not report_path.exists():
    print(f"❌ Report not found: {report_path}")
    sys.exit(1)

wb = load_workbook(report_path, data_only=True)
ws = wb.active

print("\n" + "="*80)
print("🔍 Searching for all ratio-related rows")
print("="*80 + "\n")

# Scan all rows
for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
    cell_value = ws.cell(row=row_idx, column=2).value  # Column B
    if cell_value and "สัดส่วน" in str(cell_value):
        # Get value from column C (รวมทั้งสิ้น)
        grand_total = ws.cell(row=row_idx, column=3).value
        print(f"Row {row_idx}: {cell_value}")
        print(f"  Value in Column C: {grand_total}")
        print(f"  Type: {type(grand_total)}")
        if isinstance(grand_total, (int, float)):
            print(f"  As %: {grand_total*100:.2f}%")
        print()

print("="*80)

# Now check the specific ones we care about
print("\n" + "="*80)
print("📊 Checking specific calculation rows")
print("="*80 + "\n")

target_rows = [
    "รายได้บริการ",
    "     1. ต้นทุนบริการรวม",
    "         สัดส่วนต่อรายได้",
]

for label in target_rows:
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value == label:
            grand_total = ws.cell(row=row_idx, column=3).value
            print(f"{label}:")
            print(f"  Row {row_idx}, Value: {grand_total}")
            if isinstance(grand_total, (int, float)):
                if "สัดส่วน" in label:
                    print(f"  As %: {grand_total*100:.2f}%")
            print()
            break
