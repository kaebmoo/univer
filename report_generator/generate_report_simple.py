#!/usr/bin/env python3
"""
Simple Report Generator - Test script
Generates Excel report from CSV data using new DataAggregator
"""
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.data_loader import CSVLoader, DataProcessor, DataAggregator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.row_order import ROW_ORDER
from config.data_mapping import get_group_sub_group, is_calculated_row
from config.settings import settings
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def generate_simple_report(csv_path: Path, output_path: Path):
    """
    Generate simplified P&L Excel report

    Args:
        csv_path: Path to CSV file
        output_path: Path to save Excel file
    """
    logger.info(f"Loading CSV: {csv_path}")

    # Load CSV
    csv_loader = CSVLoader(encoding='tis-620')
    df = csv_loader.load_csv(csv_path)

    logger.info(f"Loaded {len(df)} rows")
    logger.info(f"Columns: {df.columns.tolist()}")

    # Process data
    data_processor = DataProcessor()
    df = data_processor.process_data(df)

    logger.info(f"Processing complete")
    logger.info(f"Unique GROUPs: {df['GROUP'].unique()[:10]}")

    # Create aggregator
    aggregator = DataAggregator(df)

    # Get unique BUs and service groups
    bu_list = data_processor.get_unique_business_units(df)
    logger.info(f"BUs: {bu_list}")

    # Build service group dict
    service_group_dict = {}
    for bu in bu_list:
        service_group_dict[bu] = data_processor.get_unique_service_groups(df, bu)

    logger.info(f"Service groups per BU: {[(bu, len(sgs)) for bu, sgs in service_group_dict.items()]}")

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    # Font settings
    font_name = settings.excel_font_name
    font_size = settings.excel_font_size

    # Start positions
    start_row = 5  # Row 6 (0-indexed = 5)
    start_col = 1  # Column B (0-indexed = 1)

    # Header rows (rows 1-5)
    # Row 2: Company name
    cell = ws.cell(row=2, column=2)
    cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Row 3: Report name
    cell = ws.cell(row=3, column=2)
    period_desc, months = data_processor.get_period_description(df, "MTH" if "MTH" in csv_path.name else "YTD")
    report_type = "มิติประเภทต้นทุน" if "COSTTYPE" in csv_path.name else "มิติหมวดบัญชี"
    cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {report_type}"
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Row 4: Period
    cell = ws.cell(row=4, column=2)
    cell.value = period_desc
    cell.font = Font(name=font_name, size=font_size, bold=True)

    # Build column structure
    current_col = start_col

    # Column 1: รายละเอียด
    ws.cell(row=start_row + 1, column=current_col + 1).value = "รายละเอียด"
    ws.column_dimensions[get_column_letter(current_col + 1)].width = 50
    current_col += 1

    column_mapping = {}  # Map column index to (bu, service_group) or "TOTAL"

    # For each BU
    for bu in bu_list:
        # BU total column
        bu_total_col = current_col
        ws.cell(row=start_row + 1, column=bu_total_col + 1).value = f"รวม {bu}"
        ws.column_dimensions[get_column_letter(bu_total_col + 1)].width = 18
        column_mapping[bu_total_col] = ("BU_TOTAL", bu, None)
        current_col += 1

        # Service groups
        for sg in service_group_dict.get(bu, []):
            sg_col = current_col
            ws.cell(row=start_row + 1, column=sg_col + 1).value = sg
            ws.column_dimensions[get_column_letter(sg_col + 1)].width = 18
            column_mapping[sg_col] = ("SERVICE_GROUP", bu, sg)
            current_col += 1

    # Grand total column
    grand_total_col = current_col
    ws.cell(row=start_row + 1, column=grand_total_col + 1).value = "รวมทั้งสิ้น"
    ws.column_dimensions[get_column_letter(grand_total_col + 1)].width = 20
    column_mapping[grand_total_col] = ("GRAND_TOTAL", None, None)

    # Format header row
    for col_idx in range(start_col, current_col + 1):
        cell = ws.cell(row=start_row + 1, column=col_idx + 1)
        cell.font = Font(name=font_name, size=font_size, bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # Write data rows
    current_row = start_row + 1
    all_row_data = {}  # Store for calculated rows

    for level, label, is_calc, formula, is_bold in ROW_ORDER:
        if not label:  # Skip empty rows
            current_row += 1
            continue

        # Write label
        cell = ws.cell(row=current_row + 1, column=start_col + 1)
        cell.value = label
        cell.font = Font(name=font_name, size=font_size, bold=is_bold)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Color for section headers
        if is_bold and level == 0:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        # Get row data
        if is_calculated_row(label):
            # Calculate from previous rows
            row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
        else:
            # Get from CSV
            row_data = aggregator.get_row_data(label, bu_list, service_group_dict)

        # Store for later calculations
        all_row_data[label] = row_data

        # Write values to columns
        for col_idx, (col_type, bu, sg) in column_mapping.items():
            if col_type == "BU_TOTAL":
                key = f"BU_TOTAL_{bu}"
            elif col_type == "SERVICE_GROUP":
                key = f"{bu}_{sg}"
            else:  # GRAND_TOTAL
                key = "GRAND_TOTAL"

            value = row_data.get(key, 0)

            cell = ws.cell(row=current_row + 1, column=col_idx + 1)

            # Handle ratio cells (they might be None for division by zero)
            if value is None:
                cell.value = ""  # Leave blank
            elif isinstance(value, float) and "สัดส่วน" in label:
                # Percentage format
                cell.value = value
                cell.number_format = '0.00%'
            else:
                cell.value = value
                # Number format: positive, (negative) in red, zero as blank
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'

            cell.font = Font(name=font_name, size=font_size, bold=is_bold)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        current_row += 1

    # Freeze panes (at first data row, grand total column)
    ws.freeze_panes = f"{get_column_letter(grand_total_col + 1)}{start_row + 2}"

    # Save
    wb.save(output_path)
    logger.info(f"Report saved to: {output_path}")


if __name__ == "__main__":
    # Test with actual file
    csv_path = Path("../data/TRN_PL_COSTTYPE_NT_MTH_TABLE_20251031.csv")
    output_path = Path("./output/test_report.xlsx")

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)

    output_path.parent.mkdir(exist_ok=True)

    generate_simple_report(csv_path, output_path)
    logger.info("Done!")
