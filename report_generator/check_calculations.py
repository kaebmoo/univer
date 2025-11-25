#!/usr/bin/env python3
"""
Check calculation values from the report
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

# Load the generated report
report_path = Path("./output/correct_report.xlsx")

if not report_path.exists():
    print(f"❌ Report not found: {report_path}")
    sys.exit(1)

wb = load_workbook(report_path, data_only=True)
ws = wb.active

# Find specific rows
target_rows = {
    "รายได้รวม": None,
    "ค่าใช้จ่ายรวม (ไม่รวมต้นทุนทางการเงิน)": None,
    "ค่าใช้จ่ายรวม (รวมต้นทุนทางการเงิน)": None,
    "EBITDA": None,
    "รายได้บริการ": None,
    "     1. ต้นทุนบริการรวม": None,
    "         สัดส่วนต่อรายได้": None,
    "     2. ต้นทุนบริการ - ค่าเสื่อมราคาฯ": None,
    "     3. ต้นทุนบริการ - ไม่รวมค่าใช้จ่ายบุคลากรและค่าเสื่อมราคาฯ": None,
}

# Scan rows to find targets
for row_idx in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row_idx, column=2).value  # Column B
    if cell_value in target_rows:
        # Get value from column C (รวมทั้งสิ้น)
        grand_total = ws.cell(row=row_idx, column=3).value
        target_rows[cell_value] = grand_total

# Print results
print("\n" + "="*80)
print("📊 Calculation Check Results")
print("="*80 + "\n")

for label, value in target_rows.items():
    if value is not None:
        if "สัดส่วน" in label:
            print(f"{label:<60} {value:>15.2%}")
        else:
            print(f"{label:<60} {value:>18,.2f}")
    else:
        print(f"{label:<60} {'NOT FOUND':>18}")

print("\n" + "="*80)
print("\n✅ Check complete!")
