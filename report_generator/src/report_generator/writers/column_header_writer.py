"""
Column Header Writer - EXACT COPY of main_generator.py logic
ทำตาม main_generator.py ทุกอย่าง 100%
"""
from typing import List
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..columns.base_column_builder import ColumnDef
import logging

logger = logging.getLogger(__name__)


class ColumnHeaderWriter:
    """Write column headers - EXACT match to main_generator.py"""

    def __init__(self, config, formatter):
        self.config = config
        self.formatter = formatter

    @staticmethod
    def _remove_leading_zero(text: str) -> str:
        """
        Remove leading zero from BU/SG names for display

        Examples:
            '01.กลุ่มธุรกิจ HARD INFRASTRUCTURE' -> '1.กลุ่มธุรกิจ HARD INFRASTRUCTURE'
            '1.1 กลุ่มบริการ...' -> '1.1 กลุ่มบริการ...' (no change)
            'รวม 01.กลุ่มธุรกิจ...' -> 'รวม 1.กลุ่มธุรกิจ...'

        Args:
            text: Text that may contain leading zeros

        Returns:
            Text with leading zeros removed
        """
        if not text:
            return text

        # Handle "รวม XX.กลุ่มธุรกิจ..." format
        if text.startswith('รวม ') and len(text) > 7:
            prefix = 'รวม '
            main_text = text[5:]  # After "รวม "
        else:
            prefix = ''
            main_text = text

        # Check if starts with "0X." pattern
        if len(main_text) >= 3 and main_text[0] == '0' and main_text[1].isdigit() and main_text[2] == '.':
            # Remove leading zero: "01." -> "1."
            main_text = main_text[1:]

        return prefix + main_text
    
    def write(self, ws, columns: List[ColumnDef]):
        """
        Write column headers inline - EXACT logic from main_generator.py
        
        Algorithm:
        1. Write individual columns inline
        2. Track BU and SG ranges
        3. Write BU/SG headers AFTER products done
        """
        start_row = self.config.start_row
        start_col = self.config.start_col
        font_name = self.config.font_name
        font_size = self.config.font_size
        
        current_col = start_col
        column_idx = 0
        
        # Track for merged headers
        current_bu = None
        bu_first_sg_col = None
        bu_end_col = None
        bu_color = None
        
        for col in columns:
            col_index = current_col
            
            # Label column
            if col.col_type == 'label':
                cell = ws.cell(row=start_row + 1, column=col_index + 1)
                cell.value = "รายละเอียด"
                cell.font = Font(name=font_name, size=font_size, bold=True)
                cell.fill = PatternFill(start_color="F4DEDC", end_color="F4DEDC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 1, start_column=col_index + 1,
                               end_row=start_row + 4, end_column=col_index + 1)
                ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                current_col += 1
                column_idx += 1
                continue
            
            # Grand total column
            if col.col_type == 'grand_total':
                # For grand total with common size, we need 2-level header
                if self.config.include_common_size:
                    # Row 1: "รวมทั้งสิ้น" merged across amount + common size
                    header_cell = ws.cell(row=start_row + 1, column=col_index + 1)
                    header_cell.value = "รวมทั้งสิ้น"
                    header_cell.font = Font(name=font_name, size=font_size, bold=True)
                    header_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    header_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Merge row 1 across grand_total and its common_size
                    ws.merge_cells(start_row=start_row + 1, start_column=col_index + 1,
                                   end_row=start_row + 1, end_column=col_index + 2)
                    
                    # Row 2-4: "จำนวนเงิน" for amount column
                    amount_cell = ws.cell(row=start_row + 2, column=col_index + 1)
                    amount_cell.value = "จำนวนเงิน"
                    amount_cell.font = Font(name=font_name, size=font_size, bold=True)
                    amount_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    amount_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    amount_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws.merge_cells(start_row=start_row + 2, start_column=col_index + 1,
                                   end_row=start_row + 4, end_column=col_index + 1)
                    ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                    
                    current_col += 1
                    column_idx += 1
                else:
                    # Original behavior: single merged cell
                    cell = ws.cell(row=start_row + 1, column=col_index + 1)
                    cell.value = "รวมทั้งสิ้น"
                    cell.font = Font(name=font_name, size=font_size, bold=True)
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws.merge_cells(start_row=start_row + 1, start_column=col_index + 1,
                                   end_row=start_row + 4, end_column=col_index + 1)
                    ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                    current_col += 1
                    column_idx += 1
                continue
            
            # Common Size column
            if col.col_type == 'common_size':
                # Common Size is always row 2-4 (sub-header under BU or Grand Total)
                cell = ws.cell(row=start_row + 2, column=col_index + 1)
                cell.value = col.name
                cell.font = Font(name=font_name, size=font_size, bold=True)
                cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 2, start_column=col_index + 1,
                               end_row=start_row + 4, end_column=col_index + 1)
                ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                current_col += 1
                column_idx += 1
                continue
            
            # BU total column
            if col.col_type == 'bu_total':
                # Write previous BU header if exists
                if current_bu and bu_first_sg_col is not None and bu_end_col is not None:
                    if bu_end_col >= bu_first_sg_col:
                        bu_header_cell = ws.cell(row=start_row + 1, column=bu_first_sg_col + 1)
                        bu_header_cell.value = self._remove_leading_zero(current_bu)
                        bu_header_cell.font = Font(name=font_name, size=font_size, bold=True)
                        bu_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                        bu_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        bu_header_cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        ws.merge_cells(start_row=start_row + 1, start_column=bu_first_sg_col + 1,
                                       end_row=start_row + 1, end_column=bu_end_col + 1)
                
                # Write BU total column
                if self.config.include_common_size:
                    # Row 1: BU name (without "รวม") merged across amount + common size
                    header_cell = ws.cell(row=start_row + 1, column=col_index + 1)
                    # Remove "รวม " prefix from col.name
                    bu_display_name = self._remove_leading_zero(col.bu)  # Use BU name directly without leading zero
                    header_cell.value = bu_display_name
                    header_cell.font = Font(name=font_name, size=font_size, bold=True)
                    header_cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    header_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Merge row 1 across bu_total and its common_size
                    ws.merge_cells(start_row=start_row + 1, start_column=col_index + 1,
                                   end_row=start_row + 1, end_column=col_index + 2)
                    
                    # Set row height for BU header row
                    ws.row_dimensions[start_row + 1].height = 55
                    
                    # Row 2-4: "จำนวนเงิน" for amount column
                    amount_cell = ws.cell(row=start_row + 2, column=col_index + 1)
                    amount_cell.value = "จำนวนเงิน"
                    amount_cell.font = Font(name=font_name, size=font_size, bold=True)
                    amount_cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                    amount_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    amount_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws.merge_cells(start_row=start_row + 2, start_column=col_index + 1,
                                   end_row=start_row + 4, end_column=col_index + 1)
                    ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                    
                    current_col += 1
                    column_idx += 1
                else:
                    # Original behavior: single merged cell
                    cell = ws.cell(row=start_row + 1, column=col_index + 1)
                    cell.value = self._remove_leading_zero(col.name)
                    cell.font = Font(name=font_name, size=font_size, bold=True)
                    cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws.merge_cells(start_row=start_row + 1, start_column=col_index + 1,
                                   end_row=start_row + 4, end_column=col_index + 1)
                    ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width
                    
                    current_col += 1
                    column_idx += 1
                
                # Start new BU tracking
                current_bu = col.bu
                bu_first_sg_col = col_index + 1  # First SG column (will be updated if has common size)
                bu_end_col = None
                bu_color = col.color
                
                continue
            
            # SG total column + products
            if col.col_type == 'sg_total':
                sg_total_col = col_index
                sg_bu = col.bu
                sg_name = col.service_group
                sg_color = col.color
                
                # Write SG total column
                cell = ws.cell(row=start_row + 2, column=sg_total_col + 1)
                cell.value = col.name
                cell.font = Font(name=font_name, size=font_size, bold=True)
                cell.fill = PatternFill(start_color=sg_color, end_color=sg_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 1,
                               end_row=start_row + 4, end_column=sg_total_col + 1)
                ws.column_dimensions[get_column_letter(sg_total_col + 1)].width = col.width
                
                current_col += 1
                column_idx += 1
                
                # Now write products for this SG
                products_written = 0
                for prod_idx in range(column_idx, len(columns)):
                    prod_col = columns[prod_idx]
                    
                    # Check if still same SG
                    if prod_col.col_type != 'product':
                        break
                    if prod_col.bu != sg_bu or prod_col.service_group != sg_name:
                        break
                    
                    # Write product columns
                    prod_col_index = current_col
                    
                    # Row 3: Product key
                    pk_cell = ws.cell(row=start_row + 3, column=prod_col_index + 1)
                    pk_cell.value = prod_col.product_key
                    pk_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pk_cell.fill = PatternFill(start_color=sg_color, end_color=sg_color, fill_type="solid")
                    pk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pk_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    # Row 4: Product name
                    pn_cell = ws.cell(row=start_row + 4, column=prod_col_index + 1)
                    pn_cell.value = prod_col.product_name
                    pn_cell.font = Font(name=font_name, size=font_size, bold=True)
                    pn_cell.fill = PatternFill(start_color=sg_color, end_color=sg_color, fill_type="solid")
                    pn_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    pn_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    ws.column_dimensions[get_column_letter(prod_col_index + 1)].width = prod_col.width
                    
                    current_col += 1
                    column_idx += 1
                    products_written += 1
                
                sg_end_col = current_col - 1
                
                # Write SG header (row 2 only, merged across products)
                if products_written > 0:
                    sg_header_cell = ws.cell(row=start_row + 2, column=sg_total_col + 2)
                    sg_header_cell.value = self._remove_leading_zero(sg_name)  # ← ชื่อ SG ไม่ใช่ "รวม SG"
                    sg_header_cell.font = Font(name=font_name, size=font_size, bold=True)
                    sg_header_cell.fill = PatternFill(start_color=sg_color, end_color=sg_color, fill_type="solid")
                    sg_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sg_header_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Merge only row 2 (row 7)
                    ws.merge_cells(start_row=start_row + 2, start_column=sg_total_col + 2,
                                   end_row=start_row + 2, end_column=sg_end_col + 1)
                
                # Update BU end column
                bu_end_col = sg_end_col
                continue
            
            # Skip product columns (already handled above)
            if col.col_type == 'product':
                continue
            
            # SG column (for BU+SG builder)
            if col.col_type == 'sg':
                cell = ws.cell(row=start_row + 2, column=col_index + 1)
                cell.value = col.name
                cell.font = Font(name=font_name, size=font_size, bold=True)
                cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 2, start_column=col_index + 1,
                               end_row=start_row + 4, end_column=col_index + 1)
                ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width

                bu_end_col = col_index
                current_col += 1
                column_idx += 1
                continue

            # SATELLITE summary column (virtual column showing sum of SATELLITE groups)
            if col.col_type == 'satellite_summary':
                cell = ws.cell(row=start_row + 2, column=col_index + 1)
                cell.value = col.name
                cell.font = Font(name=font_name, size=font_size, bold=True)
                cell.fill = PatternFill(start_color=col.color, end_color=col.color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 2, start_column=col_index + 1,
                               end_row=start_row + 4, end_column=col_index + 1)
                ws.column_dimensions[get_column_letter(col_index + 1)].width = col.width

                bu_end_col = col_index
                current_col += 1
                column_idx += 1
                continue
        
        # Write last BU header
        if current_bu and bu_first_sg_col is not None and bu_end_col is not None:
            if bu_end_col >= bu_first_sg_col:
                bu_header_cell = ws.cell(row=start_row + 1, column=bu_first_sg_col + 1)
                bu_header_cell.value = self._remove_leading_zero(current_bu)
                bu_header_cell.font = Font(name=font_name, size=font_size, bold=True)
                bu_header_cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type="solid")
                bu_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                bu_header_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws.merge_cells(start_row=start_row + 1, start_column=bu_first_sg_col + 1,
                               end_row=start_row + 1, end_column=bu_end_col + 1)
        
        # Set row height for BU header row (row 1 of header = start_row + 1)
        if self.config.include_common_size:
            ws.row_dimensions[start_row + 1].height = 55
        
        logger.info(f"Wrote {len(columns)} column headers")
