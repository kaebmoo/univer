"""
Cell Formatter
Apply Excel formatting to cells (fonts, colors, borders, alignment)
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
import logging

logger = logging.getLogger(__name__)


class CellFormatter:
    """Format Excel cells"""
    
    def __init__(self, config):
        """
        Initialize formatter
        
        Args:
            config: ReportConfig instance
        """
        self.config = config
    
    def create_font(
        self,
        bold: bool = False,
        size: Optional[int] = None,
        color: str = "000000"
    ) -> Font:
        """
        Create font object
        
        Args:
            bold: Whether font is bold
            size: Font size (None = use default)
            color: Font color in hex (without #)
        
        Returns:
            Font object
        """
        return Font(
            name=self.config.font_name,
            size=size or self.config.font_size,
            bold=bold,
            color=color
        )
    
    def create_fill(self, color: str) -> PatternFill:
        """
        Create fill object
        
        Args:
            color: Fill color in hex (without #)
        
        Returns:
            PatternFill object
        """
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )
    
    def create_border(self, style: str = 'thin') -> Border:
        """
        Create border object for all sides
        
        Args:
            style: Border style (thin, medium, thick)
        
        Returns:
            Border object
        """
        side = Side(style=style)
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
    
    def create_alignment(
        self,
        horizontal: str = 'left',
        vertical: str = 'center',
        wrap_text: bool = False
    ) -> Alignment:
        """
        Create alignment object
        
        Args:
            horizontal: Horizontal alignment (left, center, right)
            vertical: Vertical alignment (top, center, bottom)
            wrap_text: Whether to wrap text
        
        Returns:
            Alignment object
        """
        return Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=wrap_text
        )
    
    def format_header_cell(
        self,
        cell,
        bg_color: str,
        bold: bool = True,
        font_size: Optional[int] = None
    ):
        """
        Format header cell
        
        Args:
            cell: Cell object
            bg_color: Background color (hex without #)
            bold: Whether text is bold
            font_size: Font size (None = use config)
        """
        cell.font = self.create_font(
            bold=bold,
            size=font_size or self.config.header_font_size
        )
        cell.fill = self.create_fill(bg_color)
        cell.border = self.create_border()
        cell.alignment = self.create_alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
    
    def format_data_cell(
        self,
        cell,
        value,
        is_bold: bool = False,
        bg_color: Optional[str] = None,
        is_percentage: bool = False
    ):
        """
        Format data cell
        
        Args:
            cell: Cell object
            value: Cell value
            is_bold: Whether text is bold
            bg_color: Background color (optional)
            is_percentage: Whether value is percentage
        """
        cell.font = self.create_font(bold=is_bold)
        
        if bg_color:
            cell.fill = self.create_fill(bg_color)
        
        cell.border = self.create_border()
        
        # Set value
        if value is None:
            cell.value = ""
        else:
            cell.value = value
        
        # Number format
        if value is not None:
            if is_percentage:
                cell.number_format = '0.00%'
                cell.alignment = self.create_alignment(
                    horizontal='right',
                    vertical='center'
                )
            else:
                # Positive: 1,234.00
                # Negative: (1,234.00) in red
                # Zero: blank
                cell.number_format = '#,##0.00;[Red](#,##0.00);""'
                cell.alignment = self.create_alignment(
                    horizontal='right',
                    vertical='center'
                )
        else:
            cell.alignment = self.create_alignment(
                horizontal='right',
                vertical='center'
            )
    
    def format_label_cell(
        self,
        cell,
        label: str,
        is_bold: bool = False,
        bg_color: Optional[str] = None
    ):
        """
        Format label cell (รายละเอียด column)
        
        Args:
            cell: Cell object
            label: Label text
            is_bold: Whether text is bold
            bg_color: Background color (optional)
        """
        cell.value = label
        cell.font = self.create_font(bold=is_bold)
        
        if bg_color:
            cell.fill = self.create_fill(bg_color)
        
        cell.border = self.create_border()
        cell.alignment = self.create_alignment(
            horizontal='left',
            vertical='center'
        )
    
    def format_info_box_cell(
        self,
        cell,
        text: str,
        is_title: bool = False
    ):
        """
        Format info box cell
        
        Args:
            cell: Cell object
            text: Cell text
            is_title: Whether this is title line
        """
        cell.value = text
        cell.font = self.create_font(
            bold=is_title,
            size=self.config.remark_font_size
        )
        cell.fill = self.create_fill(
            self.config.row_colors.get('info_box', 'F8CBAD')
        )
        cell.border = self.create_border()
        cell.alignment = self.create_alignment(
            horizontal='left',
            vertical='top',
            wrap_text=False
        )
    
    def format_remark_cell(
        self,
        cell,
        text: str,
        is_title: bool = False
    ):
        """
        Format remark cell
        
        Args:
            cell: Cell object
            text: Cell text
            is_title: Whether this is title ("หมายเหตุ")
        """
        cell.value = text
        cell.font = self.create_font(
            bold=is_title,
            size=self.config.remark_font_size
        )
        cell.alignment = self.create_alignment(
            horizontal='left',
            vertical='top',
            wrap_text=False
        )
    
    def set_column_width(self, ws, column_index: int, width: float):
        """
        Set column width
        
        Args:
            ws: Worksheet
            column_index: Column index (0-indexed)
            width: Width in characters
        """
        col_letter = get_column_letter(column_index + 1)
        ws.column_dimensions[col_letter].width = width
    
    def set_row_height(self, ws, row_index: int, height: float):
        """
        Set row height
        
        Args:
            ws: Worksheet
            row_index: Row index (0-indexed)
            height: Height in points
        """
        ws.row_dimensions[row_index + 1].height = height
    
    def set_freeze_panes(self, ws, row: int, col: int):
        """
        Set freeze panes
        
        Args:
            ws: Worksheet
            row: Row to freeze at (0-indexed)
            col: Column to freeze at (0-indexed)
        """
        cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
        ws.freeze_panes = cell_ref
        logger.info(f"Freeze panes set at: {cell_ref}")
