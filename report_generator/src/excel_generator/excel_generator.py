"""
Excel Generator - Main module for generating P&L Excel reports
"""
import pandas as pd
from openpyxl import Workbook
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging
from datetime import datetime

from .excel_formatter import ExcelFormatter
from .excel_calculator import ExcelCalculator
from ..data_loader import CSVLoader, DataProcessor, DataAggregator
from config.data_mapping import get_group_sub_group, is_calculated_row

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate formatted Excel P&L reports"""

    # Info box content
    INFO_BOX_CONTENT = """วัตถุประสงค์ของรายงาน :
เพื่อทราบผลดำเนินงานรายเดือนหลังหักค่าใช้จ่ายรวมทั้งหมด ประกอบด้วย ประเภทต้นทุน ขายและบริหาร
ใช้รายงานเพื่อการบริหารค่าใช้จ่ายแต่ละประเภทให้มีประสิทธิภาพ
ไม่เหมาะสมที่จะใช้เพื่อวัด Performance ส่วนงาน เนื่องจากเป็นการแสดงมิติค่าใช้จ่ายแบบรวม ไม่ได้แสดงเป็นค่าใช้จ่ายที่ส่วนงานมีอำนาจควบคุมได้หรือไม่ได้
ไม่เหมาะสมที่จะใช้เพื่อพิจารณาการยกเลิกบริการ เนื่องจากไม่ได้แสดงมิติต้นทุนผันแปรและคงที่ หากยกเลิกบริการแล้ว ควรทราบว่าต้นทุนใดบ้างลดลงได้ ต้นทุนใดบ้างยังคงอยู่ เพื่อเปรียบกับรายได้ที่จะหายไป"""

    def __init__(
        self,
        settings: Optional[dict] = None,
        bu_colors: Optional[dict] = None,
        row_colors: Optional[dict] = None
    ):
        """
        Initialize Excel Generator

        Args:
            settings: Settings dictionary
            bu_colors: BU color mapping
            row_colors: Row color mapping
        """
        self.settings = settings or {}
        self.bu_colors = bu_colors or {}
        self.row_colors = row_colors or {}

        # Initialize modules
        self.formatter = ExcelFormatter(
            font_name=self.settings.get('font_name', 'TH Sarabun New'),
            font_size=self.settings.get('font_size', 18),
            header_font_size=self.settings.get('header_font_size', 18),
            remark_font_size=self.settings.get('remark_font_size', 14)
        )
        self.calculator = ExcelCalculator()
        self.data_processor = DataProcessor()

        # Report structure
        self.start_row = self.settings.get('start_row', 5)  # 0-indexed (Row 6)
        self.start_col = self.settings.get('start_col', 1)  # 0-indexed (Column B)
        self.header_row = self.settings.get('header_row', 1)  # 0-indexed (Row 2)
        self.info_box_col = self.settings.get('info_box_col', 6)  # 0-indexed (Column G)
        self.info_box_row = self.settings.get('info_box_row', 1)  # 0-indexed (Row 2)

    def generate_report(
        self,
        data: pd.DataFrame,
        output_path: Path,
        report_type: str = "COSTTYPE",
        period_type: str = "MTH",
        remark_content: str = ""
    ) -> Path:
        """
        Generate complete P&L Excel report

        Args:
            data: Processed dataframe with P&L data
            output_path: Path to save Excel file
            report_type: Report type (COSTTYPE or GLGROUP)
            period_type: Period type (MTH or YTD)
            remark_content: Remark text content

        Returns:
            Path to generated file
        """
        logger.info(f"Generating {report_type} {period_type} report...")

        # Create data aggregator
        aggregator = DataAggregator(data)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L Report"

        # Get report structure
        columns = self._build_column_structure(data)
        rows = self._build_row_structure(data, report_type)

        # Write header (3 lines)
        self._write_header(ws, data, report_type, period_type)

        # Write info box (top right)
        self._write_info_box(ws)

        # Write column headers
        current_row = self._write_column_headers(ws, columns)

        # Write data rows
        current_row = self._write_data_rows(ws, data, aggregator, columns, rows, current_row)

        # Write remarks (bottom)
        if remark_content:
            self._write_remarks(ws, remark_content, current_row + 2)

        # Apply formatting
        self._apply_final_formatting(ws, columns, len(rows))

        # Set freeze panes (at first data row, first total column)
        freeze_row = self.start_row + 1  # After headers
        freeze_col = len(columns) - 1  # At "รวมทั้งสิ้น" column
        self.formatter.apply_freeze_panes(ws, freeze_row, freeze_col)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Report saved to: {output_path}")

        return output_path

    def _build_column_structure(self, data: pd.DataFrame) -> List[Dict]:
        """
        Build column structure with BU, SERVICE_GROUP, and totals

        Args:
            data: Dataframe with data

        Returns:
            List of column definitions
        """
        columns = []

        # Column 1: รายละเอียด
        columns.append({
            'name': 'รายละเอียด',
            'type': 'label',
            'width': 50
        })

        # Get unique BUs and Service Groups
        bus = self.data_processor.get_unique_business_units(data)

        for bu in bus:
            # Add BU total column (รวม + BU name)
            columns.append({
                'name': f'รวม {bu}',
                'type': 'bu_total',
                'bu': bu,
                'width': 18,
                'color': self.bu_colors.get(bu, '#FFFFFF')
            })

            # Get service groups for this BU
            service_groups = self.data_processor.get_unique_service_groups(data, bu)

            for sg in service_groups:
                columns.append({
                    'name': sg,
                    'type': 'service_group',
                    'bu': bu,
                    'service_group': sg,
                    'width': 18,
                    'color': self.bu_colors.get(bu, '#FFFFFF')
                })

                # Check if we need PRODUCT_KEY and PRODUCT_NAME columns
                # For now, we'll aggregate at SERVICE_GROUP level
                # Can be extended later if needed

        # Add grand total column
        columns.append({
            'name': 'รวมทั้งสิ้น',
            'type': 'grand_total',
            'width': 20,
            'color': '#FFD966'  # Yellow for total
        })

        return columns

    def _build_row_structure(
        self,
        data: pd.DataFrame,
        report_type: str
    ) -> List[Dict]:
        """
        Build row structure from row_order configuration

        Args:
            data: Dataframe with data
            report_type: Report type (COSTTYPE or GLGROUP)

        Returns:
            List of row definitions
        """
        from config.row_order import ROW_ORDER

        rows = []
        for level, label, is_calc, formula, is_bold in ROW_ORDER:
            rows.append({
                'level': level,
                'label': label,
                'is_calculated': is_calc,
                'formula': formula,
                'is_bold': is_bold,
                'color': self.row_colors.get('section_header', '#F8CBAD') if is_bold and level == 0 else None
            })

        return rows

    def _write_header(
        self,
        ws,
        data: pd.DataFrame,
        report_type: str,
        period_type: str
    ):
        """
        Write 3-line header at top

        Args:
            ws: Worksheet
            data: Dataframe with data
            report_type: Report type
            period_type: Period type
        """
        row = self.header_row

        # Line 1: Company name
        cell = ws.cell(row=row + 1, column=self.start_col + 1)
        cell.value = "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)"
        cell.font = self.formatter.create_font(bold=True, size=18)

        # Line 2: Report name
        dimension_name = "มิติประเภทต้นทุน" if "COSTTYPE" in report_type.upper() else "มิติหมวดบัญชี"
        cell = ws.cell(row=row + 2, column=self.start_col + 1)
        cell.value = f"รายงานผลดำเนินงานกลุ่มธุรกิจ/กลุ่มบริการ/รายบริการ - {dimension_name}"
        cell.font = self.formatter.create_font(bold=True, size=18)

        # Line 3: Period
        period_desc, _ = self.data_processor.get_period_description(data, period_type)
        cell = ws.cell(row=row + 3, column=self.start_col + 1)
        cell.value = period_desc
        cell.font = self.formatter.create_font(bold=True, size=18)

    def _write_info_box(self, ws):
        """
        Write info box at top right

        Args:
            ws: Worksheet
        """
        self.formatter.apply_info_box_style(
            ws,
            self.info_box_row,
            self.info_box_col,
            self.INFO_BOX_CONTENT,
            bg_color=self.settings.get('info_box_color', '#F8CBAD')
        )

    def _write_column_headers(
        self,
        ws,
        columns: List[Dict]
    ) -> int:
        """
        Write column headers

        Args:
            ws: Worksheet
            columns: Column structure

        Returns:
            Next row index
        """
        row = self.start_row
        col = self.start_col

        for i, col_def in enumerate(columns):
            cell = ws.cell(row=row + 1, column=col + i + 1)
            cell.value = col_def['name']

            # Apply styling
            if col_def['type'] == 'label':
                # "รายละเอียด" cell
                cell.fill = self.formatter.create_fill(self.row_colors.get('รายละเอียด', '#F4DEDC'))
                cell.font = self.formatter.create_font(bold=True)
            else:
                # Data columns
                color = col_def.get('color', '#FFFFFF')
                cell.fill = self.formatter.create_fill(color)
                cell.font = self.formatter.create_font(bold=True)

            cell.border = self.formatter.create_border()
            cell.alignment = self.formatter.create_alignment(horizontal='center', vertical='center', wrap_text=True)

            # Set column width
            self.formatter.set_column_width(ws, col + i, col_def.get('width', 18))

        return row + 1

    def _write_data_rows(
        self,
        ws,
        data: pd.DataFrame,
        aggregator: DataAggregator,
        columns: List[Dict],
        rows: List[Dict],
        start_row: int
    ) -> int:
        """
        Write all data rows

        Args:
            ws: Worksheet
            data: Dataframe with data
            aggregator: DataAggregator instance
            columns: Column structure
            rows: Row structure
            start_row: Starting row index

        Returns:
            Next row index after all rows written
        """
        current_row = start_row

        # Build BU list and service group dict for aggregator
        bu_list = self.data_processor.get_unique_business_units(data)
        service_group_dict = {}
        for bu in bu_list:
            service_group_dict[bu] = self.data_processor.get_unique_service_groups(data, bu)

        # Store all row data for calculated rows
        all_row_data = {}

        for row_def in rows:
            label = row_def['label']

            if not label:  # Empty row
                current_row += 1
                continue

            # Write row label
            cell = ws.cell(row=current_row + 1, column=self.start_col + 1)
            cell.value = label

            # Apply row styling
            is_bold = row_def['is_bold']
            row_color = row_def.get('color')

            if row_color:
                cell.fill = self.formatter.create_fill(row_color)

            cell.font = self.formatter.create_font(bold=is_bold)
            cell.border = self.formatter.create_border()
            cell.alignment = self.formatter.create_alignment(horizontal='left')

            # Get row data from aggregator
            if is_calculated_row(label):
                row_data = aggregator.calculate_summary_row(label, bu_list, service_group_dict, all_row_data)
            else:
                row_data = aggregator.get_row_data(label, bu_list, service_group_dict)

            # Store for later calculations
            all_row_data[label] = row_data

            # Write data cells
            for col_idx, col_def in enumerate(columns[1:], 1):  # Skip label column
                cell = ws.cell(row=current_row + 1, column=self.start_col + col_idx + 1)

                # Get value for this cell
                value = self._get_cell_value_from_row_data(row_data, col_def, label)

                if value is not None:
                    if isinstance(value, float) and "สัดส่วน" in label:
                        # Percentage format
                        cell.value = value
                        cell.number_format = '0.00%'
                    else:
                        cell.value = value
                        # Number format: positive, (negative) in red, zero as blank
                        cell.number_format = '#,##0.00;[Red](#,##0.00);""'

                # Apply cell styling
                color = col_def.get('color') if not row_color else row_color
                self.formatter.apply_data_cell_style(
                    cell,
                    bg_color=color,
                    bold=is_bold,
                    is_number=True
                )

            current_row += 1

        return current_row

    def _get_cell_value_from_row_data(
        self,
        row_data: Dict[str, float],
        col_def: Dict,
        label: str
    ) -> Optional[float]:
        """
        Get value for a specific cell from row data

        Args:
            row_data: Row data dictionary from aggregator
            col_def: Column definition
            label: Row label

        Returns:
            Cell value or None
        """
        col_type = col_def['type']

        if col_type == 'bu_total':
            key = f"BU_TOTAL_{col_def['bu']}"
        elif col_type == 'service_group':
            key = f"{col_def['bu']}_{col_def['service_group']}"
        elif col_type == 'grand_total':
            key = "GRAND_TOTAL"
        else:
            return None

        value = row_data.get(key, 0)

        # Handle None for division by zero
        if value is None:
            return None

        return value

    def _write_remarks(
        self,
        ws,
        remark_content: str,
        start_row: int
    ):
        """
        Write remarks section

        Args:
            ws: Worksheet
            remark_content: Remark text
            start_row: Starting row
        """
        self.formatter.apply_remark_style(ws, start_row, remark_content)

    def _apply_final_formatting(
        self,
        ws,
        columns: List[Dict],
        num_rows: int
    ):
        """
        Apply final formatting touches

        Args:
            ws: Worksheet
            columns: Column structure
            num_rows: Number of data rows
        """
        # Additional formatting can be added here
        pass
