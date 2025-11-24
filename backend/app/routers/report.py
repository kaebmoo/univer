"""
Report Router
"""

import logging
from fastapi import APIRouter, Depends, HTTPException, status
from typing import Dict, Any

from app.models.auth import UserInfo
from app.models.report import (
    ReportFilter,
    FilterOptions,
    UniverWorkbook
)
from app.routers.auth import get_current_user
from app.services.data_loader import data_loader
from app.services.report_calculator import report_calculator
from app.services.univer_converter import univer_converter
from app.services.univer_converter_crosstab import univer_converter_crosstab

logger = logging.getLogger(__name__)

# Create router
router = APIRouter(prefix="/report", tags=["Report"])


@router.get("/filters", response_model=FilterOptions)
async def get_filter_options(
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Get available filter options

    Returns available years, business groups, and services
    that can be used to filter reports.

    Args:
        current_user: Current authenticated user

    Returns:
        FilterOptions with available years, business groups, and services
    """
    logger.info(f"Filter options requested by {current_user.email}")

    try:
        filter_options = data_loader.get_available_filters()
        return filter_options

    except FileNotFoundError as e:
        logger.error(f"Data file not found: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data files not found. Please ensure CSV files are uploaded."
        )
    except Exception as e:
        logger.error(f"Error getting filter options: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to load filter options"
        )


@router.post("/generate")
async def generate_report(
    filter: ReportFilter,
    current_user: UserInfo = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Generate P&L report with specified filters

    Generate a complete Profit & Loss report based on the provided filters.
    Returns report data with revenue, costs, and calculated metrics.

    Args:
        filter: Report filter parameters (year, months, business_groups)
        current_user: Current authenticated user

    Returns:
        Complete report data with metrics and calculations

    Raises:
        HTTPException 400: If invalid filter parameters
        HTTPException 404: If no data found for the specified filters
        HTTPException 500: If report generation fails
    """
    logger.info(
        f"Report generation requested by {current_user.email}: "
        f"year={filter.year}, months={filter.months}, "
        f"groups={filter.business_groups or 'all'}"
    )

    try:
        # Generate report
        report = report_calculator.generate_full_report(
            year=filter.year,
            months=filter.months,
            business_groups=filter.business_groups,
            view_type='monthly',
            display_type='both',
            show_common_size=True
        )

        logger.info(
            f"Report generated successfully: "
            f"Revenue={report['data']['revenue'].get('Total', 0):,.2f}"
        )

        return report

    except ValueError as e:
        logger.error(f"Invalid filter parameters: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except FileNotFoundError as e:
        logger.error(f"Data file not found: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Data files not found"
        )
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate report: {str(e)}"
        )


@router.post("/univer")
async def generate_univer_snapshot(
    filter: ReportFilter,
    current_user: UserInfo = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Generate Univer snapshot for spreadsheet display

    Generate a Univer-compatible snapshot that can be loaded directly
    into the Univer spreadsheet component on the frontend.

    Args:
        filter: Report filter parameters (year, months, business_groups)
        current_user: Current authenticated user

    Returns:
        Univer snapshot object with sheets, cells, and styles

    Raises:
        HTTPException 400: If invalid filter parameters
        HTTPException 404: If no data found for the specified filters
        HTTPException 500: If snapshot generation fails
    """
    logger.info(
        f"Univer snapshot requested by {current_user.email}: "
        f"year={filter.year}, months={filter.months}"
    )

    try:
        # Use new crosstab converter instead of old converter
        snapshot = univer_converter_crosstab.convert_to_snapshot(
            year=filter.year,
            months=filter.months,
            business_groups=filter.business_groups,
            workbook_name=f"รายงานผลดำเนินงาน {filter.year}"
        )

        logger.info(
            f"Univer snapshot generated successfully: "
            f"{len(snapshot.get('sheets', {}))} sheets"
        )

        return snapshot

    except ValueError as e:
        logger.error(f"Invalid filter parameters: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except FileNotFoundError as e:
        logger.error(f"Data file not found: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Data files not found"
        )
    except Exception as e:
        logger.error(f"Error generating Univer snapshot: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Univer snapshot: {str(e)}"
        )


@router.post("/reload-data")
async def reload_data(
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Reload data from CSV files

    Force reload of data from CSV files, clearing any cached data.
    Useful when CSV files have been updated.

    Args:
        current_user: Current authenticated user

    Returns:
        Success message with data statistics
    """
    logger.info(f"Data reload requested by {current_user.email}")

    try:
        data_loader.reload_data()

        # Get data statistics
        df = data_loader.load_profit_loss_data()

        return {
            "message": "Data reloaded successfully",
            "statistics": {
                "total_rows": len(df),
                "years": sorted(df['YEAR'].unique().tolist()),
                "business_groups": len(df['BUSINESS_GROUP'].unique()),
                "last_load_time": data_loader.last_load_time.isoformat() if data_loader.last_load_time else None
            }
        }

    except Exception as e:
        logger.error(f"Error reloading data: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to reload data: {str(e)}"
        )


@router.post("/export")
async def export_to_excel(
    filter: ReportFilter,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Export report to Excel file in crosstab format

    Generate P&L report and export as Excel (.xlsx) file with
    crosstab layout showing cost centers across columns.

    Args:
        filter: Report filter parameters (year, months, business_groups)
        current_user: Current authenticated user

    Returns:
        Excel file download response

    Raises:
        HTTPException 400: If invalid filter parameters
        HTTPException 404: If no data found
        HTTPException 500: If export fails
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import pandas as pd
    from datetime import datetime as dt
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.services.group_order import sort_service_groups

    logger.info(
        f"Excel export requested by {current_user.email}: "
        f"year={filter.year}, months={filter.months}"
    )

    try:
        # Load raw data and create crosstab
        df = data_loader.filter_data(filter.year, filter.months, filter.business_groups)

        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No data found for the specified filters"
            )

        # Create Excel file in memory
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Helper function to create and format crosstab
            def create_crosstab(data, value_col, sheet_name):
                if data.empty:
                    return None

                # Create pivot table
                pivot = pd.pivot_table(
                    data,
                    index=['TYPE', 'หมวดบัญชี'],
                    columns='SERVICE_GROUP',
                    values=value_col,
                    aggfunc='sum',
                    fill_value=0
                )

                # Sort columns by service group order
                sorted_cols = sort_service_groups(pivot.columns.tolist())
                pivot = pivot[[col for col in sorted_cols if col in pivot.columns]]

                # Add Total column at the beginning
                total_col = pivot.sum(axis=1)
                pivot.insert(0, 'รวมทั้งหมด', total_col)

                # Write to Excel
                pivot.to_excel(writer, sheet_name=sheet_name, merge_cells=False)

                return writer.sheets[sheet_name]

            # Create Summary Sheet (P&L Complete)
            df['VALUE'] = df['REVENUE_VALUE'].fillna(0) + df['EXPENSE_VALUE'].fillna(0)
            summary_ws = create_crosstab(df, 'VALUE', 'สรุปรวม P&L')

            # Revenue crosstab
            df_revenue = df[df['REVENUE_VALUE'] > 0].copy()
            revenue_ws = create_crosstab(df_revenue, 'REVENUE_VALUE', 'รายได้')

            # Cost of Service crosstab
            df_cost = df[df['TYPE'] == '02 ต้นทุนบริการ'].copy()
            cost_ws = create_crosstab(df_cost, 'EXPENSE_VALUE', 'ต้นทุนบริการ')

            # Selling Expense crosstab
            df_selling = df[df['TYPE'] == '03 ค่าใช้จ่ายขายและการตลาด'].copy()
            selling_ws = create_crosstab(df_selling, 'EXPENSE_VALUE', 'ค่าใช้จ่ายขาย')

            # Admin Expense crosstab
            df_admin = df[df['TYPE'] == '04 ค่าใช้จ่ายสนับสนุน'].copy()
            admin_ws = create_crosstab(df_admin, 'EXPENSE_VALUE', 'ค่าใช้จ่ายสนับสนุน')

            # Apply formatting to all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Header formatting
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)

                # Format header rows (first 2 rows)
                for row in worksheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Format Total column (column C - รวมทั้งหมด)
                total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
                total_font = Font(bold=True)

                # Format total column header
                worksheet['C1'].fill = total_fill
                worksheet['C1'].font = Font(bold=True, color='000000', size=11)

                # Format total column data
                for row_idx in range(3, worksheet.max_row + 1):
                    cell = worksheet[f'C{row_idx}']
                    cell.fill = total_fill
                    cell.font = total_font
                    if isinstance(cell.value, (int, float)):
                        if cell.value < 0:
                            cell.number_format = '[Red](#,##0.00)'
                        else:
                            cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

                # Format other number columns
                for col_idx in range(4, worksheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(3, worksheet.max_row + 1):
                        cell = worksheet[f'{col_letter}{row_idx}']
                        if isinstance(cell.value, (int, float)):
                            if cell.value < 0:
                                cell.number_format = '[Red](#,##0.00)'
                            else:
                                cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')

                # Auto-adjust column widths
                worksheet.column_dimensions['A'].width = 35  # TYPE
                worksheet.column_dimensions['B'].width = 45  # หมวดบัญชี
                for col_idx in range(3, worksheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 18

                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border

                # Freeze panes (first 2 rows and first 2 columns)
                worksheet.freeze_panes = 'C3'

        # Prepare file for download
        output.seek(0)

        # Generate filename with month range
        sorted_months = sorted(filter.months)
        months_str = '-'.join(map(str, sorted_months))
        filename = f"P&L_Report_{filter.year}_{months_str}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        logger.info(f"Excel file generated successfully: {filename}")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except ValueError as e:
        logger.error(f"Invalid filter parameters: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export to Excel: {str(e)}"
        )


@router.get("/health")
async def report_health_check():
    """
    Health check for report service

    Returns:
        Service status and data loading status
    """
    return {
        "status": "healthy",
        "service": "Report",
        "data_loaded": data_loader.is_data_loaded,
        "last_load_time": data_loader.last_load_time.isoformat() if data_loader.last_load_time else None
    }
