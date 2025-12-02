"""
Excel to Univer JSON Converter
Reads an existing .xlsx file and converts it into a Univer-compatible JSON snapshot,
preserving styles, merged cells, and formatting.
"""

import logging
from typing import Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, Border, Alignment
from openpyxl.utils import get_column_letter
import json

logger = logging.getLogger(__name__)


class ExcelToUniverConverter:
    """
    Converts an openpyxl Workbook object to a Univer snapshot dictionary.
    """

    def __init__(self):
        self.styles_registry: Dict[str, str] = {}
        self.next_style_id: int = 0
        self.workbook: Optional[Workbook] = None

    def _register_style(self, style_definition: Dict[str, Any]) -> Optional[str]:
        """
        Registers a style definition to avoid duplicates and returns a style ID.
        Returns None if the style is empty.
        """
        if not style_definition:
            return None
            
        style_key = json.dumps(style_definition, sort_keys=True)
        if style_key in self.styles_registry:
            return self.styles_registry[style_key]

        style_id = f"s_{self.next_style_id}"
        self.styles_registry[style_key] = style_id
        self.next_style_id += 1
        return style_id

    def _build_styles_object(self) -> Dict[str, Any]:
        """
        Builds the final styles object for the Univer snapshot from the registry.
        """
        styles = {}
        for style_key, style_id in self.styles_registry.items():
            style_definition = json.loads(style_key)
            styles[style_id] = style_definition
        return styles

    def _convert_cell_style(self, cell: Cell) -> Dict[str, Any]:
        """
        Converts an openpyxl Cell's style to a Univer style definition dictionary.
        """
        if not cell.has_style:
            return {}

        style_def = {}

        # Font
        if cell.font:
            font: Font = cell.font
            if font.bold: style_def['bl'] = 1
            if font.italic: style_def['it'] = 1
            if font.underline and font.underline != 'none': style_def['ul'] = {'s': 1}
            if font.strikethrough: style_def['st'] = {'s': 1}
            if font.name: style_def['ff'] = font.name
            if font.sz: style_def['fs'] = int(font.sz)
            if font.color and font.color.value:
                # openpyxl color.value is often AARRGGBB, Univer wants #RRGGBB
                hex_color = str(font.color.value)
                if hex_color == '00000000':  # Auto/Black
                    style_def['cl'] = {'rgb': '#000000'}
                elif len(hex_color) == 8 and hex_color != 'FF000000':  # AARRGGBB
                    style_def['cl'] = {'rgb': f"#{hex_color[2:].upper()}"}
                elif len(hex_color) == 6:  # RRGGBB
                    style_def['cl'] = {'rgb': f"#{hex_color.upper()}"}

        # Fill (Background Color)
        if cell.fill:
            fill: Fill = cell.fill
            if fill.fill_type == 'solid' and fill.fgColor:
                bg_color = None

                # Try RGB first
                if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                    hex_color = str(fill.fgColor.rgb)
                    if len(hex_color) == 8 and hex_color != '00000000' and hex_color != 'FFFFFFFF':  # AARRGGBB
                        bg_color = f"#{hex_color[2:].upper()}"
                    elif len(hex_color) == 6 and hex_color != '000000':  # RRGGBB
                        bg_color = f"#{hex_color.upper()}"

                # Try theme color
                elif hasattr(fill.fgColor, 'theme') and fill.fgColor.theme is not None:
                    # Map common theme colors (simplified mapping)
                    theme_colors = {
                        0: '#FFFFFF',  # White
                        1: '#000000',  # Black
                        2: '#E7E6E6',  # Light gray
                        3: '#44546A',  # Dark blue
                        4: '#4472C4',  # Blue
                        5: '#ED7D31',  # Orange
                        6: '#A5A5A5',  # Gray
                        7: '#FFC000',  # Yellow
                        8: '#5B9BD5',  # Light blue
                        9: '#70AD47',  # Green
                    }
                    theme_idx = fill.fgColor.theme
                    if theme_idx in theme_colors:
                        bg_color = theme_colors[theme_idx]
                        # Apply tint if present
                        if hasattr(fill.fgColor, 'tint') and fill.fgColor.tint:
                            # Tint adjustment (simplified - full implementation would blend colors)
                            pass

                # Try indexed color
                elif hasattr(fill.fgColor, 'indexed') and fill.fgColor.indexed is not None:
                    # Excel indexed colors (simplified palette)
                    indexed_colors = {
                        64: '#000000',  # Black
                        65: '#FFFFFF',  # White
                        17: '#003366',  # Dark blue
                        43: '#4472C4',  # Blue
                        44: '#ED7D31',  # Orange
                        50: '#FFD966',  # Yellow
                    }
                    idx = fill.fgColor.indexed
                    if idx in indexed_colors:
                        bg_color = indexed_colors[idx]

                if bg_color:
                    style_def['bg'] = {'rgb': bg_color}

        # Alignment
        if cell.alignment:
            alignment: Alignment = cell.alignment
            # Horizontal: 1(left), 2(center), 3(right)
            if alignment.horizontal == 'center': style_def['ht'] = 2
            elif alignment.horizontal == 'right': style_def['ht'] = 3
            elif alignment.horizontal == 'left': style_def['ht'] = 1
            elif alignment.horizontal == 'general': style_def['ht'] = 1
            # Vertical: 0(top), 1(middle), 2(bottom)
            if alignment.vertical == 'top': style_def['vt'] = 0
            elif alignment.vertical == 'center': style_def['vt'] = 1
            elif alignment.vertical == 'bottom': style_def['vt'] = 2
            # Wrap text: 1(overflow), 2(wrap), 3(clip)
            # Always use wrap (2) instead of overflow to prevent text clipping
            if alignment.wrap_text:
                style_def['tb'] = 2  # Wrap text
            # Don't set tb if not wrapping - let Univer handle default

        # Borders
        if cell.border:
            border: Border = cell.border
            bd = {}

            # Helper function to convert border side
            def convert_border_side(side):
                if side and side.style:
                    border_style = {
                        'thin': {'s': 1, 'cl': {'rgb': '#000000'}},
                        'medium': {'s': 2, 'cl': {'rgb': '#000000'}},
                        'thick': {'s': 3, 'cl': {'rgb': '#000000'}},
                        'double': {'s': 6, 'cl': {'rgb': '#000000'}},
                        'hair': {'s': 1, 'cl': {'rgb': '#000000'}},
                        'dotted': {'s': 4, 'cl': {'rgb': '#000000'}},
                        'dashed': {'s': 5, 'cl': {'rgb': '#000000'}},
                    }
                    result = border_style.get(side.style, {'s': 1, 'cl': {'rgb': '#000000'}})
                    # Override color if specified
                    if side.color and side.color.value:
                        hex_color = str(side.color.value)
                        if len(hex_color) == 8:
                            result['cl'] = {'rgb': f"#{hex_color[2:].upper()}"}
                        elif len(hex_color) == 6:
                            result['cl'] = {'rgb': f"#{hex_color.upper()}"}
                    return result
                return None

            if border.top: bd['t'] = convert_border_side(border.top)
            if border.bottom: bd['b'] = convert_border_side(border.bottom)
            if border.left: bd['l'] = convert_border_side(border.left)
            if border.right: bd['r'] = convert_border_side(border.right)

            if bd:
                style_def['bd'] = bd

        # Number Format - Convert Excel patterns to Univer-compatible format
        if cell.number_format and cell.number_format != 'General':
            original_pattern = cell.number_format

            # Clean up Excel-specific codes first
            # _( and _) are spacing characters in Excel accounting format
            # * is repeat character
            pattern = original_pattern.replace('_(', '').replace('_)', '').replace('* ', '').replace('*', '')

            # Standardize quotes
            pattern = pattern.replace('"', '')

            # Determine format type and create proper Univer pattern
            if '%' in pattern:
                # PERCENTAGE FORMAT
                # Excel stores as decimal (0.1234), display as percentage (12.34%)
                # Pattern should multiply by 100 and add %

                # Extract decimal places
                if '0.00%' in pattern or '0.0%' in pattern:
                    # Has decimal places
                    decimal_count = pattern.count('0', pattern.index('.'), pattern.index('%')) if '.' in pattern else 0
                    base = '0.' + '0' * decimal_count + '%'
                else:
                    base = '0%'

                # Create pattern: positive;negative (2 sections only)
                pattern = f'{base};[Red]({base})'

            elif '#,##0' in pattern or '0.00' in pattern or '0.0' in pattern:
                # NUMBER FORMAT (with or without decimals)

                # Determine if has decimals
                if '0.00' in pattern or '#,##0.00' in pattern:
                    # 2 decimal places
                    base = '#,##0.00'
                elif '0.0' in pattern or '#,##0.0' in pattern:
                    # 1 decimal place
                    base = '#,##0.0'
                elif '0.000' in pattern or '#,##0.000' in pattern:
                    # 3+ decimal places
                    decimal_count = 0
                    if '.' in pattern:
                        # Count zeros after decimal
                        after_decimal = pattern.split('.')[1]
                        decimal_count = len([c for c in after_decimal if c == '0' or c == '#'])
                    base = '#,##0.' + '0' * decimal_count
                else:
                    # No decimals
                    base = '#,##0'

                # Create pattern: positive;negative (2 sections only)
                pattern = f'{base};[Red]({base})'

            elif pattern.startswith('0') and '.' in pattern:
                # Generic decimal format without comma
                decimal_count = len(pattern.split('.')[1].replace(';', '').strip())
                base = '0.' + '0' * decimal_count
                pattern = f'{base};[Red]({base})'

            # Log for debugging
            logger.info(f"Number format: '{original_pattern}' -> '{pattern}'")

            style_def['n'] = {'pattern': pattern}

        return style_def

    def _convert_sheet_to_json(self, worksheet: Worksheet) -> Dict[str, Any]:
        """
        Converts an openpyxl Worksheet object to a Univer sheet dictionary.
        """
        cell_data = {}
        row_data = {}
        col_data = {}

        # Process cells - include cells with value OR style
        for row in worksheet.iter_rows():
            for cell in row:
                # Skip completely empty cells (no value and no style)
                if cell.value is None and not cell.has_style:
                    continue

                univer_cell = {}

                # Value and Type
                if cell.value is not None:
                    if cell.data_type == 'n':
                        # Numeric value
                        univer_cell['t'] = 2  # Number type
                        univer_cell['v'] = cell.value

                        # Debug log for numbers with special formats
                        if cell.number_format and cell.number_format != 'General':
                            if '%' in cell.number_format or abs(cell.value) < 1:
                                logger.info(f"Number cell: value={cell.value}, format={cell.number_format}, coord={cell.coordinate}")

                    elif cell.data_type == 'f':
                        univer_cell['t'] = 2  # Formula result is usually number
                        univer_cell['f'] = f"={cell.value[1:]}" if cell.value.startswith('=') else f"={cell.value}"
                        # Store cached value if available
                        # Note: For formulas, you may need to load with data_only=False to get the formula
                        # and data_only=True to get the cached value separately
                    elif cell.data_type == 'b':
                        univer_cell['t'] = 4  # Boolean type
                        univer_cell['v'] = cell.value
                    else:  # 's' (string), 'e' (error), 'd' (date)
                        univer_cell['t'] = 1  # String type
                        univer_cell['v'] = str(cell.value)
                else:
                    # Cell has style but no value - still need to include it for styling
                    univer_cell['v'] = ''
                    univer_cell['t'] = 1  # String type (empty)

                # Style - always check for style
                style_def = self._convert_cell_style(cell)
                style_id = self._register_style(style_def)
                if style_id:
                    univer_cell['s'] = style_id

                # Only add cell if it has content or style
                if univer_cell:
                    row_key = str(cell.row - 1)
                    col_key = str(cell.column - 1)
                    if row_key not in cell_data:
                        cell_data[row_key] = {}
                    cell_data[row_key][col_key] = univer_cell

        # Process row heights
        # NOTE: We let Univer auto-size rows by NOT setting explicit heights
        # This makes rows adjust to content automatically
        for row_index, dim in worksheet.row_dimensions.items():
            row_dict = {}
            # Only set height if row is hidden (for hiding purposes)
            # Otherwise, let Univer auto-size the row
            if dim.hidden:
                row_dict["hd"] = 0
            # Optionally: If you want to preserve original Excel heights, uncomment:
            # if dim.height and not dim.hidden:
            #     height_in_px = dim.height * 1.333
            #     row_dict["h"] = height_in_px
            if row_dict:
                row_data[str(row_index - 1)] = row_dict

        # Process column widths - scan ALL columns used in worksheet
        max_col = worksheet.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_dict = {}

            # Check if this column has explicit dimension settings
            if col_letter in worksheet.column_dimensions:
                dim = worksheet.column_dimensions[col_letter]
                if dim.width:
                    # Convert Excel column width to pixels
                    # Excel width unit is character width, roughly: pixels = width * 7.5
                    width_in_px = dim.width * 7.5
                    col_dict["w"] = width_in_px
                # hd: 0 = hidden, 1 = visible (NOT hidden)
                if dim.hidden:
                    col_dict["hd"] = 0
            else:
                # No explicit width, use Excel's default column width
                default_width = worksheet.sheet_format.defaultColWidth or 8.43
                col_dict["w"] = default_width * 7.5

            if col_dict:
                col_data[str(col_idx - 1)] = col_dict

        # Process merged cells
        merges = []
        for merged_range in worksheet.merged_cells.ranges:
            merge_info = {
                "startRow": merged_range.min_row - 1,
                "endRow": merged_range.max_row - 1,
                "startColumn": merged_range.min_col - 1,
                "endColumn": merged_range.max_col - 1
            }
            merges.append(merge_info)

            # Ensure the top-left cell of merged range has style
            # Sometimes Excel only styles the first cell in a merge
            top_row = merged_range.min_row
            top_col = merged_range.min_col
            top_cell = worksheet.cell(top_row, top_col)

            # Make sure this cell is in our cell_data
            row_key = str(top_row - 1)
            col_key = str(top_col - 1)

            if row_key not in cell_data:
                cell_data[row_key] = {}

            if col_key not in cell_data[row_key]:
                # Cell wasn't processed yet, add it with style
                univer_cell = {}
                if top_cell.value is not None:
                    if top_cell.data_type == 'n':
                        univer_cell['t'] = 2
                        univer_cell['v'] = top_cell.value
                    else:
                        univer_cell['t'] = 1
                        univer_cell['v'] = str(top_cell.value)
                else:
                    univer_cell['v'] = ''
                    univer_cell['t'] = 1

                # Add style
                style_def = self._convert_cell_style(top_cell)
                style_id = self._register_style(style_def)
                if style_id:
                    univer_cell['s'] = style_id

                cell_data[row_key][col_key] = univer_cell

        # Build sheet data without freeze panes
        sheet_data = {
            "id": f"sheet_{worksheet.title.replace(' ', '_')}",
            "name": worksheet.title,
            "cellData": cell_data,
            "rowData": row_data,
            "columnData": col_data,
            "mergeData": merges,
            "rowCount": worksheet.max_row,
            "columnCount": worksheet.max_column,
            "defaultRowHeight": 25,  # Slightly larger default for better auto-fit
            "defaultColumnWidth": 64,  # Standard Excel default column width in pixels
            "zoomRatio": 1,  # Normal zoom
            "showGridlines": 1,  # Show gridlines
        }

        return sheet_data

    def convert_file_to_snapshot(self, file_path: Path) -> Dict[str, Any]:
        """
        Main conversion function. Loads an Excel file and converts it to a Univer snapshot.
        """
        logger.info(f"Starting full conversion of Excel file: {file_path}")

        try:
            # data_only=True reads cell values, not formulas. Change if formulas are needed.
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load workbook {file_path}: {e}")
            raise

        self.styles_registry = {}
        self.next_style_id = 0

        snapshot_sheets = {}
        snapshot_sheet_order = []

        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            sheet_json = self._convert_sheet_to_json(worksheet)
            sheet_id = sheet_json['id']
            snapshot_sheets[sheet_id] = sheet_json
            snapshot_sheet_order.append(sheet_id)

        snapshot = {
            "id": f"workbook-{file_path.stem}",
            "name": file_path.name,
            "sheetOrder": snapshot_sheet_order,
            "sheets": snapshot_sheets,
            "styles": self._build_styles_object(),
        }

        logger.info(f"Finished full conversion for {file_path}. Converted {len(snapshot_sheet_order)} sheets.")
        return snapshot

# Create a global instance of the converter
excel_to_univer_converter = ExcelToUniverConverter()
