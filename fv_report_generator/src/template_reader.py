"""
Template reader — parse Report_P14 sheet structure into row/column identity maps.

The template Excel already has all formatting, merged headers, and labels in place.
We extract two maps so we can write data into the right cells:

- row_map: {row_key: row_number}   where row_key = ("section", "sub1", "sub2")
                                   with None for higher-level aggregate rows
- col_map: {col_key: col_number}   where col_key is one of:
    ("GRAND_TOTAL",)
    ("BU_TOTAL", canonical_bu)
    ("SG_TOTAL", canonical_bu, canonical_sg)
    ("PRODUCT", canonical_bu, canonical_sg, canonical_product_key)

Also reports:
- derived_rows: set of row numbers whose values are computed (contribution margin, EBT, net profit,
  percentage, informational text) — these should not be overwritten from CSV pivot.
- data_rows: list of (row_number, row_key) for rows the pivot must fill.
"""
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from .normalizer import canonical, canonical_product_key


# Some template row labels use narrative text that doesn't canonicalize to the
# CSV GROUP label. Map each template canonical -> CSV canonical so both forms
# resolve to the same row_map key.
_ROW_LABEL_ALIASES = {
    # template: "5. กำไร(ขาดทุน)ก่อนต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น (3)-(4)"
    # CSV GROUP: "05.ต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น (3) (4)"
    canonical("5. กำไร(ขาดทุน)ก่อนต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น (3)-(4)"):
        canonical("05.ต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น (3) (4)"),
}


HEADER_BU_ROW = 5
HEADER_SG_ROW = 6
HEADER_SUBSG_ROW = 7
HEADER_PKEY_ROW = 8
HEADER_PNAME_ROW = 9
DATA_START_ROW = 10
LABEL_COL = 2
GRAND_TOTAL_COL = 3


@dataclass
class TemplateSchema:
    row_map: dict = field(default_factory=dict)
    col_map: dict = field(default_factory=dict)
    data_rows: list = field(default_factory=list)
    derived_rows: set = field(default_factory=set)
    informational_rows: set = field(default_factory=set)
    last_row: int = 0
    last_col: int = 0
    row_labels: dict = field(default_factory=dict)  # row_num -> original label
    col_labels: dict = field(default_factory=dict)  # col_num -> human-readable path


def _merged_value_resolver(ws):
    """Return a callable eff(row, col) that resolves merged-cell top-left values."""
    merged = {}
    for rng in ws.merged_cells.ranges:
        tl = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = tl

    def eff(r, c):
        if (r, c) in merged:
            return merged[(r, c)]
        return ws.cell(r, c).value

    return eff


def _classify_row(label: str, state: dict):
    """Classify a template row label into (row_key, kind). Mutates state for section tracking.

    row_key is a canonicalized tuple (section, sub1, sub2) where each element is
    canonical(...) or None for higher-level aggregates.
    """
    stripped = label.strip()

    # Section headers start with "N. " or "N.XXX"
    if stripped and stripped[0].isdigit() and ". " in stripped[:5]:
        state["section_raw"] = stripped
        sec = canonical(stripped)
        sec = _ROW_LABEL_ALIASES.get(sec, sec)
        state["section"] = sec
        state["sub1"] = None
        state["sub1_raw"] = None
        return (state["section"], None, None), "section"

    # Percentage marker (single line, e.g. "%กำไรส่วนเกิน (3)/(1)")
    if stripped.startswith("%"):
        return (canonical(stripped), None, None), "derived"

    # Dashed items ("    - label")
    if label.startswith(" ") and "-" in label[:10]:
        item = canonical(label)
        if state["sub1"] is not None:
            return (state["section"], state["sub1"], item), "sub2"
        return (state["section"], item, None), "sub1"

    # Non-dashed line inside an expense section → SUB_GROUP1
    section_raw = state.get("section_raw") or ""
    if any(tok in section_raw for tok in ("ค่าใช้จ่าย", "ผันแปร", "คงที่", "Variable", "Fixed")):
        state["sub1_raw"] = stripped
        state["sub1"] = canonical(stripped)
        return (state["section"], state["sub1"], None), "sub1"

    return (canonical(stripped), None, None), "informational"


def _classify_column(col: int, eff):
    """Classify a column by walking header rows 5 (BU), 6 (SG), 7 (sub-SG), 8 (PKEY or 'รวม')."""
    if col == LABEL_COL:
        return None

    r5 = eff(HEADER_BU_ROW, col)
    r6 = eff(HEADER_SG_ROW, col)
    r7 = eff(HEADER_SUBSG_ROW, col)
    r8 = eff(HEADER_PKEY_ROW, col)

    if col == GRAND_TOTAL_COL or (isinstance(r5, str) and r5.strip() == "รวมทั้งสิ้น"):
        return ("GRAND_TOTAL",)

    bu = canonical(r5)
    if not bu:
        return None

    # BU total column: r6 starts with "รวม" OR col is a single-column BU (r5 == r6).
    is_sum_col = isinstance(r6, str) and r6.strip().startswith("รวม")
    is_single_col_bu = r5 is not None and r6 is not None and str(r5) == str(r6)
    if is_sum_col or is_single_col_bu:
        return ("BU_TOTAL", bu)

    sg = canonical(r6)

    # SG-level grand total column (sums across sub-SGs): r7 == "รวม" (split case only) or r7 equals r6
    r7_str = str(r7).strip() if r7 is not None else ""
    r8_str = str(r8).strip() if r8 is not None else ""
    r7_is_sum = r7_str == "รวม"
    r7_equals_r6 = r7 is not None and r6 is not None and str(r7) == str(r6)

    if r8_str == "รวม":
        if r7_is_sum:
            # Top-level SG total spanning all sub-SGs (only exists when SG is split)
            return ("SG_TOTAL", bu, sg)
        # Non-split SG total (r7 == r6) or sub-SG total within a split SG
        subsg = sg if r7_equals_r6 else canonical(r7)
        if r7_equals_r6:
            return ("SG_TOTAL", bu, sg)
        return ("SUBSG_TOTAL", bu, sg, subsg)

    pkey = canonical_product_key(r8)
    if not pkey:
        return None
    subsg = sg if r7_equals_r6 else canonical(r7)
    return ("PRODUCT", bu, sg, subsg, pkey)


def read_template(template_path, sheet_name: str = "Report_P14") -> TemplateSchema:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb[sheet_name]
    eff = _merged_value_resolver(ws)

    schema = TemplateSchema(last_row=ws.max_row, last_col=ws.max_column)

    # --- Rows ---
    state = {"section": None, "section_raw": None, "sub1": None, "sub1_raw": None}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        label = ws.cell(r, LABEL_COL).value
        if label is None:
            continue
        row_key, kind = _classify_row(label, state)
        schema.row_map[row_key] = r
        schema.row_labels[r] = label
        if kind == "informational":
            schema.informational_rows.add(r)
        schema.data_rows.append((r, row_key, kind))

    # --- Columns ---
    for c in range(2, ws.max_column + 1):
        key = _classify_column(c, eff)
        if key is None:
            continue
        # first-wins (top-left of any merge)
        if key not in schema.col_map:
            schema.col_map[key] = c
        # keep a human-readable label for debug/reconcile
        r5 = eff(HEADER_BU_ROW, c)
        r6 = eff(HEADER_SG_ROW, c)
        r8 = eff(HEADER_PKEY_ROW, c)
        r9 = eff(HEADER_PNAME_ROW, c)
        schema.col_labels[c] = " | ".join(
            str(v).strip().replace("\n", " ") for v in (r5, r6, r8, r9) if v
        )

    return schema
