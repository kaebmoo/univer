"""Orchestrator — coordinate aggregator, builders, and writers to emit one .xlsx."""
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook

from . import aggregator, column_builder, row_builder
from .config import FVConfig
from .writers.cell_formatter import CellFormatter
from .writers.column_header_writer import ColumnHeaderWriter
from .writers.data_writer import DataWriter
from .writers.header_writer import HeaderWriter

log = logging.getLogger(__name__)


def generate_report(
    df: pd.DataFrame,
    output_path: Path,
    config: FVConfig,
    period_key: Optional[int] = None,
    sheet_name: str = "Report_FV",
) -> Path:
    """Build the FV report workbook from CSV data."""
    pivot = aggregator.build_pivot(df, period_key=period_key)
    log.info("pivot: %d (row, col) cells", len(pivot))

    columns = column_builder.build_columns(df, config, period_key=period_key)
    log.info("columns: %d", len(columns))

    rows = row_builder.build_rows(df, config, period_key=period_key)
    log.info("rows: %d", len(rows))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    formatter = CellFormatter(config)
    header_writer = HeaderWriter(config, formatter)
    column_header_writer = ColumnHeaderWriter(config, formatter)
    data_writer = DataWriter(config, formatter)

    # Column headers + widths first (so HeaderWriter knows last_col for merging)
    last_col = column_header_writer.write(ws, columns, start_xl_col=config.label_col)
    header_writer.write(ws, last_col=last_col)
    data_writer.write(ws, columns, rows, pivot, start_xl_col=config.label_col)

    # Freeze panes: just below header rows, just after grand_total column
    freeze_row = config.data_start_row
    grand_col = config.label_col + 2  # label + grand_total = 2 cols, freeze AFTER grand_total
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(grand_col)}{freeze_row}"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("saved %s (%.1f KB)", output_path, output_path.stat().st_size / 1024)
    return output_path, pivot
