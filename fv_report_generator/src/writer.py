"""
Writer — clone the Report_P14 template, clear the data region, and fill in
the freshly pivoted values. The template preserves all merged headers, fonts,
and column widths, so we only touch data cells.
"""
import shutil
from pathlib import Path

import openpyxl

from .template_reader import DATA_START_ROW


def _is_merge_origin(ws, row: int, col: int) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row and rng.min_col == col:
            return True
    return False


def _is_inside_merge(ws, row: int, col: int) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return not (rng.min_row == row and rng.min_col == col)
    return False


def write_report(
    template_path: Path,
    output_path: Path,
    cells: dict,
    schema,
    sheet_name: str = "Report_P14",
) -> Path:
    """Copy template → output, clear data area, write pivot cells.

    Args:
        template_path: path to Report_FV_Y2568(P14).XLSX (or any period)
        output_path:   target .xlsx (will be overwritten)
        cells:         {(row_num, col_num): value} from apply_pivot_to_schema
        schema:        TemplateSchema (used to bound the data region)
        sheet_name:    target sheet within the template

    Returns:
        output_path
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path, keep_vba=False)
    ws = wb[sheet_name]

    first_data_col = 3  # column C (grand total) is first data column
    last_row = max((r for r, _, _ in schema.data_rows), default=schema.last_row)
    last_col = schema.last_col

    # 1) Clear data area, skipping informational rows and cells covered by merges.
    for r in range(DATA_START_ROW, last_row + 1):
        if r in schema.informational_rows:
            continue
        for c in range(first_data_col, last_col + 1):
            if _is_inside_merge(ws, r, c):
                continue
            ws.cell(r, c).value = None

    # 2) Write pivot values.
    for (r, c), value in cells.items():
        if _is_inside_merge(ws, r, c):
            continue
        ws.cell(r, c).value = float(value)

    wb.save(output_path)
    return output_path
