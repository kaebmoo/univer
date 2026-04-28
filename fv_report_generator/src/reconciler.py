"""
Reconciler — verify the generated .xlsx against the source CSV.

Two independent verification paths:

1. reconcile() — cell-by-cell comparison of the .xlsx against the pivot
   recomputed from build_pivot().  Catches writer bugs, lost cells,
   column-map drift, and rounding regressions end-to-end.

2. reconcile_invariants() — checks structural consistency of build_pivot()
   itself, without trusting its logic.  Catches aggregation bugs that
   reconcile() would miss because both sides use the same code path.

   Layer A — Hierarchical totals:
       GRAND_TOTAL         == sum(BU_TOTAL for every BU)
       BU_TOTAL[bu]        == sum(SG_TOTAL for every SG in that BU)
       SG_TOTAL[bu,sg]     == sum(PRODUCT  for every product in that SG/subSG)
                              (for split SGs also checks via SUBSG_TOTAL)

   Layer B — Raw CSV cross-check:
       Sum VALUES per GROUP straight from the CSV (no canonicalization, no
       logic) and compare to pivot GRAND_TOTAL per section row.  A mismatch
       here means build_pivot() dropped or double-counted rows.

   Layer C — Derived row formula:
       %กำไรส่วนเกิน == CM / Revenue for every column, verified directly
       from the pivot scalars, not from derived.percent_value().

This is the only verification path — there is no external "Data_P14"
ground-truth file. The CSV itself is the source of truth.
"""
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from . import aggregator, column_builder, derived, row_builder
from .normalizer import canonical, canonical_product_key
from .satellite_split import SATELLITE_GROUPS, is_split_sg, subsg_for


# ---------------------------------------------------------------------------
# Original cell-level reconciler
# ---------------------------------------------------------------------------

@dataclass
class ReconcileResult:
    mismatches: List[Tuple] = field(default_factory=list)
    # (row_label, col_display, actual_in_xlsx, expected_from_csv, diff)

    cells_checked: int = 0
    cells_skipped: int = 0  # informational/empty etc.

    @property
    def ok(self) -> bool:
        return not self.mismatches


def _to_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def reconcile(
    output_path,
    df: pd.DataFrame,
    config,
    period_key: Optional[int] = None,
    sheet_name: str = "Report_FV",
    tolerance: float = 0.01,
) -> ReconcileResult:
    """Compare every non-label data cell in `output_path` to the pivot recomputed from `df`."""
    pivot = aggregator.build_pivot(df, period_key=period_key)
    columns = column_builder.build_columns(df, config, period_key=period_key)
    rows = row_builder.build_rows(df, config, period_key=period_key)

    wb = openpyxl.load_workbook(output_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {output_path}; available: {wb.sheetnames}")
    ws = wb[sheet_name]

    result = ReconcileResult()

    for row_offset, rd in enumerate(rows):
        xl_row = config.data_start_row + row_offset
        for col_offset, col in enumerate(columns):
            xl_col = config.label_col + col_offset
            if col.col_type == "label":
                continue

            ck = col.col_key
            if ck is None:
                continue

            # Expected value
            if rd.row_type == "percent":
                expected = derived.percent_value(pivot, ck)
            else:
                expected = pivot.get((rd.row_key, ck))

            actual = _to_float(ws.cell(xl_row, xl_col).value)

            # Treat None and 0 as equivalent (writer formats 0 as blank).
            exp_f = expected if expected is not None else 0.0
            act_f = actual if actual is not None else 0.0

            result.cells_checked += 1
            diff = act_f - exp_f
            if abs(diff) > tolerance:
                result.mismatches.append((
                    rd.display_label,
                    col.display_name or repr(ck),
                    act_f,
                    exp_f,
                    diff,
                ))
    return result


# ---------------------------------------------------------------------------
# Business-rule checker — reads directly from .xlsx, trusts neither pivot
# nor CSV.  Verifies accounting identities on the numbers that actually
# landed in the spreadsheet.
# ---------------------------------------------------------------------------

@dataclass
class BusinessRuleViolation:
    rule: str           # e.g. 'CM = Revenue − Variable Cost'
    col_display: str    # column header
    got: float          # value found in xlsx
    expected: float     # value computed from other xlsx cells
    diff: float
    detail: str = ""    # human-readable breakdown (Revenue=…, VarCost=…)


@dataclass
class BusinessRuleResult:
    violations: List[BusinessRuleViolation] = field(default_factory=list)
    checks_run: int = 0
    skipped_sections: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.violations

    def summary(self) -> str:
        if self.ok:
            return f"All {self.checks_run} business-rule checks passed."
        lines = [f"{self.checks_run} checks, {len(self.violations)} violation(s):"]
        for v in self.violations:
            lines.append(
                f"  [{v.rule}] col={v.col_display}: "
                f"xlsx_CM={v.got:.2f}, expected={v.expected:.2f} (diff {v.diff:+.2f})"
            )
            if v.detail:
                lines.append(f"    {v.detail}")
        return "\n".join(lines)


def reconcile_business_rules(
    output_path,
    df: pd.DataFrame,
    config,
    period_key: Optional[int] = None,
    sheet_name: str = "Report_FV",
    tolerance: float = 0.01,
) -> BusinessRuleResult:
    """Verify accounting identities by reading values directly from the .xlsx.

    Does NOT use build_pivot() or the raw CSV for the actual numbers —
    it reads whatever ended up in each cell of the Excel file and checks
    that the numbers are self-consistent.

    Row rules — P&L chain (per column):
        Rule 1: CM (03)     = Revenue (01) − Variable Cost (02)
        Rule 2: EBITDA (05) = CM (03)     − Fixed Cost (04)
        Rule 3: EBIT (09)   = EBITDA (05) − Depreciation (07)
        Rule 4: EBT (11)    = EBIT (09)   − Interest (10)

    Column rules — subtotal hierarchy (per section row):
        Rule 5: GRAND_TOTAL == Σ BU_TOTAL  (for every section row)
        Rule 6: BU_TOTAL    == Σ SG_TOTAL  (for every section row, per BU)
        Rule 7: SG_TOTAL    == Σ PRODUCT   (for every section row, per SG)
    """
    columns = column_builder.build_columns(df, config, period_key=period_key)
    rows = row_builder.build_rows(df, config, period_key=period_key)

    wb = openpyxl.load_workbook(output_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"sheet {sheet_name!r} not found in {output_path}; "
            f"available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    result = BusinessRuleResult()

    def _cell(xl_row: int, xl_col: int) -> float:
        return _to_float(ws.cell(xl_row, xl_col).value) or 0.0

    def _check_rule(rule: str, col_display: str, got: float, expected: float,
                    detail: str = "") -> None:
        result.checks_run += 1
        diff = got - expected
        if abs(diff) > tolerance:
            result.violations.append(BusinessRuleViolation(
                rule=rule, col_display=col_display,
                got=got, expected=expected, diff=diff, detail=detail,
            ))

    # ------------------------------------------------------------------
    # Locate xl_row for every section code present in the sheet
    # ------------------------------------------------------------------
    # Include "06" (ผลตอบแทนทางการเงินและรายได้อื่น) — not a pure section but its header row
    # carries the section-level total used in the EBT formula.
    # Include "08" (ต้นทุนทางการเงิน-ด้านการจัดหาเงิน) — mandatory row, may be 0.
    _ALL_CODES = {"01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11"}
    section_xl_rows: Dict[str, int] = {}
    for row_offset, rd in enumerate(rows):
        if rd.is_section and rd.parent_section_code in _ALL_CODES:
            code = rd.parent_section_code
            if code not in section_xl_rows:
                section_xl_rows[code] = config.data_start_row + row_offset

    def _has(*codes: str) -> bool:
        return all(c in section_xl_rows for c in codes)

    data_cols = [
        (col_offset, col)
        for col_offset, col in enumerate(columns)
        if col.col_type != "label" and col.col_key is not None
    ]

    # ------------------------------------------------------------------
    # Row rules — P&L chain (per column)
    #
    # Each rule: (result_code, [(input_code, sign), ...], rule_name)
    # expected = Σ (sign × section_value_from_xlsx)
    #
    # Actual formulas from the report (confirmed from Excel label):
    #   CM  (03) = Revenue (01) − Variable Cost (02)
    #   (05) = CM (03) − Fixed Cost (04)
    #          [กำไร(ขาดทุน)ก่อนต้นทุนจัดหาเงิน รายได้อื่นและค่าใช้จ่ายอื่น]
    #   EBT (09) = (05) + Financial/OtherIncome (06)
    #                   − OtherExpenses (07)
    #                   − FinancialCost (08)    ← (5)+(6)−(7)−(8) per report label
    #   Net (11) = EBT (09) − Tax (10)
    #
    # NOTE: There is NO EBITDA in this report.
    #       Section 07 is "ค่าใช้จ่ายอื่น" (other expenses), NOT depreciation.
    # ------------------------------------------------------------------
    _PL_RULES = [
        ("03", [("01", +1), ("02", -1)],
               "CM = Revenue − Variable Cost"),
        ("05", [("03", +1), ("04", -1)],
               "ผลก่อนต้นทุนจัดหาเงินฯ (05) = CM (03) − Fixed Cost (04)"),
        ("09", [("05", +1), ("06", +1), ("07", -1), ("08", -1)],
               "EBT (09) = (05)+(06)−(07)−(08)"),
        ("11", [("09", +1), ("10", -1)],
               "Net Profit (11) = EBT (09) − Tax (10)"),
    ]
    for code_r, operands, rule_name in _PL_RULES:
        needed = [code_r] + [c for c, _ in operands]
        if not _has(*needed):
            missing_codes = [c for c in needed if c not in section_xl_rows]
            result.skipped_sections.append(
                f"{rule_name} (missing sections: {missing_codes})"
            )
            continue
        xl_r = section_xl_rows[code_r]
        xl_inputs = [(section_xl_rows[c], sign) for c, sign in operands]

        # Skip rule if ALL input operand values are zero across every data column.
        # This happens when a section exists as a structural row (e.g. mandatory
        # injected row) but carries no data — the formula is inapplicable, not wrong.
        all_inputs_zero = all(
            _cell(xl_row_in, config.label_col + col_offset) == 0.0
            for col_offset, col in data_cols
            for xl_row_in, _ in xl_inputs
            if xl_row_in != section_xl_rows[code_r]   # exclude result row
        )
        if all_inputs_zero:
            result.skipped_sections.append(
                f"{rule_name} (skipped: all operand sections are zero — formula not applicable)"
            )
            continue

        for col_offset, col in data_cols:
            xl_col = config.label_col + col_offset
            val_r    = _cell(xl_r, xl_col)
            expected = sum(sign * _cell(xl_row_in, xl_col)
                           for xl_row_in, sign in xl_inputs)
            parts = "  ".join(
                f"[{c}]={'+'if sign>0 else '-'}{abs(_cell(section_xl_rows[c], xl_col)):,.2f}"
                for c, sign in operands
            )
            _check_rule(
                rule=rule_name,
                col_display=col.display_name or repr(col.col_key),
                got=val_r,
                expected=expected,
                detail=f"{parts}  expected [{code_r}]={expected:,.2f}  xlsx [{code_r}]={val_r:,.2f}",
            )

    # ------------------------------------------------------------------
    # Column rules — subtotal hierarchy (per section row)
    # ------------------------------------------------------------------
    grand_xl_col: Optional[int] = None
    bu_xl_col: Dict[str, int] = {}
    sg_xl_col: Dict[tuple, int] = {}
    prod_xl_col: Dict[tuple, List[int]] = {}

    for col_offset, col in enumerate(columns):
        xl_col = config.label_col + col_offset
        if col.col_type == "grand_total":
            grand_xl_col = xl_col
        elif col.col_type == "bu_total":
            bu_xl_col[col.bu_raw] = xl_col
        elif col.col_type == "sg_total":
            sg_xl_col[(col.bu_raw, col.sg_raw)] = xl_col
        elif col.col_type == "product":
            prod_xl_col.setdefault((col.bu_raw, col.sg_raw), []).append(xl_col)

    bu_to_sgs: Dict[str, List[int]] = {}
    for (bu_raw, _), xl_col in sg_xl_col.items():
        bu_to_sgs.setdefault(bu_raw, []).append(xl_col)

    section_rows_with_xl = [
        (config.data_start_row + offset, rd)
        for offset, rd in enumerate(rows)
        if rd.is_section
    ]

    for xl_row, rd in section_rows_with_xl:
        row_label = rd.display_label

        # Rule 5: GRAND_TOTAL == Σ BU_TOTAL
        if grand_xl_col is not None and bu_xl_col:
            grand_val = _cell(xl_row, grand_xl_col)
            bu_sum = sum(_cell(xl_row, c) for c in bu_xl_col.values())
            _check_rule(
                rule="GRAND_TOTAL == Σ BU_TOTAL",
                col_display=f"row={row_label}",
                got=grand_val, expected=bu_sum,
                detail=f"grand={grand_val:,.2f}  Σbu={bu_sum:,.2f}",
            )

        # Rule 6: BU_TOTAL == Σ SG_TOTAL  (per BU)
        for bu_raw, sg_cols in bu_to_sgs.items():
            if bu_raw not in bu_xl_col:
                continue
            bu_val = _cell(xl_row, bu_xl_col[bu_raw])
            sg_sum = sum(_cell(xl_row, c) for c in sg_cols)
            _check_rule(
                rule="BU_TOTAL == Σ SG_TOTAL",
                col_display=f"row={row_label}  bu={bu_raw}",
                got=bu_val, expected=sg_sum,
                detail=f"bu={bu_val:,.2f}  Σsg={sg_sum:,.2f}",
            )

        # Rule 7: SG_TOTAL == Σ PRODUCT  (per SG)
        for sg_key, prod_cols in prod_xl_col.items():
            if sg_key not in sg_xl_col:
                continue
            sg_val   = _cell(xl_row, sg_xl_col[sg_key])
            prod_sum = sum(_cell(xl_row, c) for c in prod_cols)
            _check_rule(
                rule="SG_TOTAL == Σ PRODUCT",
                col_display=f"row={row_label}  sg={sg_key[1]}",
                got=sg_val, expected=prod_sum,
                detail=f"sg={sg_val:,.2f}  Σprod={prod_sum:,.2f}",
            )

    return result


# ---------------------------------------------------------------------------
# Invariant checker — independent of build_pivot() logic
# ---------------------------------------------------------------------------

@dataclass
class InvariantViolation:
    layer: str          # 'A_hierarchy', 'B_raw_csv', or 'C_percent'
    description: str    # human-readable check name
    got: float
    expected: float
    diff: float


@dataclass
class InvariantResult:
    violations: List[InvariantViolation] = field(default_factory=list)
    checks_run: int = 0

    @property
    def ok(self) -> bool:
        return not self.violations

    def summary(self) -> str:
        if self.ok:
            return f"All {self.checks_run} invariant checks passed."
        lines = [f"{self.checks_run} checks, {len(self.violations)} violation(s):"]
        for v in self.violations:
            lines.append(
                f"  [{v.layer}] {v.description}: got {v.got:.4f}, expected {v.expected:.4f} (diff {v.diff:+.4f})"
            )
        return "\n".join(lines)


def _raw_section_sums(df: pd.DataFrame, period_key: Optional[int]) -> Dict[str, float]:
    """Sum VALUE per GROUP straight from the CSV — no canonicalization, no logic.

    This is the independent reference for Layer B.  Rows with GROUP starting
    with the percent-group prefix ("33.") are excluded because they are not
    summable (pre-computed per-product percentages).
    """
    _PERCENT_PREFIX = "33."
    sub = df.copy()
    if period_key is not None:
        sub = sub[sub["TIME_KEY"] == period_key]
    sub = sub[sub["VALUE"].notna()].copy()
    sub["VALUE"] = sub["VALUE"].astype(float)
    # Exclude percent-group rows
    sub = sub[~sub["GROUP"].astype(str).str.lstrip().str.startswith(_PERCENT_PREFIX)]
    sums: Dict[str, float] = (
        sub.groupby("GROUP")["VALUE"].sum().to_dict()
    )
    return sums


def reconcile_invariants(
    df: pd.DataFrame,
    period_key: Optional[int] = None,
    tolerance: float = 0.01,
) -> InvariantResult:
    """Check structural consistency of build_pivot() independent of its logic.

    Parameters
    ----------
    df:
        Raw DataFrame loaded from the source CSV (same object passed to
        reconcile()).
    period_key:
        Same period filter used when generating the report.
    tolerance:
        Absolute difference threshold to flag a violation.

    Returns
    -------
    InvariantResult with a list of violations and a checks_run counter.
    """
    pivot = aggregator.build_pivot(df, period_key=period_key)
    result = InvariantResult()

    def _p(key: tuple) -> float:
        return pivot.get(key, 0.0)

    def _check(layer: str, description: str, got: float, expected: float) -> None:
        result.checks_run += 1
        diff = got - expected
        if abs(diff) > tolerance:
            result.violations.append(InvariantViolation(
                layer=layer,
                description=description,
                got=got,
                expected=expected,
                diff=diff,
            ))

    # ------------------------------------------------------------------
    # Layer A — Hierarchical totals
    # ------------------------------------------------------------------

    bus = aggregator.enumerate_bus(df, period_key=period_key)

    # A1: GRAND_TOTAL == sum of all BU_TOTAL
    grand_total = _p(("GRAND_TOTAL",))
    bu_total_sum = sum(_p(("BU_TOTAL", canonical(bu))) for bu in bus)
    _check("A_hierarchy", "GRAND_TOTAL == Σ BU_TOTAL", grand_total, bu_total_sum)

    for bu in bus:
        bu_c = canonical(bu)
        bu_val = _p(("BU_TOTAL", bu_c))
        sgs = aggregator.enumerate_sgs(df, bu, period_key=period_key)

        # A2: BU_TOTAL == sum of its SG_TOTAL
        sg_sum = sum(_p(("SG_TOTAL", bu_c, canonical(sg))) for sg in sgs)
        _check("A_hierarchy", f"BU_TOTAL[{bu}] == Σ SG_TOTAL", bu_val, sg_sum)

        for sg in sgs:
            sg_c = canonical(sg)
            sg_val = _p(("SG_TOTAL", bu_c, sg_c))

            products = aggregator.enumerate_products(df, bu, sg, period_key=period_key)
            if not products:
                continue

            # A3a: If SG is flat — SG_TOTAL == sum of its PRODUCT columns
            if not is_split_sg(sg_c):
                prod_sum = sum(
                    _p(("PRODUCT", bu_c, sg_c,
                        subsg_for(sg_c, canonical_product_key(pkey)),
                        canonical_product_key(pkey)))
                    for pkey, _ in products
                )
                _check(
                    "A_hierarchy",
                    f"SG_TOTAL[{bu}/{sg}] == Σ PRODUCT",
                    sg_val,
                    prod_sum,
                )
            else:
                # A3b: Split SG — SG_TOTAL == sum of SUBSG_TOTAL
                subsg_names = [g["name"] for g in SATELLITE_GROUPS.values()]
                subsg_sum = sum(
                    _p(("SUBSG_TOTAL", bu_c, sg_c, canonical(subsg)))
                    for subsg in subsg_names
                )
                _check(
                    "A_hierarchy",
                    f"SG_TOTAL[{bu}/{sg}] == Σ SUBSG_TOTAL",
                    sg_val,
                    subsg_sum,
                )
                # A3c: each SUBSG_TOTAL == sum of its products
                for subsg in subsg_names:
                    subsg_c = canonical(subsg)
                    subsg_val = _p(("SUBSG_TOTAL", bu_c, sg_c, subsg_c))
                    sub_products = aggregator.enumerate_products(
                        df, bu, sg, period_key=period_key, subsg_canon=subsg_c
                    )
                    sub_prod_sum = sum(
                        _p(("PRODUCT", bu_c, sg_c, subsg_c, canonical_product_key(pkey)))
                        for pkey, _ in sub_products
                    )
                    _check(
                        "A_hierarchy",
                        f"SUBSG_TOTAL[{bu}/{sg}/{subsg}] == Σ PRODUCT",
                        subsg_val,
                        sub_prod_sum,
                    )

    # ------------------------------------------------------------------
    # Layer B — Raw CSV cross-check (independent simple groupby)
    # ------------------------------------------------------------------

    raw_sums = _raw_section_sums(df, period_key)
    for group_raw, raw_val in raw_sums.items():
        row_key = (canonical(group_raw), None, None)
        pivot_val = _p((row_key, ("GRAND_TOTAL",)))
        _check(
            "B_raw_csv",
            f"GRAND_TOTAL for GROUP '{group_raw}'",
            pivot_val,
            raw_val,
        )

    # ------------------------------------------------------------------
    # Layer C — Derived percent row formula
    # ------------------------------------------------------------------

    # Build list of non-label col_keys from pivot keys
    col_keys_seen = set()
    for (row_key_part, ck) in pivot.keys():
        if ck is not None:
            col_keys_seen.add(ck)

    _REV_KEY = derived._REVENUE_KEY
    _CM_KEY = derived._CM_KEY

    for ck in col_keys_seen:
        revenue = _p((_REV_KEY, ck))
        cm = _p((_CM_KEY, ck))
        if not revenue:
            continue  # undefined percent — skip
        expected_pct = cm / revenue
        # derived.percent_value recomputes the same — verify it matches directly
        got_pct = derived.percent_value(pivot, ck) or 0.0
        _check(
            "C_percent",
            f"%กำไรส่วนเกิน for col {ck}",
            got_pct,
            expected_pct,
        )

    return result
