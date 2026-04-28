"""Excel cell formatter — fonts, fills, borders, number formats."""
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_THIN_SIDE = Side(style="thin", color="808080")
THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


class CellFormatter:
    def __init__(self, config):
        self.config = config

    def font(self, bold: bool = False, size: Optional[int] = None, color: str = "000000") -> Font:
        return Font(name=self.config.font_name, size=size or self.config.font_size, bold=bold, color=color)

    def fill(self, color: Optional[str]) -> Optional[PatternFill]:
        if not color or color == "FFFFFF":
            return None
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def format_title(self, cell, text: str, bold: bool = True, size: Optional[int] = None,
                     horizontal: str = "center"):
        cell.value = text
        cell.font = self.font(bold=bold, size=size or self.config.title_font_size)
        cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    def format_header(self, cell, text: str, color: str, bold: bool = True):
        cell.value = text
        cell.font = self.font(bold=bold, size=self.config.header_font_size)
        f = self.fill(color)
        if f:
            cell.fill = f
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    def format_label(self, cell, text: str, bold: bool = False, color: Optional[str] = None):
        cell.value = text
        cell.font = self.font(bold=bold)
        f = self.fill(color)
        if f:
            cell.fill = f
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER

    def format_number(self, cell, value, bold: bool = False, color: Optional[str] = None):
        cell.font = self.font(bold=bold)
        f = self.fill(color)
        if f:
            cell.fill = f
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if value is None:
            cell.value = None
        else:
            cell.value = float(value)
        # Accounting-style: negatives in red parentheses, zero blank
        cell.number_format = '#,##0.00;[Red](#,##0.00);""'

    def format_percent(self, cell, value, bold: bool = False, color: Optional[str] = None):
        cell.font = self.font(bold=bold)
        f = self.fill(color)
        if f:
            cell.fill = f
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if value is None:
            cell.value = None
        else:
            cell.value = float(value)
        cell.number_format = '0.00%;[Red](0.00%);""'

    def set_column_width(self, ws, col_index: int, width: float):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    def set_row_height(self, ws, row_index: int, height: float):
        ws.row_dimensions[row_index].height = height
