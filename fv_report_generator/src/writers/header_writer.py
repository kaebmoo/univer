"""Title rows above the column headers."""
from openpyxl.utils import get_column_letter


class HeaderWriter:
    def __init__(self, config, formatter):
        self.config = config
        self.formatter = formatter

    def write(self, ws, last_col: int):
        merge_end = get_column_letter(max(last_col, self.config.label_col + 1))

        # Row 1
        c = ws.cell(self.config.company_row, self.config.label_col)
        self.formatter.format_title(c, "บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)", size=self.config.title_font_size, horizontal="left")
        ws.merge_cells(start_row=self.config.company_row, start_column=self.config.label_col,
                       end_row=self.config.company_row, end_column=last_col)

        # Row 2
        c = ws.cell(self.config.report_title_row, self.config.label_col)
        self.formatter.format_title(c, "รายงานต้นทุนคงที่/ผันแปร", size=self.config.title_font_size, horizontal="left")
        ws.merge_cells(start_row=self.config.report_title_row, start_column=self.config.label_col,
                       end_row=self.config.report_title_row, end_column=last_col)

        # Row 3 — period label
        c = ws.cell(self.config.period_row, self.config.label_col)
        period_text = self.config.period_label or ""
        self.formatter.format_title(c, period_text, bold=False, size=self.config.font_size, horizontal="left")
        ws.merge_cells(start_row=self.config.period_row, start_column=self.config.label_col,
                       end_row=self.config.period_row, end_column=last_col)
