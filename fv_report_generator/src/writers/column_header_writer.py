"""
Column header writer — emit a 5-row header matching Report_P14:

    row N+0 : BU                  (merged across BU's columns)
    row N+1 : SG                   (merged across SG's columns)
    row N+2 : SUBSG                (merged across SUBSG's columns; same as SG for non-split)
    row N+3 : product key
    row N+4 : product name

Where N = config.header_start_row.

Total / Summary columns occupy a span:
- label column        : merged across all 5 rows
- grand_total column  : merged across all 5 rows
- bu_total column     : merged across rows 1-5 (under BU header — but bu_total is itself the BU label position; we render as merged "รวม BU" across all 5 rows)
- sg_total column     : merged across rows 2-5 ("รวม SG")
- subsg_total column  : merged across rows 3-5 ("รวม SUBSG")
- product column      : BU/SG/SUBSG cells merged from neighbours; product_key + product_name written

Width of every column is also applied here.
"""
from typing import List

from openpyxl.utils import get_column_letter

from ..column_builder import ColumnDef


class ColumnHeaderWriter:
    def __init__(self, config, formatter):
        self.config = config
        self.formatter = formatter

    def write(self, ws, columns: List[ColumnDef], start_xl_col: int) -> int:
        """Write headers; return the last column number written (1-indexed)."""
        cf = self.formatter
        h0 = self.config.header_start_row          # BU row
        h_sg = h0 + 1
        h_subsg = h0 + 2
        h_pkey = h0 + 3
        h_pname = h0 + 4
        last_xl_col = start_xl_col - 1

        # Per-column write pass
        for idx, col in enumerate(columns):
            xl_col = start_xl_col + idx
            last_xl_col = xl_col
            cf.set_column_width(ws, xl_col, col.width)

            if col.col_type == "label":
                ws.cell(h0, xl_col).value = col.display_name
                cf.format_header(ws.cell(h0, xl_col), col.display_name, col.color)
                ws.merge_cells(start_row=h0, start_column=xl_col, end_row=h_pname, end_column=xl_col)

            elif col.col_type == "grand_total":
                cf.format_header(ws.cell(h0, xl_col), col.display_name, col.color)
                ws.merge_cells(start_row=h0, start_column=xl_col, end_row=h_pname, end_column=xl_col)

            elif col.col_type == "bu_total":
                cf.format_header(ws.cell(h0, xl_col), col.display_name, col.color)
                ws.merge_cells(start_row=h0, start_column=xl_col, end_row=h_pname, end_column=xl_col)

            elif col.col_type == "sg_total":
                cf.format_header(ws.cell(h_sg, xl_col), col.display_name, col.color)
                ws.merge_cells(start_row=h_sg, start_column=xl_col, end_row=h_pname, end_column=xl_col)

            elif col.col_type == "subsg_total":
                cf.format_header(ws.cell(h_subsg, xl_col), col.display_name, col.color)
                ws.merge_cells(start_row=h_subsg, start_column=xl_col, end_row=h_pname, end_column=xl_col)

            elif col.col_type == "product":
                # product_key on h_pkey, product_name on h_pname; BU/SG/SUBSG merged later
                cf.format_header(ws.cell(h_pkey, xl_col), col.product_key_raw or "", col.color)
                cf.format_header(ws.cell(h_pname, xl_col), col.product_name or "", col.color)
                # also fill the BU/SG/SUBSG cells with same color so merged spans display correctly
                cf.format_header(ws.cell(h0, xl_col), "", col.color)
                cf.format_header(ws.cell(h_sg, xl_col), "", col.color)
                cf.format_header(ws.cell(h_subsg, xl_col), "", col.color)

        # Pass 2: merge BU header span over consecutive same-BU columns
        self._merge_bu_spans(ws, columns, start_xl_col, h0)
        # Pass 3: merge SG header span over consecutive same-(BU,SG) product columns (non-sg_total cells)
        self._merge_sg_spans(ws, columns, start_xl_col, h_sg)
        # Pass 4: merge SUBSG header span over consecutive same-(BU,SG,SUBSG) product columns
        self._merge_subsg_spans(ws, columns, start_xl_col, h_subsg)

        return last_xl_col

    # --- merge helpers --------------------------------------------------------

    def _merge_bu_spans(self, ws, columns: List[ColumnDef], start_xl_col: int, row: int):
        """Merge row=BU across each BU's product/sg/subsg columns (excluding bu_total)."""
        i = 0
        n = len(columns)
        while i < n:
            col = columns[i]
            if col.col_type == "product" or col.col_type == "sg_total" or col.col_type == "subsg_total":
                # Find run of same BU among non-bu_total columns
                bu = col.bu_raw
                j = i
                while j < n and columns[j].col_type in ("product", "sg_total", "subsg_total") and columns[j].bu_raw == bu:
                    j += 1
                # Merge h0 from i..j-1, set BU display
                first_xl = start_xl_col + i
                last_xl = start_xl_col + j - 1
                if last_xl > first_xl:
                    self._set_and_merge(ws, row, first_xl, last_xl, _bu_display(bu), columns[i].color)
                else:
                    self.formatter.format_header(ws.cell(row, first_xl), _bu_display(bu), columns[i].color)
                i = j
            else:
                i += 1

    def _merge_sg_spans(self, ws, columns: List[ColumnDef], start_xl_col: int, row: int):
        """Merge row=SG across consecutive same-(BU,SG) columns of types product/subsg_total (skip sg_total cell itself)."""
        i = 0
        n = len(columns)
        while i < n:
            col = columns[i]
            if col.col_type in ("product", "subsg_total"):
                bu, sg = col.bu_raw, col.sg_raw
                j = i
                while j < n and columns[j].col_type in ("product", "subsg_total") and columns[j].bu_raw == bu and columns[j].sg_raw == sg:
                    j += 1
                first_xl = start_xl_col + i
                last_xl = start_xl_col + j - 1
                if last_xl > first_xl:
                    self._set_and_merge(ws, row, first_xl, last_xl, _strip_leading_zero(sg), columns[i].color)
                else:
                    self.formatter.format_header(ws.cell(row, first_xl), _strip_leading_zero(sg), columns[i].color)
                i = j
            else:
                i += 1

    def _merge_subsg_spans(self, ws, columns: List[ColumnDef], start_xl_col: int, row: int):
        """Merge row=SUBSG across consecutive same-(BU,SG,SUBSG) product columns only."""
        i = 0
        n = len(columns)
        while i < n:
            col = columns[i]
            if col.col_type == "product":
                bu, sg, subsg = col.bu_raw, col.sg_raw, col.subsg_raw
                j = i
                while j < n and columns[j].col_type == "product" and columns[j].bu_raw == bu and columns[j].sg_raw == sg and columns[j].subsg_raw == subsg:
                    j += 1
                first_xl = start_xl_col + i
                last_xl = start_xl_col + j - 1
                # Display SUBSG name (or empty if same as SG)
                disp = "" if subsg == sg else _strip_leading_zero(subsg)
                if last_xl > first_xl:
                    self._set_and_merge(ws, row, first_xl, last_xl, disp, columns[i].color)
                else:
                    self.formatter.format_header(ws.cell(row, first_xl), disp, columns[i].color)
                i = j
            else:
                i += 1

    def _set_and_merge(self, ws, row: int, first_col: int, last_col: int, text: str, color: str):
        self.formatter.format_header(ws.cell(row, first_col), text, color)
        ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)


def _strip_leading_zero(text: str) -> str:
    if not text:
        return text
    if len(text) >= 3 and text[0] == "0" and text[1].isdigit() and text[2] == ".":
        return text[1:]
    return text


def _bu_display(text: str) -> str:
    return _strip_leading_zero(text or "")
